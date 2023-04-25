import csv
import os

import openpyxl as xl
import re
import datetime
import xlsxwriter
import pandas as pd




def start_handler(path):
    directory_dbs_files = os.getcwd() + "\\dbs"
    file_db_isotp = '\\iso_tp_db.csv'
    isotp_dic = {}
    tp_dic = {}
    iso_dic = {}

    """
    Сводные списки для финальной записи в сводки по фазам.
    """
    summary_iso_tp_phase_1 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                               'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
                               'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
                               'Статус уведомлений']]

    summary_iso_tp_phase_2 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                               'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
                               'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
                               'Статус уведомлений']]

    summary_iso_tp_phase_3 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                               'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
                               'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
                               'Статус уведомлений']]

    summary_iso_tp_phase_4 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                               'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
                               'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
                               'Статус уведомлений']]

    summary_iso_tp_phase_5 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                               'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
                               'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
                               'Статус уведомлений']]

    summary_tp_phase_1 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                               'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]

    summary_tp_phase_2 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                               'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]

    summary_tp_phase_3 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                               'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]

    summary_tp_phase_4 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                           'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]

    summary_tp_phase_5 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                           'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]


    """
    Наполнение словарей по изометрии+тестпакет и по тестпакетам исходной подтверждённой информацией.
    """
    with open(directory_dbs_files + file_db_isotp, 'r') as read_db:
        readed_db = csv.reader(read_db, delimiter=';')

        for row in readed_db:
            isotp_dic[row[0]] = ['', '', '', '', '', '', '', 0, '', '', '', '', '', '', '', '', '', '']
            tp_dic[row[2]] = ['', '', '', '', '', 0, '', '', '', '', '', '']
            iso_dic[row[1]] = ['', '', '', '', '', '']

    with open(directory_dbs_files + file_db_isotp, 'r') as read_db:
        readed_db = csv.reader(read_db, delimiter=';')

        for row in readed_db:

            iso_with_tp = row[0].strip()
            isometric = row[1].strip()
            testpackage = row[2].strip()
            phase = row[3].strip()
            line = row[4].strip()
            title = row[5].strip()
            unit = row[6].strip()
            fluid = row[7].strip()
            ggn_status = row[8].strip()
            iso_length = float(row[9].strip().replace(',', '.'))

            rfi_erection, rfi_test, rfi_airblowing, rfi_reinstatement = '', '', '', ''


            if row[10]:
                rfi_erection = row[10].strip() + ' BD'
            if row[11]:
                rfi_test = row[11].strip() + ' BD'
            if row[12]:
                rfi_airblowing = row[12].strip() + ' BD'
            if row[13]:
                rfi_reinstatement = row[13].strip() + ' BD'


            type_ins = row[14].strip()
            volume_ins = row[15].strip()

            rfi_ins_cotton, rfi_ins_metall, rfi_ins_box = '', '', ''

            if row[16]:
                rfi_ins_cotton = row[16].strip() + ' BD'
            if row[17]:
                rfi_ins_metall = row[17].strip() + ' BD'
            if row[18]:
                rfi_ins_box = row[18].strip() + ' BD'

            isotp_dic[iso_with_tp] = [testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length,
                                      rfi_erection, rfi_test, rfi_airblowing,
                                      rfi_reinstatement, type_ins, volume_ins,
                                      rfi_ins_cotton, rfi_ins_metall, rfi_ins_box
                , '', '', phase]

            tp_dic[testpackage][0] = testpackage
            tp_dic[testpackage][1] = title
            tp_dic[testpackage][2] = unit
            tp_dic[testpackage][3] = fluid
            tp_dic[testpackage][4] = ggn_status
            tp_dic[testpackage][5] += iso_length
            tp_dic[testpackage][6] = rfi_erection
            tp_dic[testpackage][7] = rfi_test
            tp_dic[testpackage][8] = rfi_airblowing
            tp_dic[testpackage][9] = rfi_reinstatement
            tp_dic[testpackage][11] = phase

            iso_dic[isometric][5] = testpackage



    # Проверка Журнал заявок АИС Р2 ФАЗА 1, 2, 3, 4, 5-------------------------------------
    df = pd.read_excel('Журнал заявок общий.xlsx', engine='openpyxl')
    df = df.sort_values(by='Дата назначения инспекции / Date of scheduled inspection', ascending=True)
    df.to_excel('Журнал заявок общий.xlsx', index=0)

    print('Журнал заявок отсортирован по дате отработки инспекции.')

    wb_journal_rfi = xl.load_workbook('Журнал заявок общий.xlsx')
    sheet_journal_rfi = wb_journal_rfi['Sheet1']

    replace_pattern_1 = ['-HT', '-VT', '-PT']
    replace_pattern_2 = ['(T.T. REINSTATEMENT)', '(T.T. AIR BLOWING)', '(AIR BLOWING)', '(T.T AIR BLOWING)', '(T.T AIR BLOWING',
                         '(T.T. ERECTION)', '(T.T .ERECTION)', '(T.T.TEST)', '(T.T. AIR BLOWIHG)', '(T.T. TEST)',
                         '(T.T ERECTION)', '(T.T TEST)', '(T.T REINSTATEMENT)', '(T.T. REIINSTATEMENT)',
                          '(GPA AIR BLOWING)', '(GPA TEST)',
                         '(GPA ERECTION)', '(GPA REINSTATEMENT)', '(T.T. REISTATEMENT)', '(T.T.REINSTATEMENT)',
                         '(T.T RE-INSTATEMENT)', '( T.T AIR BLOWING )', '( T.T AIR BLOWING )', '(TT REINSTATEMENT)',
                         '(T.T.ERECTION)', '(T.T.TEST)', '(T.T.AIR BLOWING)', '(T.T.REINSTATEMENT)']


    for i in sheet_journal_rfi['B2':'AT550000']:
        if i[0].value:
            rfi_number = str(i[1].value).strip()
            tp_number = str(i[2].value).strip().replace('-HT', '').replace('-VT', '').replace('-PT', '')
            pkk = str(i[4].value).strip()

            description_rfi = str(i[18].value)
            violation = str(i[39].value)
            list_iso = str(i[8].value).replace(' ', '').split(';')
            category_cancelled = str(i[34].value).strip()
            comment = str(i[44].value)

            date_of_submissions = str(i[13].value)

            re_tp = re.findall(
                r'YMT-\d-\d\d\d-HP-\d\d\d\d\d\d-\d\d[A-Z]-\d\d[A-Z]|YMT-\d-\d\d\d-HP-\d\d\d\d\d\d-\d\d-\d\d[A-Z]|'
                r'YMT-\d-\d\d\d-HP-\d\d\d\d\d\d-\d\d[A-Z]-\d\d|'
                r'YMT-\d-\d\d-HP-\d\d\d\d\d\d-\d\d[A-Z]-\d\d[A-Z]|YMT-\d-\d\d-HP-\d\d\d\d\d\d-\d\d[A-Z]-\d\d|'
                r'YMT-\d-\d\d-HP-\d\d\d\d\d\d-\d\d-\d\d[A-Z]|YMT-\d-\d\d-HP-\d\d\d\d\d\d-\d\d-\d\d\d|'
                r'YMT-\d-\d\d-HP-\d\d\d\d\d\d-\d\d-\d\d',
                tp_number)
            if re_tp:
                tp_number = re_tp[0]


            for typo in replace_pattern_2:
                if typo in tp_number:
                    tp_number = tp_number.replace(typo, '').strip()

            for typo in replace_pattern_1:
                if typo in tp_number:
                    tp_number = tp_number.replace(typo, '').strip()

            if 'INTER' in str(i[2].value):
                tp_number = tp_number.replace('1-00', '').replace('2-00', '').replace('3-00', '').replace('4-00', '')
                for temp in replace_pattern_2:
                    tp_number = tp_number.replace(temp, '')
                tp_number = 'INTERFACE ISO'

            tp_number = tp_number.strip()

            rep_patt_for_iso = ['р.01', 'р.1', 'р.3', 'р.4', 'р.5', 'р.6', 'р.7', 'р.8', 'р.0', '\n']

            if '97011' in rfi_number:
                description_rfi = description_rfi + 'крепежа и герметизации металлического кожуха согласно изометрическим'

            if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                if 'подтвержд' in comment or 'подтвржд' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][6] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][8] = rfi_number + " ФОП"
                if 'зафиксирова' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][6] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][8] = rfi_number + " ФОП"
                if 'Принято' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][6] = rfi_number
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][8] = rfi_number
                if 'Принято с замечаниями' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][6] = rfi_number + " ПЗ"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][8] = rfi_number + " ПЗ"




            if 'испытаний на прочность и плотность' in description_rfi or\
                    'испытаний технологического трубопровода  на прочность' in description_rfi or\
                    'испытаний технологического трубопровода на прочность' in description_rfi:

                if 'Принято' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number
                if 'Принято с замечаниями' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number + " ПЗ"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number + " ПЗ"
                if 'подтвержд' in comment or 'подтвржд' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number + " ФОП"
                if 'зафиксирова' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number + " ФОП"
                if 'выдерж' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number + " ФОП"

            if 'испыт' and 'рочност' in description_rfi:
                if 'подтвержд' in comment or 'подтвржд' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number + " ФОП"
                if 'зафиксирова' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number + " ФОП"
                if 'выдерж' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number + " ФОП"
                if 'Принято' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number
                if 'Принято с замечаниями' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][7] = rfi_number + " ПЗ"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number + " ПЗ"




            if 'родувка' in description_rfi and 'еплоспутн' not in description_rfi:
                if 'подтвержд' in comment or 'подтвржд' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][8] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][10] = rfi_number + " ФОП"
                if 'зафиксирова' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][8] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][10] = rfi_number + " ФОП"
                if 'Принято' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][8] = rfi_number
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][10] = rfi_number
                if 'Принято с замечаниями' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][8] = rfi_number + " ПЗ"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][10] = rfi_number + " ПЗ"




            if 'сборки технологических трубопроводов в проект' in description_rfi or \
                    'сборки технологических трубопроводов в рамках' in description_rfi or \
                    'обратной cборки в рамках тест' in description_rfi:
                if 'подтвержд' in comment or 'подтвржд' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][9] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][11] = rfi_number + " ФОП"
                if 'зафиксирова' in comment:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][9] = rfi_number + " ФОП"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][11] = rfi_number + " ФОП"
                if 'Принято' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][9] = rfi_number
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][11] = rfi_number
                if 'Принято с замечаниями' == category_cancelled:
                    if tp_number in tp_dic.keys():
                        tp_dic[tp_number][9] = rfi_number + " ПЗ"
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                                    isotp_dic[f'{isometric.strip()}-{tp_number}'][11] = rfi_number + " ПЗ"



            if 'дополн' in description_rfi:
                if 'Принято' == category_cancelled:
                    if list_iso:
                        for isometric in list_iso:
                            for patt in rep_patt_for_iso:
                                isometric = isometric.replace(patt, '')
                                if isometric.strip() in iso_dic.keys():
                                    iso_dic[isometric.strip()][3] = rfi_number


            # ИЗОЛЯЦИЯ ПРОВЕРКА------------------------------ ТРУБА
            if 'завершении работ по теплоизоляц' in pkk:
                if 'Принято' == category_cancelled:
                    if 'крепежа и герметизации металлического кожуха согласно изометрическим' in description_rfi \
                            or ' крепежа и герметизации металлического кожуха согласно изометрическим' in description_rfi:
                        if list_iso:
                            for iso in list_iso:
                                for patt in rep_patt_for_iso:
                                    iso = iso.replace(patt, '')
                                if iso.strip() in iso_dic.keys():
                                    iso_dic[iso.strip()][1] = rfi_number
                    if 'теплоизоляционного покрытия согласно изометрическим' in description_rfi:
                        if list_iso:
                            for iso in list_iso:
                                for patt in rep_patt_for_iso:
                                    iso = iso.replace(patt, '')
                                if iso.strip() in iso_dic.keys():
                                    iso_dic[iso.strip()][0] = rfi_number
                if "Принято с замечаниями" == category_cancelled:
                    if 'крепежа и герметизации металлического кожуха согласно изометрическим' in description_rfi \
                            or ' крепежа и герметизации металлического кожуха согласно изометрическим' in description_rfi:
                        if list_iso:
                            for iso in list_iso:
                                for patt in rep_patt_for_iso:
                                    iso = iso.replace(patt, '')
                                if iso.strip() in iso_dic.keys():
                                    iso_dic[iso.strip()][1] = rfi_number + " ПЗ"
                    if 'теплоизоляционного покрытия согласно изометрическим' in description_rfi:
                        if list_iso:
                            for iso in list_iso:
                                for patt in rep_patt_for_iso:
                                    iso = iso.replace(patt, '')
                                if iso.strip() in iso_dic.keys():
                                    iso_dic[iso.strip()][0] = rfi_number + " ПЗ"


                if 'представлены не в полном объеме, представлены некорректные документы' in violation or \
                        'документы, подтверждающие качество работ' in violation\
                        or 'подтвержд' in comment or 'подтвржд' in comment:

                    if 'крепежа и герметизации металлического кожуха согласно изометрическим' in description_rfi \
                            or ' крепежа и герметизации металлического кожуха согласно изометрическим' in description_rfi:
                        if list_iso:
                            for iso in list_iso:
                                for patt in rep_patt_for_iso:
                                    iso = iso.replace(patt, '')
                                    if iso.strip() in iso_dic.keys():
                                        iso_dic[iso.strip()][1] = rfi_number + ' ФОП'

                    if 'теплоизоляционного покрытия согласно изометрическим' in description_rfi:
                        if list_iso:
                            for iso in list_iso:
                                for patt in rep_patt_for_iso:
                                    iso = iso.replace(patt, '')
                                    if iso.strip() in iso_dic.keys():
                                        iso_dic[iso.strip()][0] = rfi_number + ' ФОП'




                # КОРОБА/ЧЕХЛЫ ФЛАНЦЫ ЗРА

                    if 'смонтированного теплоизоляционного покрытия и металлического кожуха металлических коробов на ' \
                       'фланцах и ЗРА' in description_rfi or 'Проверка качества' \
                                                             ' смонтированной теплоизоляционной оболочки ( термочехлов) ' \
                                                             'согласно изометрическим' in description_rfi or 'Проверка ' \
                             'качества смонтированного теплоизоляционного покрытия и металлического ' \
                                 'кожуха металлических коробов, теплоизоляционной оболочки ' \
                                         '( термочехлов) на фланцах' in description_rfi\
                            or 'Проверка качества крепежа и герметизации металлического ' \
                               'кожуха фланцев и ЗРА согласно изометрическим чертежам' in description_rfi\
                            or 'Проверка качества крепежа и герметизации металлического' \
                               ' кожуха фланцев и ЗРА согласно изометрическим чертежам' in description_rfi\
                            or 'Проверка качества смонтированного теплоизоляционного ' \
                               'покрытия металлических коробов на фланцах и ЗРА'in description_rfi\
                            or 'Проверка качества смонтированной теплоизоляционной ' \
                               'оболочки ( термочехлов) и металлических коробов' \
                               ' согласно изометрическим чертежам' in description_rfi\
                            or 'качества смонтированного теплоизоляционного' \
                               ' покрытия металлических коробов на фланцах и ЗРА' in description_rfi\
                            or 'смонтированного теплоизоляционного покрытия и' \
                               ' металлического кожуха фланцев и ЗРА согласно изометрич' in description_rfi\
                            or 'смонтированной теплоизоляционной оболочки ( термочехлов)' \
                               ' и металлических коробов' in description_rfi\
                            or 'Проверка качества смонтированного теплоизоляционного покрытия и ' \
                               'металлического кожуха металлических коробов на фланцах и ЗРА' in description_rfi\
                            or ' качества монтажа металлических коробов коллекторов' in description_rfi:

                        if list_iso:
                            for iso in list_iso:
                                for patt in rep_patt_for_iso:
                                    iso = iso.replace(patt, '')
                                    if iso.strip() in iso_dic.keys():
                                        iso_dic[iso.strip()][2] = rfi_number + ' ФОП'



    print(f'Дата подачи последней инспекции в выгрузке - {date_of_submissions}')

    # -Проверка на уведомления-------------------------
    wb_ncr = xl.load_workbook('Реестр уведомлений.xlsx')
    sheet_ncr = wb_ncr['Предписания (Instructions)']
    iso_ncr = {}

    for i in sheet_ncr['B4':'V55000']:
        if i[0].value:
            number_ncr = str(i[0].value)
            mark_execution = str(i[16].value)
            notification_items = str(i[1].value)
            type_violation = str(i[5].value)
            content_remarks = str(i[6].value).replace(' ', '')
            content_remarks_iso = re.findall(
                r'\d-\d-\d-\d\d-\d\d\d-\w*-\d\w-\d\d\d\d-\d\d\d|'
                r'\d-\d-\d-\d\d-\d\d\d-\w*-\d\d-\d\d\d\d-\d\d\d|'
                r'\d-\d-\d-\d\d-\d\d\d-NHC3P\+-\d\d-\d\d\d\d-\d\d\d|'
                r'\d-\d-\d-\d\d-\d\d\d-NHC3\+-\d\d-\d\d\d\d-\d\d\d|'
                r'\d-\d-\d-\d\d-\d\d\d-NHC4P\+-\d\d-\d\d\d\d-\d\d\d|'
                r'\d-\d-\d-\d\d-\d\d\d-NHC5\+-\d\d-\d\d\d\d-\d\d\d|'
                r'\d-\d-\d-\d\d-\d\d\d-NHC4\+-\d\d-\d\d\d\d-\d\d\d',
                content_remarks.replace(' ', '').replace('\n', '').replace('Р', 'P').replace('С', 'C').strip())
            if 'Нет' in mark_execution:
                if content_remarks_iso:
                    for iso in content_remarks_iso:
                        try:
                            if number_ncr not in iso_dic[iso][4]:
                                iso_dic[iso][4] += number_ncr + "/"
                            if number_ncr not in tp_dic[iso_dic[iso][5]][10]:
                                tp_dic[iso_dic[iso][5]][10] += number_ncr + "/"
                        except:
                            print(f'Не нашел в словаре изометрию {iso}')
        else:
            break




    """
    Блок подготовки к записи в сводный файл
    """

    n_dic_4_30 = {'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHNGA': ['Природный газ',  0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHNGAD': ['Сухой природный газ',  0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWSU': ['Поверхностная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NOWWA': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

                  # 'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'UNMP': ['Азот СД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

    n_dic_4_110 = {
                    # 'HWBR': ['Вода котлового контура, обратная (Т21)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'HWBS': ['Вода котлового контура, прямая (Т11)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NHSGA': ['Товарный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NHSGAHP': ['Товарный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NODRA': ['Дренаж', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UABA': ['Барьерный воздух ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UHGAL': ['Топливный газ НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'UWWW': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                   'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

    n_dic_2_60 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC4P+': ['Бутановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC3P+': ['Пропановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC5+': ['С5+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC4+': ['С4+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNMP': ['Азот СД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NOWWA': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'WMMI': ['Водометанольная смесь', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWSU': ['Поверхностная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HYDV': ['Пары углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  }

    n_dic_2_70 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

    n_dic_1_60 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC4P+': ['Бутановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC3P+': ['Пропановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC5+': ['С5+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC4+': ['С4+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNMP': ['Азот СД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NOWWA': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'WMMI': ['Водометанольная смесь', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  # 'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWSU': ['Поверхностная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HYDV': ['Пары углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  }

    n_dic_1_70 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                  'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

    # isotp_dic = {}
    # tp_dic = {}
    # iso_dic = {}
    #
    # """
    # Сводные списки для финальной записи в сводки по фазам.
    # """
    # summary_iso_tp_phase_1 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
    #                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
    #                            'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
    #                            'Статус уведомлений']]
    #
    # summary_iso_tp_phase_2 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
    #                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
    #                            'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
    #                            'Статус уведомлений']]
    #
    # summary_iso_tp_phase_3 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
    #                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
    #                            'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
    #                            'Статус уведомлений']]
    #
    # summary_tp_phase_1 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
    #                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]
    #
    # summary_tp_phase_2 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
    #                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]
    #
    # summary_tp_phase_3 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
    #                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]


    n_list_4_110 = [['', f'Статус по ТП 4-110 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                     '', '', '', '', '', ''],
                    ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                     'Принят монтаж, ТП',
                     'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                     'Принята ОС, м.',
                     'Принята ОС, ТП']]
    ITOG_list_4_110 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    n_list_4_30 = [['', f'Статус по ТП 4-30 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                    '', '', '', '', '',
                    ''],
                   ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                    'Принят монтаж, ТП',
                    'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                    'Принята ОС, м.',
                    'Принята ОС, ТП']]
    ITOG_list_4_30 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    n_list_2_60 = [['', f'Статус по ТП 2-60 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                    '', '', '', '', '', ''],
                   ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                    'Принят монтаж, ТП',
                    'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                    'Принята ОС, м.',
                    'Принята ОС, ТП']]
    ITOG_list_2_60 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    n_list_2_70 = [['', f'Статус по ТП 2-70 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                    '', '', '', '', '', ''],
                   ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                    'Принят монтаж, ТП',
                    'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                    'Принята ОС, м.',
                    'Принята ОС, ТП']]
    ITOG_list_2_70 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    n_list_1_60 = [['', f'Статус по ТП 1-60 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                    '', '', '', '', '',
                    ''],
                   ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                    'Принят монтаж, ТП',
                    'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                    'Принята ОС, м.',
                    'Принята ОС, ТП']]
    ITOG_list_1_60 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    n_list_1_70 = [['', f'Статус по ТП 1-70 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                    '', '', '', '', '',
                    ''],
                   ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                    'Принят монтаж, ТП',
                    'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                    'Принята ОС, м.',
                    'Принята ОС, ТП']]
    ITOG_list_1_70 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    for isotp in isotp_dic.keys():

        isotp_dic[isotp][14] = iso_dic[isotp_dic[isotp][1]][0]
        isotp_dic[isotp][15] = iso_dic[isotp_dic[isotp][1]][1]
        isotp_dic[isotp][16] = iso_dic[isotp_dic[isotp][1]][2]
        isotp_dic[isotp][17] = iso_dic[isotp_dic[isotp][1]][3]
        isotp_dic[isotp][18] = iso_dic[isotp_dic[isotp][1]][4]


        if isotp_dic[isotp][19] == '1' and 'SUPER' not in isotp_dic[isotp][0]:
            summary_iso_tp_phase_1.append([isotp_dic[isotp][0], isotp_dic[isotp][1], isotp_dic[isotp][2],
                                           isotp_dic[isotp][3], isotp_dic[isotp][4], isotp_dic[isotp][5],
                                           isotp_dic[isotp][6], isotp_dic[isotp][7], isotp_dic[isotp][8],
                                           isotp_dic[isotp][9], isotp_dic[isotp][10], isotp_dic[isotp][11],
                                           isotp_dic[isotp][12], isotp_dic[isotp][13], isotp_dic[isotp][14],
                                           isotp_dic[isotp][15], isotp_dic[isotp][16], isotp_dic[isotp][17],
                                           isotp_dic[isotp][18]
                                           ])

        if isotp_dic[isotp][19] == '2' and 'SUPER' not in isotp_dic[isotp][0]:
            summary_iso_tp_phase_2.append([isotp_dic[isotp][0], isotp_dic[isotp][1], isotp_dic[isotp][2],
                                           isotp_dic[isotp][3], isotp_dic[isotp][4], isotp_dic[isotp][5],
                                           isotp_dic[isotp][6], isotp_dic[isotp][7], isotp_dic[isotp][8],
                                           isotp_dic[isotp][9], isotp_dic[isotp][10], isotp_dic[isotp][11],
                                           isotp_dic[isotp][12], isotp_dic[isotp][13], isotp_dic[isotp][14],
                                           isotp_dic[isotp][15], isotp_dic[isotp][16], isotp_dic[isotp][17],
                                           isotp_dic[isotp][18]
                                           ])

        if isotp_dic[isotp][19] == '3' and 'SUPER' not in isotp_dic[isotp][0]:
            summary_iso_tp_phase_3.append([isotp_dic[isotp][0], isotp_dic[isotp][1], isotp_dic[isotp][2],
                                           isotp_dic[isotp][3], isotp_dic[isotp][4], isotp_dic[isotp][5],
                                           isotp_dic[isotp][6], isotp_dic[isotp][7], isotp_dic[isotp][8],
                                           isotp_dic[isotp][9], isotp_dic[isotp][10], isotp_dic[isotp][11],
                                           isotp_dic[isotp][12], isotp_dic[isotp][13], isotp_dic[isotp][14],
                                           isotp_dic[isotp][15], isotp_dic[isotp][16], isotp_dic[isotp][17],
                                           isotp_dic[isotp][18]
                                           ])

        if isotp_dic[isotp][19] == '4' and 'SUPER' not in isotp_dic[isotp][0]:
            summary_iso_tp_phase_4.append([isotp_dic[isotp][0], isotp_dic[isotp][1], isotp_dic[isotp][2],
                                           isotp_dic[isotp][3], isotp_dic[isotp][4], isotp_dic[isotp][5],
                                           isotp_dic[isotp][6], isotp_dic[isotp][7], isotp_dic[isotp][8],
                                           isotp_dic[isotp][9], isotp_dic[isotp][10], isotp_dic[isotp][11],
                                           isotp_dic[isotp][12], isotp_dic[isotp][13], isotp_dic[isotp][14],
                                           isotp_dic[isotp][15], isotp_dic[isotp][16], isotp_dic[isotp][17],
                                           isotp_dic[isotp][18]
                                           ])

        if isotp_dic[isotp][19] == '5' and 'SUPER' not in isotp_dic[isotp][0]:
            summary_iso_tp_phase_5.append([isotp_dic[isotp][0], isotp_dic[isotp][1], isotp_dic[isotp][2],
                                           isotp_dic[isotp][3], isotp_dic[isotp][4], isotp_dic[isotp][5],
                                           isotp_dic[isotp][6], isotp_dic[isotp][7], isotp_dic[isotp][8],
                                           isotp_dic[isotp][9], isotp_dic[isotp][10], isotp_dic[isotp][11],
                                           isotp_dic[isotp][12], isotp_dic[isotp][13], isotp_dic[isotp][14],
                                           isotp_dic[isotp][15], isotp_dic[isotp][16], isotp_dic[isotp][17],
                                           isotp_dic[isotp][18]
                                           ])


    # isotp_dic[iso_with_tp] = [
    # testpackage, 0
    # isometric,1
    # line, 2
    # title, 3
    # unit, 4
    # fluid, 5
    # ggn_status,6
    # iso_length,7
    # rfi_erection, 8
    # rfi_test, 9
    # rfi_airblowing, 10
    # rfi_reinstatement, 11
    # type_ins, 12
    # volume_ins 13
    # rfi_ins_cotton, 14
    # rfi_ins_metall, 15
    # rfi_ins_box, 16
    # rfi dig 17
    # ncr 18
    # phase 19
    #
    #         tp_dic[testpackage][0] = testpackage
    #         tp_dic[testpackage][1] = title
    #         tp_dic[testpackage][2] = unit
    #         tp_dic[testpackage][3] = fluid
    #         tp_dic[testpackage][4] = ggn_status
    #         tp_dic[testpackage][5] += iso_length
    #         tp_dic[testpackage][6] = rfi_erection
    #         tp_dic[testpackage][7] = rfi_test
    #         tp_dic[testpackage][8] = rfi_airblowing
    #         tp_dic[testpackage][9] = rfi_reinstatement
    #         tp_dic[testpackage][10] = ncr_status
    #         tp_dic[testpackage][11] = phase

    summary_tp_for_db_atom = []
    for tp in tp_dic.keys():

        if 'SUPER' not in tp_dic[tp][0] and 'INTERFACE' not in tp_dic[tp][0] and "n/d" not in tp_dic[tp][0]:
            summary_tp_for_db_atom.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                       tp_dic[tp][5], tp_dic[tp][6], tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                       tp_dic[tp][10]])

        if tp_dic[tp][11] == '1' and 'SUPER' not in tp_dic[tp][0] and 'INTERFACE'\
                not in tp_dic[tp][0] and "n/d" not in tp_dic[tp][0]:
            summary_tp_phase_1.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                       tp_dic[tp][5], tp_dic[tp][6],tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                       tp_dic[tp][10]])

            if tp_dic[tp][2] == '1-60':
                n_dic_1_60[tp_dic[tp][3]][1] += tp_dic[tp][5]
                n_dic_1_60[tp_dic[tp][3]][2] += 1

                if tp_dic[tp][6]:
                    n_dic_1_60[tp_dic[tp][3]][3] += tp_dic[tp][5]
                    n_dic_1_60[tp_dic[tp][3]][4] += 1
                if tp_dic[tp][7] and tp_dic[tp][7] != 'Визуальный BD':
                    n_dic_1_60[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_1_60[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][7] == 'Визуальный BD' and tp_dic[tp][8]:
                    n_dic_1_60[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_1_60[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][8]:
                    n_dic_1_60[tp_dic[tp][3]][7] += tp_dic[tp][5]
                    n_dic_1_60[tp_dic[tp][3]][8] += 1
                if tp_dic[tp][9]:
                    n_dic_1_60[tp_dic[tp][3]][9] += tp_dic[tp][5]
                    n_dic_1_60[tp_dic[tp][3]][10] += 1


            if tp_dic[tp][2] == '1-70':
                n_dic_1_70[tp_dic[tp][3]][1] += tp_dic[tp][5]
                n_dic_1_70[tp_dic[tp][3]][2] += 1

                if tp_dic[tp][6]:
                    n_dic_1_70[tp_dic[tp][3]][3] += tp_dic[tp][5]
                    n_dic_1_70[tp_dic[tp][3]][4] += 1
                if tp_dic[tp][7] and tp_dic[tp][7] != 'Визуальный BD':
                    n_dic_1_70[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_1_70[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][7] == 'Визуальный BD' and tp_dic[tp][8]:
                    n_dic_1_70[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_1_70[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][8]:
                    n_dic_1_70[tp_dic[tp][3]][7] += tp_dic[tp][5]
                    n_dic_1_70[tp_dic[tp][3]][8] += 1
                if tp_dic[tp][9]:
                    n_dic_1_70[tp_dic[tp][3]][9] += tp_dic[tp][5]
                    n_dic_1_70[tp_dic[tp][3]][10] += 1

        if tp_dic[tp][11] == '2' and 'SUPER' not in tp_dic[tp][0] and 'INTERFACE'\
                not in tp_dic[tp][0] and "n/d" not in tp_dic[tp][0]:
            summary_tp_phase_2.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                       tp_dic[tp][5], tp_dic[tp][6],tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                       tp_dic[tp][10]])

            if tp_dic[tp][2] == '2-60':
                n_dic_2_60[tp_dic[tp][3]][1] += tp_dic[tp][5]
                n_dic_2_60[tp_dic[tp][3]][2] += 1

                if tp_dic[tp][6]:
                    n_dic_2_60[tp_dic[tp][3]][3] += tp_dic[tp][5]
                    n_dic_2_60[tp_dic[tp][3]][4] += 1
                if tp_dic[tp][7] and tp_dic[tp][7] != 'Визуальный BD':
                    n_dic_2_60[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_2_60[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][7] == 'Визуальный BD' and tp_dic[tp][8]:
                    n_dic_2_60[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_2_60[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][8]:
                    n_dic_2_60[tp_dic[tp][3]][7] += tp_dic[tp][5]
                    n_dic_2_60[tp_dic[tp][3]][8] += 1
                if tp_dic[tp][9]:
                    n_dic_2_60[tp_dic[tp][3]][9] += tp_dic[tp][5]
                    n_dic_2_60[tp_dic[tp][3]][10] += 1

            if tp_dic[tp][2] == '2-70':
                n_dic_2_70[tp_dic[tp][3]][1] += tp_dic[tp][5]
                n_dic_2_70[tp_dic[tp][3]][2] += 1

                if tp_dic[tp][6]:
                    n_dic_2_70[tp_dic[tp][3]][3] += tp_dic[tp][5]
                    n_dic_2_70[tp_dic[tp][3]][4] += 1
                if tp_dic[tp][7] and tp_dic[tp][7] != 'Визуальный BD':
                    n_dic_2_70[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_2_70[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][7] == 'Визуальный BD' and tp_dic[tp][8]:
                    n_dic_2_70[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_2_70[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][8]:
                    n_dic_2_70[tp_dic[tp][3]][7] += tp_dic[tp][5]
                    n_dic_2_70[tp_dic[tp][3]][8] += 1
                if tp_dic[tp][9]:
                    n_dic_2_70[tp_dic[tp][3]][9] += tp_dic[tp][5]
                    n_dic_2_70[tp_dic[tp][3]][10] += 1

        if tp_dic[tp][11] == '3' and 'SUPER' not in tp_dic[tp][0] and 'INTERFACE'\
                not in tp_dic[tp][0] and "n/d" not in tp_dic[tp][0]:
            summary_tp_phase_3.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                       tp_dic[tp][5], tp_dic[tp][6],tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                       tp_dic[tp][10]])
            if tp_dic[tp][2] == '4-30':
                n_dic_4_30[tp_dic[tp][3]][1] += tp_dic[tp][5]
                n_dic_4_30[tp_dic[tp][3]][2] += 1

                if tp_dic[tp][6]:
                    n_dic_4_30[tp_dic[tp][3]][3] += tp_dic[tp][5]
                    n_dic_4_30[tp_dic[tp][3]][4] += 1
                if tp_dic[tp][7] and tp_dic[tp][7] != 'Визуальный BD':
                    n_dic_4_30[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_4_30[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][7] == 'Визуальный BD' and tp_dic[tp][8]:
                    n_dic_4_30[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_4_30[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][8]:
                    n_dic_4_30[tp_dic[tp][3]][7] += tp_dic[tp][5]
                    n_dic_4_30[tp_dic[tp][3]][8] += 1
                if tp_dic[tp][9]:
                    n_dic_4_30[tp_dic[tp][3]][9] += tp_dic[tp][5]
                    n_dic_4_30[tp_dic[tp][3]][10] += 1

            if tp_dic[tp][2] == '4-110':
                n_dic_4_110[tp_dic[tp][3]][1] += tp_dic[tp][5]
                n_dic_4_110[tp_dic[tp][3]][2] += 1

                if tp_dic[tp][6]:
                    n_dic_4_110[tp_dic[tp][3]][3] += tp_dic[tp][5]
                    n_dic_4_110[tp_dic[tp][3]][4] += 1
                if tp_dic[tp][7] and tp_dic[tp][7] != 'Визуальный BD':
                    n_dic_4_110[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_4_110[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][7] == 'Визуальный BD' and tp_dic[tp][8]:
                    n_dic_4_110[tp_dic[tp][3]][5] += tp_dic[tp][5]
                    n_dic_4_110[tp_dic[tp][3]][6] += 1
                if tp_dic[tp][8]:
                    n_dic_4_110[tp_dic[tp][3]][7] += tp_dic[tp][5]
                    n_dic_4_110[tp_dic[tp][3]][8] += 1
                if tp_dic[tp][9]:
                    n_dic_4_110[tp_dic[tp][3]][9] += tp_dic[tp][5]
                    n_dic_4_110[tp_dic[tp][3]][10] += 1

        if tp_dic[tp][11] == '4' and 'SUPER' not in tp_dic[tp][0] and 'INTERFACE'\
                not in tp_dic[tp][0] and "n/d" not in tp_dic[tp][0]:
            summary_tp_phase_4.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                       tp_dic[tp][5], tp_dic[tp][6], tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                       tp_dic[tp][10]])

        if tp_dic[tp][11] == '5' and 'SUPER' not in tp_dic[tp][0] and 'INTERFACE'\
                not in tp_dic[tp][0] and "n/d" not in tp_dic[tp][0]:
            summary_tp_phase_5.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                       tp_dic[tp][5], tp_dic[tp][6], tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                       tp_dic[tp][10]])



    summary_list_units = []
    # for i in n_list_1_70:
    #     summary_list_units.append(i)
    # for key in n_dic_1_70.keys():
    #     summary_list_units.append([key, n_dic_1_70[key][0], n_dic_1_70[key][1], n_dic_1_70[key][2], n_dic_1_70[key][3],
    #                                n_dic_1_70[key][4], n_dic_1_70[key][5], n_dic_1_70[key][6], n_dic_1_70[key][7],
    #                                n_dic_1_70[key][8], n_dic_1_70[key][9], n_dic_1_70[key][10]])
    #     ITOG_list_1_70[2] += n_dic_1_70[key][1]
    #     ITOG_list_1_70[3] += n_dic_1_70[key][2]
    #     ITOG_list_1_70[4] += n_dic_1_70[key][3]
    #     ITOG_list_1_70[5] += n_dic_1_70[key][4]
    #     ITOG_list_1_70[6] += n_dic_1_70[key][5]
    #     ITOG_list_1_70[7] += n_dic_1_70[key][6]
    #     ITOG_list_1_70[8] += n_dic_1_70[key][7]
    #     ITOG_list_1_70[9] += n_dic_1_70[key][8]
    #     ITOG_list_1_70[10] += n_dic_1_70[key][9]
    #     ITOG_list_1_70[11] += n_dic_1_70[key][10]
    #
    # summary_list_units.append(ITOG_list_1_70)
    # ost_list_1_70 = ['', 'Остаток:', '', '', ITOG_list_1_70[2] - ITOG_list_1_70[4], ITOG_list_1_70[3] - ITOG_list_1_70[5],
    #                     ITOG_list_1_70[2] - ITOG_list_1_70[6], ITOG_list_1_70[3] - ITOG_list_1_70[7],
    #                     ITOG_list_1_70[2] - ITOG_list_1_70[8], ITOG_list_1_70[3] - ITOG_list_1_70[9],
    #                     ITOG_list_1_70[2] - ITOG_list_1_70[10], ITOG_list_1_70[3] - ITOG_list_1_70[11]]
    # summary_list_units.append(ost_list_1_70)
    # empty_str = [' * ', ' * ', ' * ', '  *  ', '  *  ', '  *  ', '  *  ', '  ', '  ', '  ', '  ', '']
    # summary_list_units.append(empty_str)
    #
    # for i in n_list_1_60:
    #     summary_list_units.append(i)
    # for key in n_dic_1_60.keys():
    #     summary_list_units.append([key, n_dic_1_60[key][0], n_dic_1_60[key][1], n_dic_1_60[key][2], n_dic_1_60[key][3],
    #                                n_dic_1_60[key][4], n_dic_1_60[key][5], n_dic_1_60[key][6], n_dic_1_60[key][7],
    #                                n_dic_1_60[key][8], n_dic_1_60[key][9], n_dic_1_60[key][10]])
    #     ITOG_list_1_60[2] += n_dic_1_60[key][1]
    #     ITOG_list_1_60[3] += n_dic_1_60[key][2]
    #     ITOG_list_1_60[4] += n_dic_1_60[key][3]
    #     ITOG_list_1_60[5] += n_dic_1_60[key][4]
    #     ITOG_list_1_60[6] += n_dic_1_60[key][5]
    #     ITOG_list_1_60[7] += n_dic_1_60[key][6]
    #     ITOG_list_1_60[8] += n_dic_1_60[key][7]
    #     ITOG_list_1_60[9] += n_dic_1_60[key][8]
    #     ITOG_list_1_60[10] += n_dic_1_60[key][9]
    #     ITOG_list_1_60[11] += n_dic_1_60[key][10]
    #
    # summary_list_units.append(ITOG_list_1_60)
    # ost_list_1_60 = ['', 'Остаток:', '', '', ITOG_list_1_60[2] - ITOG_list_1_60[4], ITOG_list_1_60[3] - ITOG_list_1_60[5],
    #                     ITOG_list_1_60[2] - ITOG_list_1_60[6], ITOG_list_1_60[3] - ITOG_list_1_60[7],
    #                     ITOG_list_1_60[2] - ITOG_list_1_60[8], ITOG_list_1_60[3] - ITOG_list_1_60[9],
    #                     ITOG_list_1_60[2] - ITOG_list_1_60[10], ITOG_list_1_60[3] - ITOG_list_1_60[11]]
    # summary_list_units.append(ost_list_1_60)
    # summary_list_units.append(empty_str)

    for i in n_list_2_70:
        summary_list_units.append(i)
    for key in n_dic_2_70.keys():
        summary_list_units.append([key, n_dic_2_70[key][0], n_dic_2_70[key][1], n_dic_2_70[key][2], n_dic_2_70[key][3],
                                   n_dic_2_70[key][4], n_dic_2_70[key][5], n_dic_2_70[key][6], n_dic_2_70[key][7],
                                   n_dic_2_70[key][8], n_dic_2_70[key][9], n_dic_2_70[key][10]])
        ITOG_list_2_70[2] += n_dic_2_70[key][1]
        ITOG_list_2_70[3] += n_dic_2_70[key][2]
        ITOG_list_2_70[4] += n_dic_2_70[key][3]
        ITOG_list_2_70[5] += n_dic_2_70[key][4]
        ITOG_list_2_70[6] += n_dic_2_70[key][5]
        ITOG_list_2_70[7] += n_dic_2_70[key][6]
        ITOG_list_2_70[8] += n_dic_2_70[key][7]
        ITOG_list_2_70[9] += n_dic_2_70[key][8]
        ITOG_list_2_70[10] += n_dic_2_70[key][9]
        ITOG_list_2_70[11] += n_dic_2_70[key][10]

    summary_list_units.append(ITOG_list_2_70)
    ost_list_2_70 = ['', 'Остаток:', '', '', ITOG_list_2_70[2] - ITOG_list_2_70[4], ITOG_list_2_70[3] - ITOG_list_2_70[5],
                        ITOG_list_2_70[2] - ITOG_list_2_70[6], ITOG_list_2_70[3] - ITOG_list_2_70[7],
                        ITOG_list_2_70[2] - ITOG_list_2_70[8], ITOG_list_2_70[3] - ITOG_list_2_70[9],
                        ITOG_list_2_70[2] - ITOG_list_2_70[10], ITOG_list_2_70[3] - ITOG_list_2_70[11]]
    summary_list_units.append(ost_list_2_70)
    empty_str = [' * ', ' * ', ' * ', '  *  ', '  *  ', '  *  ', '  *  ', '  ', '  ', '  ', '  ', '']
    summary_list_units.append(empty_str)

    for i in n_list_2_60:
        summary_list_units.append(i)
    for key in n_dic_2_60.keys():
        summary_list_units.append([key, n_dic_2_60[key][0], n_dic_2_60[key][1], n_dic_2_60[key][2], n_dic_2_60[key][3],
                                   n_dic_2_60[key][4], n_dic_2_60[key][5], n_dic_2_60[key][6], n_dic_2_60[key][7],
                                   n_dic_2_60[key][8], n_dic_2_60[key][9], n_dic_2_60[key][10]])
        ITOG_list_2_60[2] += n_dic_2_60[key][1]
        ITOG_list_2_60[3] += n_dic_2_60[key][2]
        ITOG_list_2_60[4] += n_dic_2_60[key][3]
        ITOG_list_2_60[5] += n_dic_2_60[key][4]
        ITOG_list_2_60[6] += n_dic_2_60[key][5]
        ITOG_list_2_60[7] += n_dic_2_60[key][6]
        ITOG_list_2_60[8] += n_dic_2_60[key][7]
        ITOG_list_2_60[9] += n_dic_2_60[key][8]
        ITOG_list_2_60[10] += n_dic_2_60[key][9]
        ITOG_list_2_60[11] += n_dic_2_60[key][10]

    summary_list_units.append(ITOG_list_2_60)
    ost_list_2_60 = ['', 'Остаток:', '', '', ITOG_list_2_60[2] - ITOG_list_2_60[4], ITOG_list_2_60[3] - ITOG_list_2_60[5],
                        ITOG_list_2_60[2] - ITOG_list_2_60[6], ITOG_list_2_60[3] - ITOG_list_2_60[7],
                        ITOG_list_2_60[2] - ITOG_list_2_60[8], ITOG_list_2_60[3] - ITOG_list_2_60[9],
                        ITOG_list_2_60[2] - ITOG_list_2_60[10], ITOG_list_2_60[3] - ITOG_list_2_60[11]]
    summary_list_units.append(ost_list_2_60)
    summary_list_units.append(empty_str)

    for i in n_list_4_110:
        summary_list_units.append(i)
    for key in n_dic_4_110.keys():
        summary_list_units.append([key, n_dic_4_110[key][0], n_dic_4_110[key][1], n_dic_4_110[key][2], n_dic_4_110[key][3],
                                   n_dic_4_110[key][4], n_dic_4_110[key][5], n_dic_4_110[key][6], n_dic_4_110[key][7],
                                   n_dic_4_110[key][8], n_dic_4_110[key][9], n_dic_4_110[key][10]])
        ITOG_list_4_110[2] += n_dic_4_110[key][1]
        ITOG_list_4_110[3] += n_dic_4_110[key][2]
        ITOG_list_4_110[4] += n_dic_4_110[key][3]
        ITOG_list_4_110[5] += n_dic_4_110[key][4]
        ITOG_list_4_110[6] += n_dic_4_110[key][5]
        ITOG_list_4_110[7] += n_dic_4_110[key][6]
        ITOG_list_4_110[8] += n_dic_4_110[key][7]
        ITOG_list_4_110[9] += n_dic_4_110[key][8]
        ITOG_list_4_110[10] += n_dic_4_110[key][9]
        ITOG_list_4_110[11] += n_dic_4_110[key][10]

    summary_list_units.append(ITOG_list_4_110)
    ost_list_4_110 = ['', 'Остаток:', '', '', ITOG_list_4_110[2] - ITOG_list_4_110[4], ITOG_list_4_110[3] - ITOG_list_4_110[5],
                        ITOG_list_4_110[2] - ITOG_list_4_110[6], ITOG_list_4_110[3] - ITOG_list_4_110[7],
                        ITOG_list_4_110[2] - ITOG_list_4_110[8], ITOG_list_4_110[3] - ITOG_list_4_110[9],
                        ITOG_list_4_110[2] - ITOG_list_4_110[10], ITOG_list_4_110[3] - ITOG_list_4_110[11]]
    summary_list_units.append(ost_list_4_110)
    summary_list_units.append(empty_str)

    for i in n_list_4_30:
        summary_list_units.append(i)
    for key in n_dic_4_30.keys():
        summary_list_units.append([key, n_dic_4_30[key][0], n_dic_4_30[key][1], n_dic_4_30[key][2], n_dic_4_30[key][3],
                                   n_dic_4_30[key][4], n_dic_4_30[key][5], n_dic_4_30[key][6], n_dic_4_30[key][7],
                                   n_dic_4_30[key][8], n_dic_4_30[key][9], n_dic_4_30[key][10]])
        ITOG_list_4_30[2] += n_dic_4_30[key][1]
        ITOG_list_4_30[3] += n_dic_4_30[key][2]
        ITOG_list_4_30[4] += n_dic_4_30[key][3]
        ITOG_list_4_30[5] += n_dic_4_30[key][4]
        ITOG_list_4_30[6] += n_dic_4_30[key][5]
        ITOG_list_4_30[7] += n_dic_4_30[key][6]
        ITOG_list_4_30[8] += n_dic_4_30[key][7]
        ITOG_list_4_30[9] += n_dic_4_30[key][8]
        ITOG_list_4_30[10] += n_dic_4_30[key][9]
        ITOG_list_4_30[11] += n_dic_4_30[key][10]

    summary_list_units.append(ITOG_list_4_30)
    ost_list_4_30 = ['', 'Остаток:', '', '', ITOG_list_4_30[2] - ITOG_list_4_30[4], ITOG_list_4_30[3] - ITOG_list_4_30[5],
                        ITOG_list_4_30[2] - ITOG_list_4_30[6], ITOG_list_4_30[3] - ITOG_list_4_30[7],
                        ITOG_list_4_30[2] - ITOG_list_4_30[8], ITOG_list_4_30[3] - ITOG_list_4_30[9],
                        ITOG_list_4_30[2] - ITOG_list_4_30[10], ITOG_list_4_30[3] - ITOG_list_4_30[11]]
    summary_list_units.append(ost_list_4_30)
    summary_list_units.append(empty_str)


    """
    Запись БД для работы скрипта AICT
    """
    with open('bd_tp_rfi.csv', 'w', newline='') as write_file:
        writed_file = csv.writer(write_file, delimiter=";")
        writed_file.writerows(summary_tp_for_db_atom)
        print('БД для АТОМа создана')


    workbook_summary = xlsxwriter.Workbook(f'{path}\\Сводка по ФАЗАМ на {datetime.datetime.now().strftime("%d.%m.%Y")}.xlsx')


    # -------------------------------------Краткая сводка
    ws0 = workbook_summary.add_worksheet('Краткая сводка по установкам')
    ws0.set_column(0, 0, 12)
    ws0.set_column(1, 1, 40)
    ws0.set_column(4, 15, 12)
    ws0.set_column(2, 2, 12)
    ws0.set_column(3, 3, 12)

    cell_format_green = workbook_summary.add_format()
    cell_format_green.set_bg_color('#99FF99')
    cell_format_blue = workbook_summary.add_format()
    cell_format_blue.set_bg_color('#99CCCC')
    cell_format_hat = workbook_summary.add_format()
    cell_format_hat.set_bg_color('#FF9966')
    cell_format_date = workbook_summary.add_format()
    cell_format_date.set_font_size(font_size=14)
    for i, (one, two, three, four, five, six, seven, eight, nine, ten, eleven, twelve) in enumerate(summary_list_units, start=2):

        if two == 'Итого:' or 'Статус' in two or 'Наименование' in two or two == 'Остаток:':
            color = cell_format_hat
            color.set_bold('bold')

        else:
            try:
                oos = float(str(four)) - float(str(twelve))

                if oos == 0:
                    color = cell_format_green
                else:
                    color = cell_format_blue
            except Exception as e:
                pass


        try:
            color.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
        except:
            pass
        ws0.write(f'A{i}', one, color)
        ws0.write(f'B{i}', two, color)
        ws0.write(f'C{i}', three, color)
        ws0.write(f'D{i}', four, color)
        ws0.write(f'E{i}', five, color)
        ws0.write(f'F{i}', six, color)
        ws0.write(f'G{i}', seven, color)
        ws0.write(f'H{i}', eight, color)
        ws0.write(f'I{i}', nine, color)
        ws0.write(f'J{i}', ten, color)
        ws0.write(f'K{i}', eleven, color)
        ws0.write(f'L{i}', twelve, color)






    # -------------------------------------------------------------------------------------------------
    ws5 = workbook_summary.add_worksheet('Сводка ТП ФАЗА 1')

    ws5.set_column(0, 0, 30)
    ws5.set_column(1, 5, 15)
    ws5.set_column(6, 11, 22)
    ws5.set_column(12, 17, 25)
    ws5.set_column(18, 18, 13)
    ws5.autofilter('A1:S11682')

    for i, (testpackage, title, unit, fluid, ggn_status, iso_length, rfi_erection, rfi_test,
            rfi_airblowing, rfi_reinstatement, ncr_status) in enumerate(summary_tp_phase_1, start=1):

        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
        elif '-CC' in rfi_reinstatement:
            color = cell_format_green
        else:
            color = cell_format_blue

        try:
            color.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
        except:
            pass

        ws5.write(f'A{i}', testpackage, color)
        ws5.write(f'B{i}', title, color)
        ws5.write(f'C{i}', unit, color)
        ws5.write(f'D{i}', fluid, color)
        ws5.write(f'E{i}', ggn_status, color)
        ws5.write(f'F{i}', iso_length, color)
        ws5.write(f'G{i}', rfi_erection, color)
        ws5.write(f'H{i}', rfi_test, color)
        ws5.write(f'I{i}', rfi_airblowing, color)
        ws5.write(f'J{i}', rfi_reinstatement, color)
        ws5.write(f'K{i}', ncr_status, color)



    # -------------------------------------------------------------------------------------------------
    ws1 = workbook_summary.add_worksheet('Сводка ТП ФАЗА 2')

    ws1.set_column(0, 0, 30)
    ws1.set_column(1, 5, 15)
    ws1.set_column(6, 11, 22)
    ws1.set_column(12, 17, 25)
    ws1.set_column(18, 18, 13)
    ws1.autofilter('A1:S1682')

    for i, (testpackage, title, unit, fluid, ggn_status, iso_length, rfi_erection, rfi_test,
            rfi_airblowing, rfi_reinstatement, ncr_status) in enumerate(summary_tp_phase_2, start=1):

        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
        elif '-CC' in rfi_reinstatement:
            color = cell_format_green
        else:
            color = cell_format_blue

        try:
            color.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
        except:
            pass

        ws1.write(f'A{i}', testpackage, color)
        ws1.write(f'B{i}', title, color)
        ws1.write(f'C{i}', unit, color)
        ws1.write(f'D{i}', fluid, color)
        ws1.write(f'E{i}', ggn_status, color)
        ws1.write(f'F{i}', iso_length, color)
        ws1.write(f'G{i}', rfi_erection, color)
        ws1.write(f'H{i}', rfi_test, color)
        ws1.write(f'I{i}', rfi_airblowing, color)
        ws1.write(f'J{i}', rfi_reinstatement, color)
        ws1.write(f'K{i}', ncr_status, color)


    # -------------------------------------------------------------------------------------------------
    ws2 = workbook_summary.add_worksheet('Сводка ТП ФАЗА 3')

    ws2.set_column(0, 0, 30)
    ws2.set_column(1, 5, 15)
    ws2.set_column(6, 11, 22)
    ws2.set_column(12, 17, 25)
    ws2.set_column(18, 18, 13)
    ws2.autofilter('A1:S1682')

    for i, (testpackage, title, unit, fluid, ggn_status, iso_length, rfi_erection, rfi_test,
            rfi_airblowing, rfi_reinstatement, ncr_status) in enumerate(summary_tp_phase_3, start=1):

        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
        elif '-CC' in rfi_reinstatement:
            color = cell_format_green
        else:
            color = cell_format_blue

        try:
            color.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
        except:
            pass

        ws2.write(f'A{i}', testpackage, color)
        ws2.write(f'B{i}', title, color)
        ws2.write(f'C{i}', unit, color)
        ws2.write(f'D{i}', fluid, color)
        ws2.write(f'E{i}', ggn_status, color)
        ws2.write(f'F{i}', iso_length, color)
        ws2.write(f'G{i}', rfi_erection, color)
        ws2.write(f'H{i}', rfi_test, color)
        ws2.write(f'I{i}', rfi_airblowing, color)
        ws2.write(f'J{i}', rfi_reinstatement, color)
        ws2.write(f'K{i}', ncr_status, color)

    # -------------------------------------------------------------------------------------------------
    ws4 = workbook_summary.add_worksheet('Сводка ТП ФАЗА 4')

    ws4.set_column(0, 0, 30)
    ws4.set_column(1, 5, 15)
    ws4.set_column(6, 11, 22)
    ws4.set_column(12, 17, 25)
    ws4.set_column(18, 18, 13)
    ws4.autofilter(f'A1:S{len(summary_tp_phase_4)}')

    for i, (testpackage, title, unit, fluid, ggn_status, iso_length, rfi_erection, rfi_test,
            rfi_airblowing, rfi_reinstatement, ncr_status) in enumerate(summary_tp_phase_4, start=1):
        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
        elif '-CC' in rfi_reinstatement:
            color = cell_format_green
        else:
            color = cell_format_blue

        try:
            color.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
        except:
            pass

        ws4.write(f'A{i}', testpackage, color)
        ws4.write(f'B{i}', title, color)
        ws4.write(f'C{i}', unit, color)
        ws4.write(f'D{i}', fluid, color)
        ws4.write(f'E{i}', ggn_status, color)
        ws4.write(f'F{i}', iso_length, color)
        ws4.write(f'G{i}', rfi_erection, color)
        ws4.write(f'H{i}', rfi_test, color)
        ws4.write(f'I{i}', rfi_airblowing, color)
        ws4.write(f'J{i}', rfi_reinstatement, color)
        ws4.write(f'K{i}', ncr_status, color)

    # -------------------------------------------------------------------------------------------------
    ws5 = workbook_summary.add_worksheet('Сводка ТП ФАЗА 5')

    ws5.set_column(0, 0, 30)
    ws5.set_column(1, 5, 15)
    ws5.set_column(6, 11, 22)
    ws5.set_column(12, 17, 25)
    ws5.set_column(18, 18, 13)
    ws5.autofilter(f'A1:S{len(summary_tp_phase_5)}')

    for i, (testpackage, title, unit, fluid, ggn_status, iso_length, rfi_erection, rfi_test,
            rfi_airblowing, rfi_reinstatement, ncr_status) in enumerate(summary_tp_phase_5, start=1):
        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
        elif '-CC' in rfi_reinstatement:
            color = cell_format_green
        else:
            color = cell_format_blue

        try:
            color.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
        except:
            pass

        ws5.write(f'A{i}', testpackage, color)
        ws5.write(f'B{i}', title, color)
        ws5.write(f'C{i}', unit, color)
        ws5.write(f'D{i}', fluid, color)
        ws5.write(f'E{i}', ggn_status, color)
        ws5.write(f'F{i}', iso_length, color)
        ws5.write(f'G{i}', rfi_erection, color)
        ws5.write(f'H{i}', rfi_test, color)
        ws5.write(f'I{i}', rfi_airblowing, color)
        ws5.write(f'J{i}', rfi_reinstatement, color)
        ws5.write(f'K{i}', ncr_status, color)


    # Сводка по изометриям

    wsi1 = workbook_summary.add_worksheet('Сводка по ИЗО ФАЗА 1')
    wsi1.set_column(0, 1, 32)
    wsi1.set_column(2, 7, 14)
    wsi1.set_column(8, 11, 20)

    wsi1.set_column(12, 13, 13)
    wsi1.set_column(14, 16, 20)
    wsi1.set_column(17, 18, 14)

    cell_format_ins = workbook_summary.add_format()
    cell_format_ins.set_bg_color('#FFEBCD')

    wsi1.autofilter('A1:S20000')
    for i, (testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length, rfi_erection,
            rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins, rfi_ins_cotton,
            rfi_ins_metall, rfi_ins_box, rfi_dig, ncr) in enumerate(summary_iso_tp_phase_1, start=1):

        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
            color.set_text_wrap(text_wrap=1)
            color_2 = cell_format_hat
            color_2.set_bold('bold')
            color_2.set_text_wrap(text_wrap=1)

        elif rfi_reinstatement:
            color = cell_format_green
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        else:
            color = cell_format_blue
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        try:
            color.set_border(style=1)
            color_2.set_border(style=1)
        except:
            pass

        wsi1.write(f'A{i}', testpackage, color)
        wsi1.write(f'B{i}', isometric, color)
        wsi1.write(f'C{i}', line, color)
        wsi1.write(f'D{i}', title, color)
        wsi1.write(f'E{i}', unit, color)
        wsi1.write(f'F{i}', fluid, color)
        wsi1.write(f'G{i}', ggn_status, color)
        wsi1.write(f'H{i}', iso_length, color)
        wsi1.write(f'I{i}', rfi_erection, color)
        wsi1.write(f'J{i}', rfi_test, color)
        wsi1.write(f'K{i}', rfi_airblowing, color)
        wsi1.write(f'L{i}', rfi_reinstatement, color)
        wsi1.write(f'M{i}', type_ins, color_2)
        wsi1.write(f'N{i}', volume_ins, color_2)
        wsi1.write(f'O{i}', rfi_ins_cotton, color_2)
        wsi1.write(f'P{i}', rfi_ins_metall, color_2)
        wsi1.write(f'Q{i}', rfi_ins_box, color_2)
        wsi1.write(f'R{i}', rfi_dig, color)
        wsi1.write(f'S{i}', ncr, color)


    wsi2 = workbook_summary.add_worksheet('Сводка по ИЗО ФАЗА 2')
    wsi2.set_column(0, 1, 32)
    wsi2.set_column(2, 7, 14)
    wsi2.set_column(8, 11, 20)

    wsi2.set_column(12, 13, 13)
    wsi2.set_column(14, 16, 20)
    wsi2.set_column(17, 18, 14)

    cell_format_ins = workbook_summary.add_format()
    cell_format_ins.set_bg_color('#FFEBCD')

    wsi2.autofilter('A1:S20000')
    for i, (testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length, rfi_erection,
            rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins, rfi_ins_cotton,
            rfi_ins_metall, rfi_ins_box, rfi_dig, ncr) in enumerate(summary_iso_tp_phase_2, start=1):

        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
            color.set_text_wrap(text_wrap=1)
            color_2 = cell_format_hat
            color_2.set_bold('bold')
            color_2.set_text_wrap(text_wrap=1)

        elif rfi_reinstatement:
            color = cell_format_green
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        else:
            color = cell_format_blue
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        try:
            color.set_border(style=1)
            color_2.set_border(style=1)
        except:
            pass

        wsi2.write(f'A{i}', testpackage, color)
        wsi2.write(f'B{i}', isometric, color)
        wsi2.write(f'C{i}', line, color)
        wsi2.write(f'D{i}', title, color)
        wsi2.write(f'E{i}', unit, color)
        wsi2.write(f'F{i}', fluid, color)
        wsi2.write(f'G{i}', ggn_status, color)
        wsi2.write(f'H{i}', iso_length, color)
        wsi2.write(f'I{i}', rfi_erection, color)
        wsi2.write(f'J{i}', rfi_test, color)
        wsi2.write(f'K{i}', rfi_airblowing, color)
        wsi2.write(f'L{i}', rfi_reinstatement, color)
        wsi2.write(f'M{i}', type_ins, color_2)
        wsi2.write(f'N{i}', volume_ins, color_2)
        wsi2.write(f'O{i}', rfi_ins_cotton, color_2)
        wsi2.write(f'P{i}', rfi_ins_metall, color_2)
        wsi2.write(f'Q{i}', rfi_ins_box, color_2)
        wsi2.write(f'R{i}', rfi_dig, color)
        wsi2.write(f'S{i}', ncr, color)


    wsi3 = workbook_summary.add_worksheet('Сводка по ИЗО ФАЗА 3')
    wsi3.set_column(0, 1, 32)
    wsi3.set_column(2, 7, 14)
    wsi3.set_column(8, 11, 20)

    wsi3.set_column(12, 13, 13)
    wsi3.set_column(14, 16, 20)
    wsi3.set_column(17, 18, 14)

    cell_format_ins = workbook_summary.add_format()
    cell_format_ins.set_bg_color('#FFEBCD')

    wsi3.autofilter('A1:S20000')
    for i, (testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length, rfi_erection,
            rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins, rfi_ins_cotton,
            rfi_ins_metall, rfi_ins_box, rfi_dig, ncr) in enumerate(summary_iso_tp_phase_3, start=1):

        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
            color.set_text_wrap(text_wrap=1)
            color_2 = cell_format_hat
            color_2.set_bold('bold')
            color_2.set_text_wrap(text_wrap=1)

        elif rfi_reinstatement:
            color = cell_format_green
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        else:
            color = cell_format_blue
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        try:
            color.set_border(style=1)
            color_2.set_border(style=1)
        except:
            pass

        wsi3.write(f'A{i}', testpackage, color)
        wsi3.write(f'B{i}', isometric, color)
        wsi3.write(f'C{i}', line, color)
        wsi3.write(f'D{i}', title, color)
        wsi3.write(f'E{i}', unit, color)
        wsi3.write(f'F{i}', fluid, color)
        wsi3.write(f'G{i}', ggn_status, color)
        wsi3.write(f'H{i}', iso_length, color)
        wsi3.write(f'I{i}', rfi_erection, color)
        wsi3.write(f'J{i}', rfi_test, color)
        wsi3.write(f'K{i}', rfi_airblowing, color)
        wsi3.write(f'L{i}', rfi_reinstatement, color)
        wsi3.write(f'M{i}', type_ins, color_2)
        wsi3.write(f'N{i}', volume_ins, color_2)
        wsi3.write(f'O{i}', rfi_ins_cotton, color_2)
        wsi3.write(f'P{i}', rfi_ins_metall, color_2)
        wsi3.write(f'Q{i}', rfi_ins_box, color_2)
        wsi3.write(f'R{i}', rfi_dig, color)
        wsi3.write(f'S{i}', ncr, color)

    # ---------------------------------------------------------------------------
    wsi4 = workbook_summary.add_worksheet('Сводка по ИЗО ФАЗА 4')
    wsi4.set_column(0, 1, 32)
    wsi4.set_column(2, 7, 14)
    wsi4.set_column(8, 11, 20)

    wsi4.set_column(12, 13, 13)
    wsi4.set_column(14, 16, 20)
    wsi4.set_column(17, 18, 14)

    cell_format_ins = workbook_summary.add_format()
    cell_format_ins.set_bg_color('#FFEBCD')

    wsi4.autofilter('A1:S20000')
    for i, (testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length, rfi_erection,
            rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins, rfi_ins_cotton,
            rfi_ins_metall, rfi_ins_box, rfi_dig, ncr) in enumerate(summary_iso_tp_phase_4, start=1):

        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
            color.set_text_wrap(text_wrap=1)
            color_2 = cell_format_hat
            color_2.set_bold('bold')
            color_2.set_text_wrap(text_wrap=1)

        elif rfi_reinstatement:
            color = cell_format_green
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        else:
            color = cell_format_blue
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        try:
            color.set_border(style=1)
            color_2.set_border(style=1)
        except:
            pass

        wsi4.write(f'A{i}', testpackage, color)
        wsi4.write(f'B{i}', isometric, color)
        wsi4.write(f'C{i}', line, color)
        wsi4.write(f'D{i}', title, color)
        wsi4.write(f'E{i}', unit, color)
        wsi4.write(f'F{i}', fluid, color)
        wsi4.write(f'G{i}', ggn_status, color)
        wsi4.write(f'H{i}', iso_length, color)
        wsi4.write(f'I{i}', rfi_erection, color)
        wsi4.write(f'J{i}', rfi_test, color)
        wsi4.write(f'K{i}', rfi_airblowing, color)
        wsi4.write(f'L{i}', rfi_reinstatement, color)
        wsi4.write(f'M{i}', type_ins, color_2)
        wsi4.write(f'N{i}', volume_ins, color_2)
        wsi4.write(f'O{i}', rfi_ins_cotton, color_2)
        wsi4.write(f'P{i}', rfi_ins_metall, color_2)
        wsi4.write(f'Q{i}', rfi_ins_box, color_2)
        wsi4.write(f'R{i}', rfi_dig, color)
        wsi4.write(f'S{i}', ncr, color)

    # ---------------------------------------------------------------------------
    wsi5 = workbook_summary.add_worksheet('Сводка по ИЗО ФАЗА 5')
    wsi5.set_column(0, 1, 32)
    wsi5.set_column(2, 7, 14)
    wsi5.set_column(8, 11, 20)

    wsi5.set_column(12, 13, 13)
    wsi5.set_column(14, 16, 20)
    wsi5.set_column(17, 18, 14)

    cell_format_ins = workbook_summary.add_format()
    cell_format_ins.set_bg_color('#FFEBCD')

    wsi5.autofilter('A1:S20000')
    for i, (testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length, rfi_erection,
            rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins, rfi_ins_cotton,
            rfi_ins_metall, rfi_ins_box, rfi_dig, ncr) in enumerate(summary_iso_tp_phase_5, start=1):

        if testpackage == 'Тестпакет':
            color = cell_format_hat
            color.set_bold('bold')
            color.set_text_wrap(text_wrap=1)
            color_2 = cell_format_hat
            color_2.set_bold('bold')
            color_2.set_text_wrap(text_wrap=1)

        elif rfi_reinstatement:
            color = cell_format_green
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        else:
            color = cell_format_blue
            color.set_text_wrap(text_wrap=0)
            color_2 = cell_format_ins
            color_2.set_text_wrap(text_wrap=0)

        try:
            color.set_border(style=1)
            color_2.set_border(style=1)
        except:
            pass

        wsi5.write(f'A{i}', testpackage, color)
        wsi5.write(f'B{i}', isometric, color)
        wsi5.write(f'C{i}', line, color)
        wsi5.write(f'D{i}', title, color)
        wsi5.write(f'E{i}', unit, color)
        wsi5.write(f'F{i}', fluid, color)
        wsi5.write(f'G{i}', ggn_status, color)
        wsi5.write(f'H{i}', iso_length, color)
        wsi5.write(f'I{i}', rfi_erection, color)
        wsi5.write(f'J{i}', rfi_test, color)
        wsi5.write(f'K{i}', rfi_airblowing, color)
        wsi5.write(f'L{i}', rfi_reinstatement, color)
        wsi5.write(f'M{i}', type_ins, color_2)
        wsi5.write(f'N{i}', volume_ins, color_2)
        wsi5.write(f'O{i}', rfi_ins_cotton, color_2)
        wsi5.write(f'P{i}', rfi_ins_metall, color_2)
        wsi5.write(f'Q{i}', rfi_ins_box, color_2)
        wsi5.write(f'R{i}', rfi_dig, color)
        wsi5.write(f'S{i}', ncr, color)

    workbook_summary.close()

    print('Всё записал.')


