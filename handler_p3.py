import csv
import os

import openpyxl as xl
import re
import datetime
import xlsxwriter
import pandas as pd


def read_rfi_journal():

    # Создание словарей для тестпакетов и изометрий
    tps = {}
    isos = {}
    isos_db = {}

    ehts = []  # греющий кабель
    hts = []  # Теплоспутник

    with open(f"{os.getcwd()}\\dbs_ti_p3\\bd_tp_p3.csv", "r") as read_file:
        readed_file = csv.reader(read_file, delimiter=";")
        for row in readed_file:
            if "Длина" not in row[3]:
                phase = row[0]
                testpackage = row[1]
                isometric = row[2]
                length_tt = round(float(row[3]), 2)
                length_eht = round(float(row[4]), 2)
                value_akz = round(float(row[5]), 2)
                value_ins = round(float(row[6]), 2)
                count_zra = int(row[7])

                key = isometric + "$" + testpackage

                isos[key] = [phase, testpackage, length_tt, "", "", "", value_akz, "", length_eht, "",
                             value_ins, count_zra, "", ""]

                isos_db[key] = [phase, testpackage, length_tt, "", "", "", "", "", "", "not used", "", "", "", "",
                                "", "", "", "", "", "", count_zra, "", "", "", ""]

                tps[row[1]] = [0, "", "", "", "", "", "", "", ""]
    read_file.close()

    # Чтение из накопленных баз
    with open(f"{os.getcwd()}\\dbs_ti_p3\\summary_accum_db_rfi.csv", "r") as read_file:
        readed_file = csv.reader(read_file, delimiter=";")
        for row in readed_file:
            isos[row[0]][3] = row[4]
            isos[row[0]][4] = row[5]
            isos[row[0]][5] = row[6]
            isos[row[0]][7] = row[8]
            isos[row[0]][9] = row[10]
            isos[row[0]][12] = row[13]
            isos[row[0]][13] = row[14]
    read_file.close()

    with open(f"{os.getcwd()}\\dbs_ti_p3\\summary_accum_db_tps_rfi.csv", "r") as read_file:
        readed_file = csv.reader(read_file, delimiter=";")
        for row in readed_file:
            tps[row[0]][1] = row[1]
            tps[row[0]][2] = row[2]
            tps[row[0]][3] = row[3]
            tps[row[0]][4] = row[4]
            tps[row[0]][5] = row[5]
            tps[row[0]][6] = row[6]
            tps[row[0]][7] = row[7]
            tps[row[0]][8] = row[8]
    read_file.close()

    with open(f"{os.getcwd()}\\dbs_ti_p3\\iso_ins_db_rfi.csv", "r") as read_file:
        readed_file = csv.reader(read_file, delimiter=";")
        for row in readed_file:
            if "Изометрия" not in row[0]:
                key = row[0] + "$" + row[1]
                isos_db[key][3] = row[2]
                isos_db[key][4] = row[3]
                isos_db[key][5] = row[4]
                isos_db[key][6] = row[5]
                isos_db[key][7] = row[6]
                isos_db[key][8] = row[7]

                isos_db[key][10] = row[9]
                isos_db[key][11] = row[10]
                isos_db[key][12] = row[11]
                isos_db[key][13] = row[12]
                isos_db[key][14] = row[13]
                isos_db[key][15] = row[14]
                isos_db[key][16] = row[15]
                isos_db[key][17] = row[16]
                isos_db[key][18] = row[17]
                isos_db[key][19] = row[18]

                isos_db[key][21] = row[20]
                isos_db[key][22] = row[21]
                isos_db[key][23] = row[22]
                isos_db[key][24] = row[23]
    read_file.close()

    with open(f"{os.getcwd()}\\dbs_ti_p3\\bd_tp_p3.csv", "r") as read_file:
        readed_file = csv.reader(read_file, delimiter=";")
        for row in readed_file:
            if "Длина" not in row[3]:
                tps[row[1]][0] += round(float(row[3]), 2)
    read_file.close()

    # Греющий кабель
    with open(f"{os.getcwd()}\\dbs_ti_p3\\bd_iso_eht_p3.csv", "r") as read_file:
        readed_file = csv.reader(read_file, delimiter=";")
        for row in readed_file:
            if "Фаза" not in row[0] and row[3] != "-":
                try:
                    isos_db[row[2].strip() + "$" + row[1]][9] = row[3]
                except:
                    pass
                ehts.append([row[3], row[2].strip() + "$" + row[1]])
    read_file.close()

    # Теплоспутник подающие манифольды
    with open(f"{os.getcwd()}\\dbs_ti_p3\\bd_iso_ht_p3.csv", "r") as read_file:
        readed_file = csv.reader(read_file, delimiter=";")
        for row in readed_file:
            if "Фаза" not in row[0]:
                hts.append([row[4], row[2].strip() + "$" + row[1]])
    read_file.close()

    # wb_journal_rfi = xl.load_workbook(f'{os.getcwd()}\\xlsx\\Журнал заявок общий.xlsx')
    # sheet_journal_rfi = wb_journal_rfi['Sheet1']

    wb_journal_rfi = xl.load_workbook(f'{os.getcwd()}\\xlsx\\Инспекции.xlsx')
    sheet_journal_rfi = wb_journal_rfi['Инспекции']

    for i in sheet_journal_rfi['B2':'AZ550000']:
        if i[0].value:
            rfi_number = str(i[1].value).strip()

            pkk = str(i[4].value).strip()

            description_rfi = str(i[18].value)
            violation = str(i[39].value)
            list_iso = str(i[8].value).replace(' ', '').split(';')
            category_cancelled = str(i[34].value).strip()
            conclusion_sk = str(i[44].value).strip()
            comment_sk = str(i[48].value).strip()

            date_of_submissions = str(i[13].value)

            # ПОИСК ПО ТЕСТПАКЕТАМ

            # МОНТАЖ ТРУБОПРОВОДА
            if "роверка" in description_rfi and "еред" in description_rfi:
                if 'Принято' == category_cancelled:
                    for tp in tps:
                        if tp in description_rfi:
                            tps[tp][1] = rfi_number
                            # tps[tp][2] = conclusion_sk + "/ \n /" + comment_sk
                            tps[tp][2] = "***"

                if "Принято с замечаниями" == category_cancelled:
                    if "роверка" in description_rfi and "еред" in description_rfi:
                        for tp in tps:
                            if tp in description_rfi:
                                tps[tp][1] = rfi_number + " ПЗ"
                                if comment_sk != 'None':
                                    tps[tp][2] = conclusion_sk + " //// " + comment_sk
                                else:
                                    tps[tp][2] = conclusion_sk

            # ИСПЫТАНИЯ ТРУБОПРОВОДА
            # ПЕРЕЧЕНЬ условий по поиску описания на ТЕСТЫ
            if "Визуальный тест" in description_rfi or "невматическое испытан" in description_rfi or \
                    "изуальный осмот" in description_rfi or "изуального Осмотр" in description_rfi or \
                    "изуального испытан" in description_rfi or "изуального осмотр" in description_rfi or \
                    "ИЗУАЛЬНОГО ИСПЫТАН" in description_rfi or "невматические испытан" in description_rfi or \
                    "влических испытан" in description_rfi or "лическое испытан" in description_rfi or \
                    "тическое испытан" in description_rfi or "лические испытан" in description_rfi or \
                    "тического испытан" in description_rfi or "зуального  испытан" in description_rfi or \
                    "зуальных  испытан" in description_rfi or "тических испытан" in description_rfi or \
                    "гидроиспытания" in description_rfi or "зуальное испытан" in description_rfi or \
                    "зуальный испытан" in description_rfi or "пневмоиспытан" in description_rfi or \
                    "тическое  испытан" in description_rfi or "зуальных испытан" in description_rfi or \
                    "Hydrostatic" in description_rfi or "гидро испытан" in description_rfi or \
                    "лических  испытан" in description_rfi or "HYDRO  TEST" in description_rfi or \
                    "HYDRO TEST" in description_rfi or "VISUAL  TEST" in description_rfi or \
                    "VISUAL TEST" in description_rfi or "тического  испытан" in description_rfi or \
                    "зуальные испытан" in description_rfi or "тические  испытан" in description_rfi or \
                    "изуальному осмотр" in description_rfi or "гидроиспытан" in description_rfi or \
                    "тические испытан" in description_rfi:

                if 'Принято' == category_cancelled:
                    for tp in tps:
                        if tp in description_rfi:
                            tps[tp][3] = rfi_number
                            tps[tp][4] = "***"

                if "Принято с замечаниями" == category_cancelled:
                    for tp in tps:
                        if tp in description_rfi:
                            tps[tp][3] = rfi_number + " ПЗ"
                            if comment_sk != "None":
                                tps[tp][4] = conclusion_sk + " //// " + comment_sk
                            else:
                                tps[tp][4] = conclusion_sk

            # ПРОДУВКА ТРУБОПРОВОДА
            # if "родувка" in description_rfi:
            #     if 'Принято' == category_cancelled:
            #         for tp in tps:
            #             if tp in description_rfi:
            #                 tps[tp][1] = rfi_number
            #                 # tps[tp][2] = conclusion_sk + "/ \n /" + comment_sk
            #                 tps[tp][2] = "***"
            #
            #                 print(tp, " - blowing - ", rfi_number)
            #
            #     if "Принято с замечаниями" == category_cancelled:
            #         for tp in tps:
            #             if tp in description_rfi:
            #                 tps[tp][1] = rfi_number + " ПЗ"
            #                 tps[tp][2] = conclusion_sk + "/ \n /" + comment_sk

            if "ратная сборк" in description_rfi or "ратной сборк" in description_rfi or \
                    "einstatement" in description_rfi:
                if 'Принято' == category_cancelled:
                    for tp in tps:
                        if tp in description_rfi:
                            tps[tp][7] = rfi_number
                            tps[tp][8] = "***"

                if "Принято с замечаниями" == category_cancelled:
                    for tp in tps:
                        if tp in description_rfi:
                            tps[tp][7] = rfi_number + " ПЗ"
                            if comment_sk != "None":
                                tps[tp][8] = conclusion_sk + " //// " + comment_sk
                            else:
                                tps[tp][8] = conclusion_sk

            # ПОИСК ПО ИЗОМЕТРИЯМ
            # ФИНАЛЬНЫЕ ИНСПЕКЦИИ АКЗ трубопровода

            if "inal acceptance of ACP pipe" in description_rfi or "инальная приёмка АКЗ трубо" in description_rfi or \
                    "приёмка АКЗ подземных трубоп" in description_rfi:
                if 'Принято' in category_cancelled:
                    for key in isos:
                        iso = key.split("$")[0]
                        if iso in description_rfi:
                            if "Принято" == category_cancelled:
                                isos[key][7] = rfi_number

                                isos_db[key][5] = rfi_number
                                isos_db[key][6] = "***"
                            if "Принято с замечаниями" == category_cancelled:
                                isos[key][7] = rfi_number + " ПЗ"

                                isos_db[key][5] = rfi_number + " ПЗ"
                                if comment_sk != "None":
                                    isos_db[key][6] = conclusion_sk + " //// " + comment_sk
                                else:
                                    isos_db[key][6] = conclusion_sk

            # ИНСПЕКЦИИ НА МОНТАЖ ГРЕЮЩЕГО КАБЕЛЯ
            # поиск по тэгу кабеля, привязка к изометрии

            if "роверка кабель" in description_rfi or "онтаж греющ" in description_rfi or \
                    "роверка  кабельн" in description_rfi or "онтаж кабельн" in description_rfi or \
                    "роверка греющ" in description_rfi or "роверка кабел" in description_rfi or \
                    "онтаж  кабельн" in description_rfi:
                if 'Принято' in category_cancelled:
                    for eht in ehts:
                        if eht[0] in description_rfi:
                            if "Принято" == category_cancelled:
                                isos[eht[1]][9] = rfi_number

                                isos_db[eht[1]][10] = rfi_number
                                isos_db[eht[1]][11] = "***"

                            if "Принято с замечаниями" == category_cancelled:
                                isos[eht[1]][9] = rfi_number + " ПЗ"
                                isos_db[eht[1]][10] = rfi_number + " ПЗ"
                                if comment_sk != "None":
                                    isos_db[eht[1]][11] = conclusion_sk + " //// " + comment_sk
                                else:
                                    isos_db[eht[1]][11] = conclusion_sk

            # ИНСПЕКЦИИ НА АКЗ ТЕПЛОСПУТНИКА
            # Поиск по подающему манифольду, привязка к изометрии
            # "приёмка АКЗ теплоспутник" - описание для поиска АКЗ спутников

            if "АКЗ теплоспутник" in description_rfi:
                if 'Принято' in category_cancelled:
                    for ht in hts:
                        if ht[0] in description_rfi:
                            if "Принято" == category_cancelled:
                                isos_db[ht[1]][12] = rfi_number
                                isos_db[ht[1]][13] = "***"
                            if "Принято с замечаниями" == category_cancelled:
                                isos_db[ht[1]][12] = rfi_number + " ПЗ"
                                if comment_sk != "None":
                                    isos_db[ht[1]][13] = conclusion_sk + " //// " + comment_sk
                                else:
                                    isos_db[ht[1]][13] = conclusion_sk

            # ИЗОЛЯЦИЯ ТРУБОПРОВОДА
            if "еплоизоляц" in description_rfi:
                ins_without_ss = "n/d"
                ins_onli_ss = "n/d"
                ins_with_ss = "n/d"
                ins_box = "n/d"
                final_ins = "n/d"

                if "свидетельствование теплоизоляции технологических трубопровод" in description_rfi and \
                        "цинкован" in description_rfi and "за исключ" in description_rfi:
                    ins_without_ss = rfi_number

                if "плоизоляции участков технологических трубопроводов в местах" in description_rfi:
                    ins_onli_ss = rfi_number

                if "свидетельствование теплоизоляции технологических трубопровод" in description_rfi and \
                        "цинкован" in description_rfi and "за исключ" not in description_rfi:
                    ins_with_ss = rfi_number
                if "теплоизоляции запорной арматуры и фланцевых соединений" in description_rfi:
                    ins_box = rfi_number
                if "инальная приёмка теплоизоляции технологич" in description_rfi:
                    final_ins = rfi_number

                if 'Принято' in category_cancelled:
                    for key in isos:
                        iso = key.split("$")[0]
                        if iso in description_rfi or iso in list_iso:
                            if "Принято" == category_cancelled:
                                if ins_without_ss != "n/d":
                                    isos_db[key][14] = ins_without_ss
                                    isos_db[key][15] = "***"
                                if ins_onli_ss != "n/d":
                                    isos_db[key][16] = ins_onli_ss
                                    isos_db[key][17] = "***"
                                if ins_with_ss != "n/d":
                                    isos_db[key][18] = ins_with_ss
                                    isos_db[key][19] = "***"

                                if ins_box != "n/d":
                                    isos[key][12] = ins_box
                                    isos_db[key][21] = ins_box
                                    isos_db[key][22] = "***"
                                if final_ins != "n/d":
                                    isos[key][13] = final_ins
                                    isos_db[key][23] = final_ins
                                    isos_db[key][24] = "***"

                            if "Принято с замечаниями" == category_cancelled:
                                if ins_without_ss != "n/d":
                                    isos_db[key][14] = ins_without_ss + " ПЗ"
                                    if comment_sk != "None":
                                        isos_db[key][15] = conclusion_sk + " //// " + comment_sk
                                    else:
                                        isos_db[key][15] = conclusion_sk
                                if ins_onli_ss != "n/d":
                                    isos_db[key][16] = ins_onli_ss + " ПЗ"
                                    if comment_sk != "None":
                                        isos_db[key][17] = conclusion_sk + " //// " + comment_sk
                                    else:
                                        isos_db[key][17] = conclusion_sk
                                if ins_with_ss != "n/d":
                                    isos_db[key][18] = ins_with_ss + " ПЗ"
                                    if comment_sk != "None":
                                        isos_db[key][19] = conclusion_sk + " //// " + comment_sk
                                    else:
                                        isos_db[key][19] = conclusion_sk

                                if ins_box != "n/d":
                                    isos[key][12] = ins_box + " ПЗ"
                                    isos_db[key][21] = ins_box + " ПЗ"
                                    if comment_sk != "None":
                                        isos_db[key][22] = conclusion_sk + " //// " + comment_sk
                                    else:
                                        isos_db[key][22] = conclusion_sk
                                if final_ins != "n/d":
                                    isos[key][13] = final_ins + " ПЗ"
                                    isos_db[key][23] = ins_box + " ПЗ"
                                    if comment_sk != "None":
                                        isos_db[key][24] = conclusion_sk + " //// " + comment_sk
                                    else:
                                        isos_db[key][24] = conclusion_sk

                            # print(f"{iso} without SS - {ins_without_ss}, only SS - {ins_onli_ss}, "
                            #       f"with SS - {ins_with_ss}, BOX - {ins_box}, final - {final_ins}")

    # Создание БД для накопительной
    summary_accum_db_rfi = []
    summary_accum_db_tps_rfi = []

    for tp in tps:
        summary_accum_db_tps_rfi.append([tp, tps[tp][1], tps[tp][2], tps[tp][3], tps[tp][4], tps[tp][5], tps[tp][6],
                                         tps[tp][7], tps[tp][8]])

    for iso in isos:
        isos[iso][3] = tps[isos[iso][1]][1]
        isos[iso][4] = tps[isos[iso][1]][3]
        isos[iso][5] = tps[isos[iso][1]][7]

        isos_db[iso][3] = tps[isos_db[iso][1]][3]
        isos_db[iso][4] = tps[isos_db[iso][1]][4]
        isos_db[iso][7] = tps[isos_db[iso][1]][7]
        isos_db[iso][8] = tps[isos_db[iso][1]][8]

        summary_accum_db_rfi.append([iso, isos[iso][0], isos[iso][1], isos[iso][2], isos[iso][3], isos[iso][4],
                                     isos[iso][5], isos[iso][6], isos[iso][7], isos[iso][8], isos[iso][9],
                                     isos[iso][10], isos[iso][11], isos[iso][12], isos[iso][13]])

    with open(f"{os.getcwd()}\\dbs_ti_p3\\summary_accum_db_rfi.csv", "w", newline="") as write_file:
        writing_file = csv.writer(write_file, delimiter=";")
        writing_file.writerows(summary_accum_db_rfi)
    print("summary_accum_db_rfi.csv writing succsessful!")

    with open(f"{os.getcwd()}\\dbs_ti_p3\\summary_accum_db_tps_rfi.csv", "w", newline="") as write_file:
        writing_file = csv.writer(write_file, delimiter=";")
        writing_file.writerows(summary_accum_db_tps_rfi)
    print("summary_accum_db_tps_rfi.csv writing succsessful!")

    # СОЗДАНИE БД ПОД ПРОВЕРКУ

    iso_summary_ins_write_db = [["Изометрия", "Тестпакет", "RFI TEST", "З + К", "RFI АКЗ", "З + К", "RFI OC", "З + К",
                                 "EHT-TAG", "RFI EHT", "З + К", "АКЗ теплосп.", "З + К",
                                 "RFI INS TT", "З + К", "RFI INS СС", "З + К", "RFI INS TT+CC", "З + К",
                                 "Кол-во ЗРА", "RFI INS ЗРА", "З + К", "RFI INS FINAL", "З + К"]]
    for key in isos_db:
        iso = key.split("$")[0]
        tp = key.split("$")[1]
        phase = isos_db[key][0]

        rfi_test = isos_db[key][3]
        test_com = isos_db[key][4]

        rfi_akz = isos_db[key][5]
        akz_com = isos_db[key][6]

        rfi_reinst = isos_db[key][7]
        reinst_com = isos_db[key][8]

        eht_tag = isos_db[key][9]
        rfi_eht = isos_db[key][10]
        eht_com = isos_db[key][11]

        rfi_akz_ht = isos_db[key][12]
        akz_ht_com = isos_db[key][13]

        rfi_ins_tt = isos_db[key][14]
        ins_tt_com = isos_db[key][15]

        rfi_ins_ss = isos_db[key][16]
        ins_ss_com = isos_db[key][17]

        rfi_ins_ttss = isos_db[key][18]
        ins_ttss_com = isos_db[key][19]

        count_zra = isos_db[key][20]

        rfi_ins_zra = isos_db[key][21]
        ins_zra_com = isos_db[key][22]

        rfi_ins_final = isos_db[key][23]
        ins_final_com = isos_db[key][24]

        iso_summary_ins_write_db.append([iso, tp, rfi_test, test_com, rfi_akz, akz_com, rfi_reinst, reinst_com,
                                         eht_tag, rfi_eht, eht_com, rfi_akz_ht, akz_ht_com,
                                         rfi_ins_tt, ins_tt_com, rfi_ins_ss, ins_ss_com, rfi_ins_ttss, ins_ttss_com,
                                         count_zra, rfi_ins_zra, ins_zra_com, rfi_ins_final, ins_final_com])

    with open(f"{os.getcwd()}\\dbs_ti_p3\\iso_ins_db_rfi.csv", "w", newline="") as write_file:
        writing_file = csv.writer(write_file, delimiter=";")
        writing_file.writerows(iso_summary_ins_write_db)
    print("iso_ins_db_rfi.csv writing succsessful!")

    return isos


# -------------Запись базы Р3
def create_summary_p3(path: str):
    summary_iso_tp_phase_1 = [["Фаза", "Тестпакет", "Изометрия", "Длина ТТ, м.", "RFI монтаж", "RFI TEST", "RFI OC",
                               "Площадь АКЗ, м2", "RFI АКЗ", "Длина греющ. каб., м.", "RFI греющ. каб.",
                               "Объём ТИ, м2", "Кол-во ЗРА", "RFI ТИ ЗРА/фланц", "RFI ФИНАЛ ТИ"]]
    summary_iso_tp_phase_2 = [["Фаза", "Тестпакет", "Изометрия", "Длина ТТ, м.", "RFI монтаж", "RFI TEST", "RFI OC",
                               "Площадь АКЗ, м2", "RFI АКЗ", "Длина греющ. каб., м.", "RFI греющ. каб.",
                               "Объём ТИ, м2", "Кол-во ЗРА", "RFI ТИ ЗРА/фланц", "RFI ФИНАЛ ТИ"]]
    summary_iso_tp_phase_3 = [["Фаза", "Тестпакет", "Изометрия", "Длина ТТ, м.", "RFI монтаж", "RFI TEST", "RFI OC",
                               "Площадь АКЗ, м2", "RFI АКЗ", "Длина греющ. каб., м.", "RFI греющ. каб.",
                               "Объём ТИ, м2", "Кол-во ЗРА", "RFI ТИ ЗРА/фланц", "RFI ФИНАЛ ТИ"]]
    summary_iso_tp_phase_4 = [["Фаза", "Тестпакет", "Изометрия", "Длина ТТ, м.", "RFI монтаж", "RFI TEST", "RFI OC",
                               "Площадь АКЗ, м2", "RFI АКЗ", "Длина греющ. каб., м.", "RFI греющ. каб.",
                               "Объём ТИ, м2", "Кол-во ЗРА", "RFI ТИ ЗРА/фланц", "RFI ФИНАЛ ТИ"]]
    summary_iso_tp_phase_5 = [["Фаза", "Тестпакет", "Изометрия", "Длина ТТ, м.", "RFI монтаж", "RFI TEST", "RFI OC",
                               "Площадь АКЗ, м2", "RFI АКЗ", "Длина греющ. каб., м.", "RFI греющ. каб.",
                               "Объём ТИ, м2", "Кол-во ЗРА", "RFI ТИ ЗРА/фланц", "RFI ФИНАЛ ТИ"]]

    isolist = read_rfi_journal()

    for key in isolist:
        iso = key.split("$")[0]
        tp = key.split("$")[1]

        phase = isolist[key][0]
        length_tt = isolist[key][2]
        rfi_erection = isolist[key][3]
        rfi_test = isolist[key][4]
        rfi_reinst = isolist[key][5]
        value_akz = isolist[key][6]
        rfi_akz = isolist[key][7]
        length_eht = isolist[key][8]
        rfi_eht = isolist[key][9]
        value_ins = isolist[key][10]
        count_zra = isolist[key][11]
        rfi_ins_zra = isolist[key][12]
        rfi_ins_tt_final = isolist[key][13]

        if phase == "Phase 1":
            summary_iso_tp_phase_1.append([phase, tp, iso, length_tt, rfi_erection, rfi_test, rfi_reinst,
                                           value_akz, rfi_akz, length_eht, rfi_eht, value_ins, count_zra,
                                           rfi_ins_zra, rfi_ins_tt_final])
        if phase == "Phase 2":
            summary_iso_tp_phase_2.append([phase, tp, iso, length_tt, rfi_erection, rfi_test, rfi_reinst,
                                           value_akz, rfi_akz, length_eht, rfi_eht, value_ins, count_zra,
                                           rfi_ins_zra, rfi_ins_tt_final])
        if phase == "Phase 3":
            summary_iso_tp_phase_3.append([phase, tp, iso, length_tt, rfi_erection, rfi_test, rfi_reinst,
                                           value_akz, rfi_akz, length_eht, rfi_eht, value_ins, count_zra,
                                           rfi_ins_zra, rfi_ins_tt_final])
        if phase == "Phase 4":
            summary_iso_tp_phase_4.append([phase, tp, iso, length_tt, rfi_erection, rfi_test, rfi_reinst,
                                           value_akz, rfi_akz, length_eht, rfi_eht, value_ins, count_zra,
                                           rfi_ins_zra, rfi_ins_tt_final])
        if phase == "Phase 5":
            summary_iso_tp_phase_5.append([phase, tp, iso, length_tt, rfi_erection, rfi_test, rfi_reinst,
                                           value_akz, rfi_akz, length_eht, rfi_eht, value_ins, count_zra,
                                           rfi_ins_zra, rfi_ins_tt_final])

    workbook_summary = xlsxwriter.Workbook(
        f'{path}\\Сводка Р3 по ФАЗАМ на {datetime.datetime.now().strftime("%d.%m.%Y")}.xlsx')

    # -------------Фаза 1
    ws1 = workbook_summary.add_worksheet('Phase 1')
    ws1.set_column(0, 0, 9)
    ws1.set_column(1, 2, 35)
    ws1.set_column(3, 16, 14)

    ws1.autofilter(f'A1:S{len(summary_iso_tp_phase_1)}')

    cell_format_green = workbook_summary.add_format()
    cell_format_green.set_bg_color('#99FF99')
    cell_format_blue = workbook_summary.add_format()
    cell_format_blue.set_bg_color('#99CCCC')
    cell_format_hat = workbook_summary.add_format()
    cell_format_hat.set_bg_color('#FF9966')
    cell_format_date = workbook_summary.add_format()
    cell_format_date.set_font_size(font_size=14)

    cell_format_inf = workbook_summary.add_format()
    cell_format_inf.set_bg_color("72d7ed")

    for i, (one, two, three, four, five, six, seven, eight, nine, ten, elev, twelve,
            thirteen, fourteen, fiveteen) in enumerate(summary_iso_tp_phase_1, start=1):

        if two == 'Тестпакет':
            color = cell_format_hat
            color_inf = cell_format_hat
            color.set_bold('bold')
        # elif '-CC' in rfi_reinstatement:
        #     color = cell_format_green
        else:
            color = cell_format_blue
            color_inf = cell_format_inf

        try:
            color.set_border(style=1)
            color_inf.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
            color_inf.set_text_wrap(text_wrap=1)

        except:
            pass
        ws1.write(f'A{i}', one, color_inf)
        ws1.write(f'B{i}', two, color_inf)
        ws1.write(f'C{i}', three, color_inf)
        ws1.write(f'D{i}', four, color_inf)
        ws1.write(f'E{i}', five, color)
        ws1.write(f'F{i}', six, color)
        ws1.write(f'G{i}', seven, color)
        ws1.write(f'H{i}', eight, color_inf)
        ws1.write(f'I{i}', nine, color)
        ws1.write(f'J{i}', ten, color_inf)
        ws1.write(f'K{i}', elev, color)
        ws1.write(f'L{i}', twelve, color_inf)
        ws1.write(f'M{i}', thirteen, color_inf)
        ws1.write(f'N{i}', fourteen, color)
        ws1.write(f'O{i}', fiveteen, color)

    # -------------Фаза 2
    ws2 = workbook_summary.add_worksheet('Phase 2')
    ws2.set_column(0, 0, 9)
    ws2.set_column(1, 2, 35)
    ws2.set_column(3, 16, 14)

    ws2.autofilter(f'A1:S{len(summary_iso_tp_phase_2)}')

    for i, (one, two, three, four, five, six, seven, eight, nine, ten, elev, twelve,
            thirteen, fourteen, fiveteen) in enumerate(summary_iso_tp_phase_2, start=1):

        if two == 'Тестпакет':
            color = cell_format_hat
            color_inf = cell_format_hat
            color.set_bold('bold')
        # elif '-CC' in rfi_reinstatement:
        #     color = cell_format_green
        else:
            color = cell_format_blue
            color_inf = cell_format_inf

        try:
            color.set_border(style=1)
            color_inf.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
            color_inf.set_text_wrap(text_wrap=1)
        except:
            pass
        ws2.write(f'A{i}', one, color_inf)
        ws2.write(f'B{i}', two, color_inf)
        ws2.write(f'C{i}', three, color_inf)
        ws2.write(f'D{i}', four, color_inf)
        ws2.write(f'E{i}', five, color)
        ws2.write(f'F{i}', six, color)
        ws2.write(f'G{i}', seven, color_inf)
        ws2.write(f'H{i}', eight, color)
        ws2.write(f'I{i}', nine, color)
        ws2.write(f'J{i}', ten, color_inf)
        ws2.write(f'K{i}', elev, color)
        ws2.write(f'L{i}', twelve, color_inf)
        ws2.write(f'M{i}', thirteen, color_inf)
        ws2.write(f'N{i}', fourteen, color)
        ws2.write(f'O{i}', fiveteen, color)

    # -------------Фаза 3
    ws3 = workbook_summary.add_worksheet('Phase 3')
    ws3.set_column(0, 0, 9)
    ws3.set_column(1, 2, 35)
    ws3.set_column(3, 16, 14)

    ws3.autofilter(f'A1:S{len(summary_iso_tp_phase_3)}')

    for i, (one, two, three, four, five, six, seven, eight, nine, ten, elev, twelve,
            thirteen, fourteen, fiveteen) in enumerate(summary_iso_tp_phase_3, start=1):

        if two == 'Тестпакет':
            color = cell_format_hat
            color_inf = cell_format_hat
            color.set_bold('bold')
        # elif '-CC' in rfi_reinstatement:
        #     color = cell_format_green
        else:
            color = cell_format_blue
            color_inf = cell_format_inf

        try:
            color.set_border(style=1)
            color_inf.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
            color_inf.set_text_wrap(text_wrap=1)
        except:
            pass
        ws3.write(f'A{i}', one, color_inf)
        ws3.write(f'B{i}', two, color_inf)
        ws3.write(f'C{i}', three, color_inf)
        ws3.write(f'D{i}', four, color_inf)
        ws3.write(f'E{i}', five, color)
        ws3.write(f'F{i}', six, color)
        ws3.write(f'G{i}', seven, color)
        ws3.write(f'H{i}', eight, color_inf)
        ws3.write(f'I{i}', nine, color)
        ws3.write(f'J{i}', ten, color_inf)
        ws3.write(f'K{i}', elev, color)
        ws3.write(f'L{i}', twelve, color_inf)
        ws3.write(f'M{i}', thirteen, color_inf)
        ws3.write(f'N{i}', fourteen, color)
        ws3.write(f'O{i}', fiveteen, color)

    # -------------Фаза 4
    ws4 = workbook_summary.add_worksheet('Phase 4')
    ws4.set_column(0, 0, 9)
    ws4.set_column(1, 2, 35)
    ws4.set_column(3, 16, 14)

    ws4.autofilter(f'A1:S{len(summary_iso_tp_phase_4)}')

    for i, (one, two, three, four, five, six, seven, eight, nine, ten, elev, twelve,
            thirteen, fourteen, fiveteen) in enumerate(summary_iso_tp_phase_4, start=1):

        if two == 'Тестпакет':
            color = cell_format_hat
            color_inf = cell_format_hat
            color.set_bold('bold')
        # elif '-CC' in rfi_reinstatement:
        #     color = cell_format_green
        else:
            color = cell_format_blue
            color_inf = cell_format_inf

        try:
            color.set_border(style=1)
            color_inf.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
            color_inf.set_text_wrap(text_wrap=1)
        except:
            pass
        ws4.write(f'A{i}', one, color_inf)
        ws4.write(f'B{i}', two, color_inf)
        ws4.write(f'C{i}', three, color_inf)
        ws4.write(f'D{i}', four, color_inf)
        ws4.write(f'E{i}', five, color)
        ws4.write(f'F{i}', six, color)
        ws4.write(f'G{i}', seven, color)
        ws4.write(f'H{i}', eight, color_inf)
        ws4.write(f'I{i}', nine, color)
        ws4.write(f'J{i}', ten, color_inf)
        ws4.write(f'K{i}', elev, color)
        ws4.write(f'L{i}', twelve, color_inf)
        ws4.write(f'M{i}', thirteen, color_inf)
        ws4.write(f'N{i}', fourteen, color)
        ws4.write(f'O{i}', fiveteen, color)

    # -------------Фаза 5
    ws5 = workbook_summary.add_worksheet('Phase 5')
    ws5.set_column(0, 0, 9)
    ws5.set_column(1, 2, 35)
    ws5.set_column(3, 16, 14)

    ws5.autofilter(f'A1:S{len(summary_iso_tp_phase_5)}')

    for i, (one, two, three, four, five, six, seven, eight, nine, ten, elev, twelve,
            thirteen, fourteen, fiveteen) in enumerate(summary_iso_tp_phase_5, start=1):

        if two == 'Тестпакет':
            color = cell_format_hat
            color_inf = cell_format_hat
            color.set_bold('bold')
        # elif '-CC' in rfi_reinstatement:
        #     color = cell_format_green
        else:
            color = cell_format_blue
            color_inf = cell_format_inf

        try:
            color.set_border(style=1)
            color_inf.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
            color_inf.set_text_wrap(text_wrap=1)
        except:
            pass
        ws5.write(f'A{i}', one, color_inf)
        ws5.write(f'B{i}', two, color_inf)
        ws5.write(f'C{i}', three, color_inf)
        ws5.write(f'D{i}', four, color_inf)
        ws5.write(f'E{i}', five, color)
        ws5.write(f'F{i}', six, color)
        ws5.write(f'G{i}', seven, color)
        ws5.write(f'H{i}', eight, color_inf)
        ws5.write(f'I{i}', nine, color)
        ws5.write(f'J{i}', ten, color_inf)
        ws5.write(f'K{i}', elev, color)
        ws5.write(f'L{i}', twelve, color_inf)
        ws5.write(f'M{i}', thirteen, color_inf)
        ws5.write(f'N{i}', fourteen, color)
        ws5.write(f'O{i}', fiveteen, color)

    workbook_summary.close()
