import os
import csv
import math
import openpyxl as xl
import xlsxwriter
from datetime import datetime as dt





journal_folder = os.getcwd() + "\\out_files_for_dbs\\"
db_folder = os.getcwd() + "\\dbs\\"

journal_rt_sg = journal_folder + "Журнал P2.xlsx"
journal_ut_sg = journal_folder + "Журнал УЗК.xlsx"

wl_phase_23 = journal_folder + "WeldLog_Yamata_Phase 2-3.xlsx"
wl_phase_1 = journal_folder + "WeldLog_Yamata_Phase-1.xlsx"
wl_phase_4 = journal_folder + "WeldLog_Yamata_Phase 4.xlsx"

summary_WL_phase_1 = [['Линия', 'Процент контроля по проекту', 'Сварено стыков',
              'Необходим контроль ПО', 'Проконтролировано ПО',
              'Необходим НК СГ, ст.', 'Проведён НК СГ ст.', 'Статус % НК СГ']]

summary_WL_phase_2 = [['Линия', 'Тестпакет', 'Установка', 'Процент контроля по проекту', 'Сварено стыков',
              'Необходим контроль ПО', 'Проконтролировано ПО',
              'Необходим НК СГ, ст.', 'Проведён НК СГ ст.', 'Статус % НК СГ']]

summary_WL_phase_4 = [['Линия', 'Тестпакет', 'Установка', 'Процент контроля по проекту', 'Сварено стыков',
              'Необходим контроль ПО', 'Проконтролировано ПО',
              'Необходим НК СГ, ст.', 'Проведён НК СГ ст.', 'Статус % НК СГ']]

"""
Сбор инфы на линии
"""
def create_ll_dic():
    dic_inf_ll = {}

    with open(db_folder + "ll_all.csv", 'r') as r_file:
        readed_file = csv.reader(r_file, delimiter=";")
        for row in readed_file:
            dic_inf_ll[row[0].strip()] = [row[1].strip(), row[2].strip(), row[3].strip(),
                                          row[4].strip(), row[5].strip(), row[6].strip(), row[7].strip()]

    return dic_inf_ll



def check_sg_journals():
    # Журнал РК
    sg_control_dic = {}

    wb_rt_sg = xl.load_workbook(journal_rt_sg)
    sheet_sg = wb_rt_sg['лог']

    for k in sheet_sg['H2':'X240000']:
        if k[0].value:
            iso_number_sg = str(k[0].value).strip()
            joint_number_sg = str(k[5].value).strip().replace(',', '.')
            if 'R' in str(k[6].value):
                joint_number_sg = str(k[5].value).strip().replace(',', '.') + str(k[6].value).strip()
            date_rt_sg = dt.strptime(str(k[15].value), "%Y-%m-%d %H:%M:%S")

            sg_control_dic[iso_number_sg+joint_number_sg] = [iso_number_sg, joint_number_sg, date_rt_sg]

    print('сг RT закончил')
    wb_rt_sg.close()

    #Журнал УЗК------------------------------------------

    wb_ut_sg = xl.load_workbook(journal_ut_sg)
    sheet_ut_sg = wb_ut_sg['P2']

    for i in sheet_ut_sg['D3':'P300000']:
        if i[0].value:
            iso_number_ut = str(i[0].value).strip()
            joint_number_ut = str(i[4].value).strip().replace(',', '.')
            if 'R' in str(i[5].value):
                joint_number_ut = str(i[4].value).strip().replace(',', '.') + str(i[5].value).strip()
            date_ut_sg = dt.strptime(str(i[11].value), "%Y-%m-%d %H:%M:%S")

            sg_control_dic[iso_number_ut+joint_number_ut] = [iso_number_ut, joint_number_ut, date_ut_sg]

    print('сг UT закончил')
    wb_ut_sg.close()

    return sg_control_dic




# Check WeldLog'а Phase 1-----------------
def check_wl_p1(check_sg_control):
    wb_WL = xl.load_workbook(wl_phase_1)
    sheet = wb_WL['WELDLOG']

    line_check_p1 = {}

    for i in sheet['A3':'CJ100000']:
        if i[0].value:
            if not i[44].value:
                line_number = str(i[3].value).strip()
                line_check_p1[line_number] = [0, 0, 0]



    for i in sheet['A3':'CJ100000']:
        if i[0].value:
            if '4.1.4.60' in str(i[1].value) or '4.1.4.70' in str(i[1].value):
                if not i[44].value:
                    line_number = str(i[3].value).strip()
                    isometric = str(i[6].value).strip()
                    number_joint = str(i[12].value).strip() + \
                                   str(i[13].value).replace(',', '.').strip()

                    if i[14].value:
                        number_joint = str(i[12].value).strip() + \
                                       str(i[13].value).replace(',', '.').strip() + str(i[14].value).strip()

                    elem1 = str(i[55].value).strip()
                    elem2 = str(i[57].value).strip()


                    if 'BW' in str(i[15].value).strip() and 'FLANGE SW' not in elem1 and 'FLANGE SW' not in elem2:
                        line_check_p1[line_number][0] += 1

                        if i[30].value or i[33].value:
                            line_check_p1[line_number][1] += 1

                        if isometric + number_joint in check_sg_control.keys():
                            line_check_p1[line_number][2] += 1
    print('WL_1 закончил')
    return line_check_p1


# Check WeldLog'а Phase 2-----------------
def check_wl_p2(check_sg_control):
    wb_WL = xl.load_workbook(wl_phase_23)
    sheet = wb_WL['WELDLOG']

    line_check_p2 = {}

    for i in sheet['A5':'CK500000']:
        if i[0].value:
            if '4.2.4.60' in str(i[1].value) or '4.2.4.70' in str(i[1].value) or '4.3.3.11' in str(
                    i[1].value) or '4.3.3.30' in str(i[1].value) or '4.3.3.20' in str(i[1].value):
                if not i[44].value:
                    line_number = str(i[3].value).strip()
                    testpackage_number = str(i[87].value).strip()
                    line_check_p2[line_number + "/" + testpackage_number] = [0, 0, 0, '']

    for i in sheet['A5':'CK500000']:
        if i[0].value:
            if '4.2.4.60' in str(i[1].value) or '4.2.4.70' in str(i[1].value) or '4.3.3.11' in str(i[1].value) or '4.3.3.30' in str(i[1].value) or '4.3.3.20' in str(i[1].value):
                if not i[44].value:
                    line_number = str(i[3].value).strip()
                    isometric = str(i[6].value).strip()
                    number_joint = str(i[12].value).strip() + \
                                   str(i[13].value).replace(',', '.').strip()
                    testpackage_number = str(i[87].value).strip()

                    if i[14].value:
                        number_joint = str(i[12].value).strip() + \
                                       str(i[13].value).replace(',', '.').strip() + str(i[14].value).strip()

                    elem1 = str(i[55].value).strip()
                    elem2 = str(i[57].value).strip()

                    if 'BW' in str(i[15].value).strip() and 'FLANGE SW' not in elem1 and 'FLANGE SW' not in elem2:
                        line_check_p2[line_number + "/" + testpackage_number][0] += 1

                        if i[30].value or i[33].value:
                            line_check_p2[line_number + "/" + testpackage_number][1] += 1

                        if isometric + number_joint in check_sg_control.keys():
                            line_check_p2[line_number + "/" + testpackage_number][2] += 1

                        if '4.2.4.60' in str(i[1].value):
                            line_check_p2[line_number + "/" + testpackage_number][3] = '2-60'
                        if '4.2.4.70' in str(i[1].value):
                            line_check_p2[line_number + "/" + testpackage_number][3] = '2-70'
                        if '4.3.3.11' in str(i[1].value):
                            line_check_p2[line_number + "/" + testpackage_number][3] = '4-110'
                        if '4.3.3.30' in str(i[1].value):
                            line_check_p2[line_number + "/" + testpackage_number][3] = '4-30'
                        if '4.3.3.20' in str(i[1].value):
                            line_check_p2[line_number + "/" + testpackage_number][3] = '4-20'
            else:
                pass


    print('WL_2 закончил')
    return line_check_p2


def check_wl_p4(check_sg_control):
    wb_WL = xl.load_workbook(wl_phase_4)
    sheet = wb_WL['WELDLOG']

    line_check_p4 = {}

    for i in sheet['A5':'CK500000']:
        if i[0].value:
            if not i[45].value:
                line_number = str(i[3].value).strip()
                testpackage_number = str(i[88].value).strip()
                line_check_p4[line_number + "/" + testpackage_number] = [0, 0, 0, '']

    for i in sheet['A5':'CK500000']:
        if i[0].value:
            if not i[45].value:
                line_number = str(i[3].value).strip()
                isometric = str(i[6].value).strip()
                number_joint = str(i[13].value).strip() + \
                               str(i[14].value).replace(',', '.').strip()
                testpackage_number = str(i[88].value).strip()

                if i[15].value:
                    number_joint = str(i[13].value).strip() + \
                                   str(i[14].value).replace(',', '.').strip() + str(i[15].value).strip()

                elem1 = str(i[56].value).strip()
                elem2 = str(i[58].value).strip()

                if 'BW' in str(i[16].value).strip() and 'FLANGE SW' not in elem1 and 'FLANGE SW' not in elem2:
                    line_check_p4[line_number + "/" + testpackage_number][0] += 1

                    if i[31].value or i[34].value:
                        line_check_p4[line_number + "/" + testpackage_number][1] += 1

                    if isometric + number_joint in check_sg_control.keys():
                        line_check_p4[line_number + "/" + testpackage_number][2] += 1

                    if '4.4.3.11' in str(i[1].value):
                        line_check_p4[line_number + "/" + testpackage_number][3] = '5-100'
                    if '4.4.3.30' in str(i[1].value):
                        line_check_p4[line_number + "/" + testpackage_number][3] = '5-30'
                    if '4.4.4.60' in str(i[1].value):
                        line_check_p4[line_number + "/" + testpackage_number][3] = '3-60'
                    if '4.4.4.70' in str(i[1].value):
                        line_check_p4[line_number + "/" + testpackage_number][3] = '3-70'
                    if '4.5.3.11' in str(i[1].value):
                        line_check_p4[line_number + "/" + testpackage_number][3] = '6-110'
            else:
                pass


    print('WL_4 закончил')
    return line_check_p4


def create_summary_nkdk(path):
    check_sg_control = check_sg_journals()

    # lines_p1 = check_wl_p1(check_sg_control)
    tp_lines_p2 = check_wl_p2(check_sg_control)
    tp_lines_p4 = check_wl_p4(check_sg_control)

    # for i in lines_p1.keys():
    #     ll_dic = create_ll_dic()
    #     status_sg = 'Недобор'
    #
    #     try:
    #         line_control_percent = float(ll_dic[i][3])
    #         sg_control_percent = float(ll_dic[i][3]) * 0.1
    #
    #
    #         need_control_po = math.ceil(float(lines_p1[i][0]) * line_control_percent)
    #         need_control_sg = math.ceil(float(lines_p1[i][0]) * sg_control_percent)
    #
    #         summ_control_po = lines_p1[i][1]
    #         summ_control_sg = lines_p1[i][2]
    #
    #         if summ_control_sg >= need_control_sg:
    #             status_sg = 'OK'
    #
    #         summary_WL_phase_1.append([i, line_control_percent, lines_p1[i][0], need_control_po, summ_control_po,
    #                                    need_control_sg, summ_control_sg, status_sg])
    #
    #     except:
    #         print(f'Не нашел в лайн-листе линию  {i}')
    ll_dic = create_ll_dic()

    for i in tp_lines_p2.keys():

        status_sg = 'Недобор'

        line = i.split("/")[0].strip()
        testpackage_number = i.split("/")[1]
        unit = tp_lines_p2[i][3]

        try:
            line_control_percent = float(ll_dic[line][3])
            sg_control_percent = float(ll_dic[line][3]) * 0.1

            need_control_po = math.ceil(float(tp_lines_p2[i][0]) * line_control_percent)
            need_control_sg = math.ceil(float(tp_lines_p2[i][0]) * sg_control_percent)

            summ_control_po = tp_lines_p2[i][1]
            summ_control_sg = tp_lines_p2[i][2]

            if summ_control_sg >= need_control_sg:
                status_sg = 'OK'

            summary_WL_phase_2.append([line, testpackage_number, unit, line_control_percent, tp_lines_p2[i][0],
                                       need_control_po, summ_control_po,
                                       need_control_sg, summ_control_sg, status_sg])
        except:
            print(f'Не нашел в лайн-листе линию  {line}')

    for i in tp_lines_p4.keys():
        status_sg = 'Недобор'

        line = i.split("/")[0].strip()
        testpackage_number = i.split("/")[1].strip()
        unit = tp_lines_p4[i][3].strip()

        try:
            line_control_percent = float(ll_dic[line][3])
            sg_control_percent = float(ll_dic[line][3]) * 0.1

            need_control_po = math.ceil(float(tp_lines_p4[i][0]) * line_control_percent)
            need_control_sg = math.ceil(float(tp_lines_p4[i][0]) * sg_control_percent)

            summ_control_po = tp_lines_p4[i][1]
            summ_control_sg = tp_lines_p4[i][2]

            if summ_control_sg >= need_control_sg:
                status_sg = 'OK'

            summary_WL_phase_4.append([line, testpackage_number, unit, line_control_percent, tp_lines_p4[i][0],
                                       need_control_po, summ_control_po,
                                       need_control_sg, summ_control_sg, status_sg])
        except Exception as e:
            print(e)
            print(f'Не нашел в лайн-листе линию  {line}')


    #Запись в файл выходных данных
    workbook_wl = xlsxwriter.Workbook(f'{path}\Сводка по %ДК СГ на  {dt.now().strftime("%d.%m.%Y")}.xlsx')

    # ws = workbook_wl.add_worksheet('Phase_1')
    #
    # ws.set_column(0, 7, 13)
    # ws.autofilter(f'A1:I500000')
    #
    #
    cell_format_green = workbook_wl.add_format()
    cell_format_green.set_bg_color('#99FF99')
    cell_format_blue = workbook_wl.add_format()
    cell_format_blue.set_bg_color('#99CCCC')
    cell_format_hat = workbook_wl.add_format()
    cell_format_hat.set_bg_color('#FF9966')
    #
    # for i, (short_tp, number_line, number_iso, number_joint, percent_control, number_ut, res_ut,
    #         number_xr) in enumerate(summary_WL_phase_1, start=1):
    #
    #     if short_tp == 'Линия':
    #         color = cell_format_hat
    #         color.set_bold('bold')
    #
    #     else:
    #         try:
    #             if number_xr == 'OK':
    #                 color = cell_format_green
    #             else:
    #                 color = cell_format_blue
    #         except Exception as e:
    #             print(e)
    #
    #
    #     try:
    #         color.set_border(style=1)
    #         color.set_text_wrap(text_wrap=1)
    #     except:
    #         pass
    #
    #     ws.write(f'A{i}', short_tp, color)
    #     ws.write(f'B{i}', number_line, color)
    #     ws.write(f'C{i}', number_iso, color)
    #     ws.write(f'D{i}', number_joint, color)
    #     ws.write(f'E{i}', percent_control, color)
    #     ws.write(f'F{i}', number_ut, color)
    #     ws.write(f'G{i}', res_ut, color)
    #     ws.write(f'H{i}', number_xr, color)

    ws2 = workbook_wl.add_worksheet('Phase_2_3')

    ws2.set_column(0, 0, 13)
    ws2.set_column(1, 1, 30)
    ws2.set_column(2, 9, 13)

    ws2.autofilter(f'A1:J100000')

    cell_form = workbook_wl.add_format()
    cell_form.set_text_wrap(text_wrap=1)

    for i, (short_tp, number_line, number_iso, number_joint, percent_control, number_ut, res_ut,
            number_xr, nine, ten) in enumerate(summary_WL_phase_2, start=1):

        if short_tp == 'Линия':
            color = cell_format_hat
            color.set_bold('bold')

        else:
            try:
                if ten == 'OK':
                    color = cell_format_green
                else:
                    color = cell_format_blue
            except Exception as e:
                print(e)


        try:
            color.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
        except:
            pass

        ws2.write(f'A{i}', short_tp, color)
        ws2.write(f'B{i}', number_line, color)
        ws2.write(f'C{i}', number_iso, color)
        ws2.write(f'D{i}', number_joint, color)
        ws2.write(f'E{i}', percent_control, color)
        ws2.write(f'F{i}', number_ut, color)
        ws2.write(f'G{i}', res_ut, color)
        ws2.write(f'H{i}', number_xr, color)
        ws2.write(f'I{i}', nine, color)
        ws2.write(f'J{i}', ten, color)

    ws3 = workbook_wl.add_worksheet('Phase_4')

    ws3.set_column(0, 0, 13)
    ws3.set_column(1, 1, 30)
    ws3.set_column(2, 9, 13)

    ws3.autofilter(f'A1:J100000')

    cell_form = workbook_wl.add_format()
    cell_form.set_text_wrap(text_wrap=1)

    for i, (short_tp, number_line, number_iso, number_joint, percent_control, number_ut, res_ut,
            number_xr, nine, ten) in enumerate(summary_WL_phase_4, start=1):

        if short_tp == 'Линия':
            color = cell_format_hat
            color.set_bold('bold')

        else:
            try:
                if ten == 'OK':
                    color = cell_format_green
                else:
                    color = cell_format_blue
            except Exception as e:
                print(e)

        try:
            color.set_border(style=1)
            color.set_text_wrap(text_wrap=1)
        except:
            pass

        ws3.write(f'A{i}', short_tp, color)
        ws3.write(f'B{i}', number_line, color)
        ws3.write(f'C{i}', number_iso, color)
        ws3.write(f'D{i}', number_joint, color)
        ws3.write(f'E{i}', percent_control, color)
        ws3.write(f'F{i}', number_ut, color)
        ws3.write(f'G{i}', res_ut, color)
        ws3.write(f'H{i}', number_xr, color)
        ws3.write(f'I{i}', nine, color)
        ws3.write(f'J{i}', ten, color)


    workbook_wl.close()
    print('Всё записал.')


