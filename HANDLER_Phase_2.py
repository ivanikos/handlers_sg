# -*- coding: cp1251 -*-

import openpyxl as xl
import re
import xlsxwriter
from dateutil import parser
import datetime, time
import pandas as pd
from playsound import playsound


# Создание общего словаря ТП-------------------------------------------
wb_phase2 = xl.load_workbook('БД ТП ФАЗА 1, 2.xlsx')
sheet_phase2_TP = wb_phase2['Сводная информация']
sheet_phase2_ISO = wb_phase2['TP']

list_categories_ae = ['Ба-I/ Ba-I', 'Б(а)-I', 'A(б)-I / A(b)-I', 'Б(а)-I / B(a)-I', 'B-I / C-I', 'Б(б)-I / B(b)-I',
                      'В-I', 'Б(а)-I/B(a)-I', 'A(б)-II / A(b)-II', 'Ба-I/Ba-I', 'Ба-II/ Ba-II', 'Бб-I/ Bb-I',
                      'B-I/ V-I', 'B-I/V-I', 'A(б)-I/A(b)-I', 'Б(б)-I', 'Б(б)-I/B(b)-I', 'B-I/C-I', 'Ба-II/Ba-II',
                      'Бб-I/Bb-I']

testpackages = {}
testpackages_ae = {}
for l in sheet_phase2_TP['A5':'K5000']:
    if l[0].value:
        tp_short_code_BD = str(l[0].value).strip()
        title = str(l[1].value).strip()
        fluid = str(l[3].value).strip()
        category_tp = str(l[7].value)
        vt_tp = str(l[4].value)
        try:
            length = round(float(str(l[5].value)), 3)
        except:
            pass
        testpackages[tp_short_code_BD] = [title, fluid, category_tp, length]

        if category_tp in list_categories_ae and 'Визуальный' not in vt_tp:
            testpackages_ae[tp_short_code_BD] = 1
        else:
            testpackages_ae[tp_short_code_BD] = 0
    else:
        break

print(f'{len(testpackages.keys())} тестпакетов')

isotpdic = {}
sc_isotpdic = {}

for i in sheet_phase2_ISO['A3':'K12117']:
    if i[0].value:
        sc_iso_tp = str(i[1].value).strip()
        isometric = str(i[2].value).strip()
        testpack = str(i[0].value).strip()
        isolength_max = round(float(str(i[6].value)), 3)
        isolength = round(float(str(i[7].value)), 3)
        iso_type_ins = str(i[8].value).strip()
        iso_ins_volume = float(str(i[9].value))
        title_iso = str(i[3].value)
        isotpdic[isometric] = [testpack, isolength_max, '', '', '', '', iso_type_ins, iso_ins_volume, '', '', '', '',
                               '']
        sc_isotpdic[sc_iso_tp] = [isometric, testpack, isolength, '', '', '', '', title_iso]
    else:
        break

print(f'Общий словарь создан. \n Уникальных изометрий - {len(isotpdic.keys())} '
      f'\n переходящих изометрий - {len(sc_isotpdic.keys()) - len(isotpdic.keys())}')
# Общий словарь создан--------------------------------------------------


for key in testpackages.keys():
    testpackages[key].append('n/p')
    testpackages[key].append('n/p')
    testpackages[key].append('n/p')


# ------------------------------------------------------------
for key in testpackages.keys():
    testpackages[key].append('')
    testpackages[key].append(0)
    testpackages[key].append('')
    testpackages[key].append('')
    testpackages[key].append('')
    testpackages[key].append('')
# ------------------------------------------------------------------


# Проверка Журнал заявок АИС Р2 ФАЗА2-------------------------------------
df = pd.read_excel('Журнал заявок общий.xlsx')
df = df.sort_values(by='Дата подачи / Date of submission', ascending=True)
df.to_excel('Журнал заявок общий.xlsx', index=0)

wb_journal_rfi = xl.load_workbook('Журнал заявок общий.xlsx')
sheet = wb_journal_rfi['Sheet1']

replace_pattern_1 = ['-HT', '-VT', '-PT']
replace_pattern_2 = ['(T.T. REINSTATEMENT)', '(T.T. AIR BLOWING)', '(AIR BLOWING)', '(T.T AIR BLOWING',
                     '(T.T. ERECTION)', '(T.T.TEST)', '(T.T. TEST)',
                     '(T.T ERECTION)', '(T.T TEST)', '(T.T REINSTATEMENT)',
                     '(T.T AIR BLOWING)', '(GPA AIR BLOWING)', '(GPA TEST)',
                     '(GPA ERECTION)', '(GPA REINSTATEMENT)', '(T.T. REISTATEMENT)', '(T.T.REINSTATEMENT)',
                     '(T.T RE-INSTATEMENT)', '( T.T AIR BLOWING )', '( T.T AIR BLOWING )',
                     '(T.T.ERECTION)', '(T.T.TEST)', '(T.T.AIR BLOWING)', '(T.T.REINSTATEMENT)']
res_summary = {}
for i in sheet['B2':'AO550000']:
    if i[0].value:
        rfi_number = str(i[1].value)
        tp_number = str(i[2].value)
        pkk = str(i[4].value)
        tp_shortname = ''
        tp_shortname_1 = ''
        for l in replace_pattern_2:
            if l in tp_number:
                tp_shortname_1 = tp_number.replace(l, '').strip()
            else:
                pass
        if '-HT' in tp_shortname_1:
            tp_shortname = tp_shortname_1.replace('-HT', '')
        elif '-PT' in tp_shortname_1:
            tp_shortname = tp_shortname_1.replace('-PT', '')
        elif '-VT' in tp_shortname_1:
            tp_shortname = tp_shortname_1.replace('-VT', '')
        else:
            tp_shortname = tp_shortname_1

        description_rfi = str(i[16].value)
        violation = str(i[35].value)
        name_insp = str(i[26].value)
        list_iso = str(i[8].value)
        volume_meter = re.sub(r'[^0-9.]', '', str(i[18].value))
        category_cancelled = str(i[31].value)
        comment = str(i[39].value)


        if tp_shortname in testpackages.keys():
            if 'Принято' in category_cancelled:
                if 'братной сборки технологических трубопроводов ГПА' in description_rfi:
                    testpackages[tp_shortname][11] = rfi_number
                if 'онтаж технологического трубопровода ГПА' in description_rfi:
                    testpackages[tp_shortname][9] = rfi_number
                if 'испытаний технологического трубопровода ГПА' in description_rfi:
                    testpackages[tp_shortname][10] = rfi_number
                if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                    testpackages[tp_shortname][9] = rfi_number
                if 'испытаний на прочность и плотность' in description_rfi:
                    testpackages[tp_shortname][10] = rfi_number
                if 'испытаний технологического трубопровода на прочность' in description_rfi:
                    testpackages[tp_shortname][10] = rfi_number
                if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                    testpackages[tp_shortname][10] = rfi_number
                if 'сборки технологических трубопроводов в проект' in description_rfi:
                    testpackages[tp_shortname][11] = rfi_number
                if 'Продувка' in description_rfi:
                    testpackages[tp_shortname][12] = rfi_number
            else:
                if 'подтвержд' in comment:
                    if 'сборки технологических трубопроводов ГПА' in description_rfi:
                        testpackages[tp_shortname][11] = rfi_number + ' ФОП'
                    if 'онтаж технологического трубопровода ГПА' in description_rfi:
                        testpackages[tp_shortname][9] = rfi_number + ' ФОП'
                    if 'испытаний технологического трубопровода ГПА' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ФОП'
                    if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                        testpackages[tp_shortname][9] = rfi_number + ' ФОП'
                    if 'испытаний на прочность и плотность' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ФОП'
                    if 'испытаний технологического трубопровода на прочность' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number
                    if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ФОП'
                    if 'сборки технологических трубопроводов в проект' in description_rfi:
                        testpackages[tp_shortname][11] = rfi_number + ' ФОП'
                    if 'родувка' in description_rfi:
                        testpackages[tp_shortname][12] = rfi_number + ' ФОП'
                if 'зафиксирован' in comment:
                    if 'испытаний технологического трубопровода ГПА' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ФОП'
                    if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                        testpackages[tp_shortname][9] = rfi_number + ' ФОП'
                    if 'испытаний на прочность и плотность' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ФОП'
                    if 'испытаний технологического трубопровода на прочность' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number
                    if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ФОП'
                    if 'сборки технологических трубопроводов в проект' in description_rfi:
                        testpackages[tp_shortname][11] = rfi_number + ' ФОП'
                    if 'Продувка' in description_rfi:
                        testpackages[tp_shortname][12] = rfi_number + ' ФОП'

        for isom in list_iso.split(';'):
            if rfi_number == 'CPECC-CC-64685/1':
                isom = isom.replace(' р. 0', '').replace(' р. 1', '').replace(' р. 3', '')

            if isom.strip() in isotpdic.keys():
                if 'Принято' in category_cancelled:
                    if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                        isotpdic[isom.strip()][2] = rfi_number
                    if 'испытаний на прочность и плотность' in description_rfi:
                        isotpdic[isom.strip()][3] = rfi_number
                    if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                        isotpdic[isom.strip()][3] = rfi_number
                    if 'сборки технологических трубопроводов в проект' in description_rfi:
                        isotpdic[isom.strip()][5] = rfi_number
                    if 'родувка' in description_rfi and 'еплоспутн' not in description_rfi:
                        isotpdic[isom.strip()][4] = rfi_number
                    if 'дополнительных испытаний' in description_rfi:
                        isotpdic[isom.strip()][11] = rfi_number
                        isotpdic[isom.strip()][10] = 'Испытан'
                else:
                    if 'подтвержд' in comment or 'подтвржд' in comment:
                        if 'сборки технологических трубопроводов ГПА' in description_rfi:
                            isotpdic[isom.strip()][5] = rfi_number + ' ФОП'
                        if 'онтаж технологического трубопровода ГПА' in description_rfi:
                            isotpdic[isom.strip()][2] = rfi_number + ' ФОП'
                        if 'испытаний технологического трубопровода ГПА' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ФОП'
                        if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                            isotpdic[isom.strip()][2] = rfi_number + ' ФОП'
                        if 'испытаний на прочность и плотность' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ФОП'
                        if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ФОП'
                        if 'сборки технологических трубопроводов в проект' in description_rfi:
                            isotpdic[isom.strip()][5] = rfi_number + ' ФОП'
                        if 'родувка' in description_rfi and 'еплоспутн' not in description_rfi:
                            isotpdic[isom.strip()][4] = rfi_number + ' ФОП'
                        if 'покраск' in description_rfi:
                            isotpdic[isom.strip()][12] = rfi_number + ' ФОП'
                    if 'зафиксирован' in comment:
                        if 'испытаний технологического трубопровода ГПА' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ФОП'
                        if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                            isotpdic[isom.strip()][2] = rfi_number + ' ФОП'
                        if 'испытаний на прочность и плотность' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ФОП'
                        if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ФОП'
                        if 'сборки технологических трубопроводов в проект' in description_rfi:
                            isotpdic[isom.strip()][5] = rfi_number + ' ФОП'
                        if 'родувка' in description_rfi and 'еплоспутн' not in description_rfi:
                            isotpdic[isom.strip()][4] = rfi_number + ' ФОП'

            # Проверка изометрий сцепка с теспаком. Исключающее условие переходящих изометрий.
            if tp_shortname + isom.strip() in sc_isotpdic.keys():
                if 'Принято' in category_cancelled:
                    if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][3] = rfi_number
                    if 'испытаний на прочность и плотность' in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number
                    if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number
                    if 'сборки технологических трубопроводов в проект' in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][6] = rfi_number
                    if 'родувка' in description_rfi and 'еплоспутн' not in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][5] = rfi_number
                else:
                    if 'подтвержд' or 'подтвржд' in comment:
                        if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][3] = rfi_number + ' ФОП'
                        if 'испытаний на прочность и плотность' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number + ' ФОП'
                        if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number + ' ФОП'
                        if 'сборки технологических трубопроводов в проект' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][6] = rfi_number + ' ФОП'
                        if 'родувка' in description_rfi and 'еплоспутн' not in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][5] = rfi_number + ' ФОП'
                    if 'зафиксирован' in comment:
                        if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][3] = rfi_number + ' ФОП'
                        if 'испытаний на прочность и плотность' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number + ' ФОП'
                        if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number + ' ФОП'
                        if 'сборки технологических трубопроводов в проект' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][6] = rfi_number + ' ФОП'
                        if 'родувка' in description_rfi and 'еплоспутн' not in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][5] = rfi_number + ' ФОП'

        if '75530/2' or '75823' in rfi_number:
            violation = 'представлены не в полном объеме, представлены некорректные документы'

        # ИЗОЛЯЦИЯ ПРОВЕРКА------------------------------
        if 'завершении работ по теплоизоляц' in pkk:
            if 'представлены не в полном объеме, представлены некорректные документы' in violation:
                if 'металлического кожуха согласно изометрическим' in description_rfi:
                    for iso in list_iso.split(';'):
                        try:
                            isotpdic[iso.strip()][9] = rfi_number
                        except:
                            pass
                if 'теплоизоляционного покрытия согласно изометрическим' in description_rfi:
                    for iso in list_iso.split(';'):
                        try:
                            isotpdic[iso.strip()][8] = rfi_number
                        except:
                            pass
            if 'документы, подтверждающие качество работ' in violation:
                if 'подтвержд' in comment:
                    if 'металлического кожуха согласно изометрическим' in description_rfi:
                        for iso in list_iso.split(';'):
                            try:
                                isotpdic[iso.strip()][9] = rfi_number
                            except:
                                pass
                    if 'теплоизоляционного покрытия согласно изометрическим' in description_rfi:
                        for iso in list_iso.split(';'):
                            try:
                                isotpdic[iso.strip()][8] = rfi_number
                            except:
                                pass
    else:
        break

print('Журнал заявок проверен. Информация добавлена.')
wb_journal_rfi.close()
# ------------------------------------------------------------------------
# -Проверка на уведомления-------------------------
wb_ncr = xl.load_workbook('Реестр уведомлений.xlsx')
sheet_ncr = wb_ncr['Предписания (Instructions)']
iso_ncr = {}
iso_ncr_iso = {}
for i in sheet_ncr['B4':'V55000']:
    if i[0].value:
        number_ncr = str(i[0].value)
        mark_execution = str(i[16].value)
        notification_items = str(i[1].value)
        type_violation = str(i[5].value)
        content_remarks = str(i[6].value).replace(' ', '')
        content_remarks_iso = re.findall(
            r'\d-\d-\d-\d\d-\d\d\d-\w*-\d\w-\d\d\d\d-\d\d\d|\d-\d-\d-\d\d-\d\d\d-\w*-\d\d-\d\d\d\d-\d\d\d|\d-\d-\d-\d\d-\d\d\d-NHC3P\+-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC3\+-\d\d-\d\d\d\d-\d\d\d|\d-\d-\d-\d\d-\d\d\d-NHC4P\+-\d\d-\d\d\d\d-\d\d\d|\d-\d-\d-\d\d-\d\d\d-NHC5\+-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC4\+-\d\d-\d\d\d\d-\d\d\d',
            content_remarks.replace(' ', '').replace('\n', '').replace('Р', 'P').replace('С', 'C').strip())
        if 'Нет' in mark_execution:
            if content_remarks_iso:
                for l in content_remarks_iso:
                    try:
                        iso_ncr_iso[l] = number_ncr
                        iso_ncr[isotpdic[l][0]] = number_ncr
                    except:
                        pass
    else:
        break
for key in testpackages.keys():
    if key in iso_ncr.keys():
        testpackages[key].append(iso_ncr[key])
    else:
        testpackages[key].append('Уведомлений нет')
wb_ncr.close()
print('Реестр уведомлений проверен. Информация добавлена.')
# --------------------------------------------------------

n_dic_3_110 = {'HWBR': ['Вода котлового контура, обратная (Т21)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWBS': ['Вода котлового контура, прямая (Т11)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHSGA': ['Товарный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHSGAHP': ['Товарный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NODRA': ['Дренаж', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UABA': ['Барьерный воздух ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHGAL': ['Топливный газ НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWWW': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_3_110_a = {'HWBR': ['Вода котлового контура, обратная (Т21)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWBS': ['Вода котлового контура, прямая (Т11)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHSGA': ['Товарный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHSGAHP': ['Товарный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NODRA': ['Дренаж', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UABA': ['Барьерный воздух ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHGAL': ['Топливный газ НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWWW': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}
n_dic_3_110_b = {'HWBR': ['Вода котлового контура, обратная (Т21)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWBS': ['Вода котлового контура, прямая (Т11)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHSGA': ['Товарный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHSGAHP': ['Товарный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NODRA': ['Дренаж', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UABA': ['Барьерный воздух ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHGAL': ['Топливный газ НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWWW': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_3_30 = {'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSU': ['Поверхностная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOWWA': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNMP': ['Азот СД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_2_60 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC4P+': ['Бутановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3P+': ['Пропановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC5+': ['С5+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC4+': ['С4+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNMP': ['Азот СД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOWWA': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WMMI': ['Водометанольная смесь', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSU': ['Поверхностная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HYDV': ['Пары углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
              }

n_dic_2_70 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_3_20 = {'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

info_summary_table_phase2 = [['Тестпакет', 'Установка', 'Среда', 'Категория', 'Длина', 'Заключение по ИД',
                              'Требуется АЭ', 'Проведен АЭ',
                              'RFI ERECTION', 'RFI TEST', 'RFI AIR BLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений',
                              'Статус проверки И.Д.', 'Акт испытаний', 'Акт продувки', 'Ремонтов (НГ)']]


for key in testpackages.keys():
    status_akt_id = testpackages[key][6]
    status_akt_test = 'не подписан'
    status_akt_blowing = 'не подписан'

    if 'CPECC' in testpackages[key][9]:
        status_akt_id = 'n/p'
    if 'CPECC' in testpackages[key][12] and 'ФОП' not in testpackages[key][12]:
        status_akt_test = 'Акт подписан'
    if 'CPECC' in testpackages[key][12] and 'ФОП' not in testpackages[key][11]:
        status_akt_blowing = 'Акт подписан'
    repairs = 0

    need_ae = 'Не требуется'
    done_ae = 'Не проводился'
    if testpackages_ae[key] == 1:
        need_ae = 'Требуется'


    info_summary_table_phase2.append(
        [key, testpackages[key][0], testpackages[key][1], testpackages[key][2], testpackages[key][3],
         testpackages[key][6], need_ae, done_ae,
         testpackages[key][9], testpackages[key][10], testpackages[key][12], testpackages[key][11],
         testpackages[key][13],
         status_akt_id, status_akt_test, status_akt_blowing, repairs])

    if testpackages[key][0] == '3-110':
        n_dic_3_110[testpackages[key][1]][1] += testpackages[key][3]
        n_dic_3_110[testpackages[key][1]][2] += 1
        if testpackages[key][9]:
            n_dic_3_110[testpackages[key][1]][3] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][4] += 1
        if testpackages[key][10]:
            n_dic_3_110[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][6] += 1
        if not testpackages[key][10] and testpackages[key][12]:
            n_dic_3_110[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][6] += 1
        if testpackages[key][12]:
            n_dic_3_110[testpackages[key][1]][7] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][8] += 1
        if testpackages[key][11]:
            n_dic_3_110[testpackages[key][1]][9] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][10] += 1
        n_dic_3_110[testpackages[key][1]][11] = n_dic_3_110[testpackages[key][1]][1] - \
                                                n_dic_3_110[testpackages[key][1]][9]
        n_dic_3_110[testpackages[key][1]][12] = n_dic_3_110[testpackages[key][1]][2] - \
                                                n_dic_3_110[testpackages[key][1]][10]
        if testpackages_ae[key] == 1:
            n_dic_3_110[testpackages[key][1]][13] += 1

        if 'A' in key[20:23]:
            n_dic_3_110_a[testpackages[key][1]][1] += testpackages[key][3]
            n_dic_3_110_a[testpackages[key][1]][2] += 1
            if testpackages[key][9]:
                n_dic_3_110_a[testpackages[key][1]][3] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][4] += 1
            if testpackages[key][10]:
                n_dic_3_110_a[testpackages[key][1]][5] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][6] += 1
            if not testpackages[key][10] and testpackages[key][12]:
                n_dic_3_110_a[testpackages[key][1]][5] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][6] += 1
            if testpackages[key][12]:
                n_dic_3_110_a[testpackages[key][1]][7] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][8] += 1
            if testpackages[key][11]:
                n_dic_3_110_a[testpackages[key][1]][9] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][10] += 1
            n_dic_3_110_a[testpackages[key][1]][11] = n_dic_3_110_a[testpackages[key][1]][1] - \
                                                      n_dic_3_110_a[testpackages[key][1]][9]
            n_dic_3_110_a[testpackages[key][1]][12] = n_dic_3_110_a[testpackages[key][1]][2] - \
                                                      n_dic_3_110_a[testpackages[key][1]][10]
            if testpackages_ae[key] == 1:
                n_dic_3_110_a[testpackages[key][1]][13] += 1

        if 'B' in key[20:23]:
            n_dic_3_110_b[testpackages[key][1]][1] += testpackages[key][3]
            n_dic_3_110_b[testpackages[key][1]][2] += 1
            if testpackages[key][9]:
                n_dic_3_110_b[testpackages[key][1]][3] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][4] += 1
            if testpackages[key][10]:
                n_dic_3_110_b[testpackages[key][1]][5] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][6] += 1
            if not testpackages[key][10] and testpackages[key][12]:
                n_dic_3_110_b[testpackages[key][1]][5] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][6] += 1
            if testpackages[key][12]:
                n_dic_3_110_b[testpackages[key][1]][7] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][8] += 1
            if testpackages[key][11]:
                n_dic_3_110_b[testpackages[key][1]][9] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][10] += 1
            n_dic_3_110_b[testpackages[key][1]][11] = n_dic_3_110_b[testpackages[key][1]][1] - \
                                                      n_dic_3_110_b[testpackages[key][1]][9]
            n_dic_3_110_b[testpackages[key][1]][12] = n_dic_3_110_b[testpackages[key][1]][2] - \
                                                      n_dic_3_110_b[testpackages[key][1]][10]
            if testpackages_ae[key] == 1:
                n_dic_3_110_b[testpackages[key][1]][13] += 1

    if testpackages[key][0] == '3-30':
        n_dic_3_30[testpackages[key][1]][1] += testpackages[key][3]
        n_dic_3_30[testpackages[key][1]][2] += 1
        if testpackages[key][9]:
            n_dic_3_30[testpackages[key][1]][3] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][4] += 1
        if testpackages[key][10]:
            n_dic_3_30[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][6] += 1

        if not testpackages[key][10] and testpackages[key][12]:
            n_dic_3_30[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][6] += 1

        if testpackages[key][12]:
            n_dic_3_30[testpackages[key][1]][7] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][8] += 1
        if testpackages[key][11]:
            n_dic_3_30[testpackages[key][1]][9] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][10] += 1
        n_dic_3_30[testpackages[key][1]][11] = n_dic_3_30[testpackages[key][1]][1] - \
                                               n_dic_3_30[testpackages[key][1]][9]
        n_dic_3_30[testpackages[key][1]][12] = n_dic_3_30[testpackages[key][1]][2] - \
                                               n_dic_3_30[testpackages[key][1]][10]
        if testpackages_ae[key] == 1:
            n_dic_3_30[testpackages[key][1]][13] += 1

    if testpackages[key][0] == '2-60':
        n_dic_2_60[testpackages[key][1]][1] += testpackages[key][3]
        n_dic_2_60[testpackages[key][1]][2] += 1
        if testpackages[key][9]:
            n_dic_2_60[testpackages[key][1]][3] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][4] += 1
        if testpackages[key][10]:
            n_dic_2_60[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][6] += 1

        if not testpackages[key][10] and testpackages[key][12]:
            n_dic_2_60[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][6] += 1

        if testpackages[key][12]:
            n_dic_2_60[testpackages[key][1]][7] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][8] += 1
        if testpackages[key][11]:
            n_dic_2_60[testpackages[key][1]][9] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][10] += 1
        n_dic_2_60[testpackages[key][1]][11] = n_dic_2_60[testpackages[key][1]][1] - \
                                               n_dic_2_60[testpackages[key][1]][9]
        n_dic_2_60[testpackages[key][1]][12] = n_dic_2_60[testpackages[key][1]][2] - \
                                               n_dic_2_60[testpackages[key][1]][10]
        if testpackages_ae[key] == 1:
            n_dic_2_60[testpackages[key][1]][13] += 1

    if testpackages[key][0] == '2-70':
        n_dic_2_70[testpackages[key][1]][1] += testpackages[key][3]
        n_dic_2_70[testpackages[key][1]][2] += 1
        if testpackages[key][9]:
            n_dic_2_70[testpackages[key][1]][3] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][4] += 1
        if testpackages[key][10]:
            n_dic_2_70[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][6] += 1

        if not testpackages[key][10] and testpackages[key][12]:
            n_dic_2_70[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][6] += 1

        if testpackages[key][12]:
            n_dic_2_70[testpackages[key][1]][7] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][8] += 1
        if testpackages[key][11]:
            n_dic_2_70[testpackages[key][1]][9] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][10] += 1
        n_dic_2_70[testpackages[key][1]][11] = n_dic_2_70[testpackages[key][1]][1] - \
                                               n_dic_2_70[testpackages[key][1]][9]
        n_dic_2_70[testpackages[key][1]][12] = n_dic_2_70[testpackages[key][1]][2] - \
                                               n_dic_2_70[testpackages[key][1]][10]
        if testpackages_ae[key] == 1:
            n_dic_2_70[testpackages[key][1]][13] += 1


n_list_3_110_a = [['', f'Статус по ТП 3-110 A на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                   '', '', '', '', '',
                   '', '', '-', '', ''],
                  ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                   'Принят монтаж, ТП',
                   'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                   'Принята ОС, м.',
                   'Принята ОС, ТП', 'Остаток ОС, м.', 'Остаток ОС, ТП', 'Требуется АЭ', 'Проведен АЭ']]
ITOG_list_a = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_3_110_b = [['', f'Статус по ТП 3-110 B на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                   '', '', '', '', '',
                   '', '', '-', '', ''],
                  ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                   'Принят монтаж, ТП',
                   'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                   'Принята ОС, м.',
                   'Принята ОС, ТП', 'Остаток ОС, м.', 'Остаток ОС, ТП', 'Требуется АЭ', 'Проведен АЭ']]
ITOG_list_b = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_3_110 = [['', f'Статус по ТП 3-110 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                 '', '', '', '', '',
                 '', '', '-', '', ''],
                ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                 'Принят монтаж, ТП',
                 'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                 'Принята ОС, м.',
                 'Принята ОС, ТП', 'Остаток ОС, м.', 'Остаток ОС, ТП', 'Требуется АЭ', 'Проведен АЭ']]
ITOG_list = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_3_30 = [['', f'Статус по ТП 3-30 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                '', '', '', '', '',
                '', '', '-', '', ''],
               ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                'Принят монтаж, ТП',
                'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                'Принята ОС, м.',
                'Принята ОС, ТП', 'Остаток ОС, м.', 'Остаток ОС, ТП', 'Требуется АЭ', 'Проведен АЭ']]
ITOG_list_3_30 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_2_60 = [['', f'Статус по ТП 2-60 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                '', '', '', '', '',
                '', '', '-', '', ''],
               ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                'Принят монтаж, ТП',
                'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                'Принята ОС, м.',
                'Принята ОС, ТП', 'Остаток ОС, м.', 'Остаток ОС, ТП', 'Требуется АЭ', 'Проведен АЭ']]
ITOG_list_2_60 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_2_70 = [['', f'Статус по ТП 2-70 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                '', '', '', '', '',
                '', '', '-', '', ''],
               ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                'Принят монтаж, ТП',
                'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.', 'Принята продувка, ТП',
                'Принята ОС, м.',
                'Принята ОС, ТП', 'Остаток ОС, м.', 'Остаток ОС, ТП', 'Требуется АЭ', 'Проведен АЭ']]
ITOG_list_2_70 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

summary_list_units = []

for key in n_dic_3_110.keys():
    n_list_3_110.append([key, n_dic_3_110[key][0], n_dic_3_110[key][1], n_dic_3_110[key][2], n_dic_3_110[key][3],
                         n_dic_3_110[key][4], n_dic_3_110[key][5], n_dic_3_110[key][6], n_dic_3_110[key][7],
                         n_dic_3_110[key][8], n_dic_3_110[key][9], n_dic_3_110[key][10], n_dic_3_110[key][11],
                         n_dic_3_110[key][12], n_dic_3_110[key][13], n_dic_3_110[key][14]])
    ITOG_list[2] += n_dic_3_110[key][1]
    ITOG_list[3] += n_dic_3_110[key][2]
    ITOG_list[4] += n_dic_3_110[key][3]
    ITOG_list[5] += n_dic_3_110[key][4]
    ITOG_list[6] += n_dic_3_110[key][5]
    ITOG_list[7] += n_dic_3_110[key][6]
    ITOG_list[8] += n_dic_3_110[key][7]
    ITOG_list[9] += n_dic_3_110[key][8]
    ITOG_list[10] += n_dic_3_110[key][9]
    ITOG_list[11] += n_dic_3_110[key][10]
    ITOG_list[12] += n_dic_3_110[key][11]
    ITOG_list[13] += n_dic_3_110[key][12]
    ITOG_list[14] += n_dic_3_110[key][13]
    ITOG_list[15] += n_dic_3_110[key][14]
n_list_3_110.append(ITOG_list)
ost_list_3_110 = ['', 'Остаток:', '', '', ITOG_list[2] - ITOG_list[4], ITOG_list[3] - ITOG_list[5],
                  ITOG_list[2] - ITOG_list[6], ITOG_list[3] - ITOG_list[7],
                  ITOG_list[2] - ITOG_list[8], ITOG_list[3] - ITOG_list[9], ITOG_list[2] - ITOG_list[10],
                  ITOG_list[3] - ITOG_list[11], '', '', ITOG_list[14] - ITOG_list[15], '']

empty_str = [' * ', ' * ', ' * ', '  *  ', '  *  ', '  *  ', '  *  ', '  ', '  ', '  ', '  ', '', '', '-', '', '']



for key in n_dic_3_30.keys():
    n_list_3_30.append([key, n_dic_3_30[key][0], n_dic_3_30[key][1], n_dic_3_30[key][2], n_dic_3_30[key][3],
                        n_dic_3_30[key][4], n_dic_3_30[key][5], n_dic_3_30[key][6], n_dic_3_30[key][7],
                        n_dic_3_30[key][8], n_dic_3_30[key][9], n_dic_3_30[key][10], n_dic_3_30[key][11],
                        n_dic_3_30[key][12], n_dic_3_30[key][13], n_dic_3_30[key][14]])
    ITOG_list_3_30[2] += n_dic_3_30[key][1]
    ITOG_list_3_30[3] += n_dic_3_30[key][2]
    ITOG_list_3_30[4] += n_dic_3_30[key][3]
    ITOG_list_3_30[5] += n_dic_3_30[key][4]
    ITOG_list_3_30[6] += n_dic_3_30[key][5]
    ITOG_list_3_30[7] += n_dic_3_30[key][6]
    ITOG_list_3_30[8] += n_dic_3_30[key][7]
    ITOG_list_3_30[9] += n_dic_3_30[key][8]
    ITOG_list_3_30[10] += n_dic_3_30[key][9]
    ITOG_list_3_30[11] += n_dic_3_30[key][10]
    ITOG_list_3_30[12] += n_dic_3_30[key][11]
    ITOG_list_3_30[13] += n_dic_3_30[key][12]
    ITOG_list_3_30[14] += n_dic_3_30[key][13]
    ITOG_list_3_30[15] += n_dic_3_30[key][14]
n_list_3_30.append(ITOG_list_3_30)
ost_list_3_30 = ['', 'Остаток:', '', '', ITOG_list_3_30[2] - ITOG_list_3_30[4], ITOG_list_3_30[3] - ITOG_list_3_30[5],
                 ITOG_list_3_30[2] - ITOG_list_3_30[6], ITOG_list_3_30[3] - ITOG_list_3_30[7],
                 ITOG_list_3_30[2] - ITOG_list_3_30[8], ITOG_list_3_30[3] - ITOG_list_3_30[9],
                 ITOG_list_3_30[2] - ITOG_list_3_30[10],
                 ITOG_list_3_30[3] - ITOG_list_3_30[11], '', '', ITOG_list_3_30[14] - ITOG_list_3_30[15], '']
for i in n_list_3_30:
    summary_list_units.append(i)
summary_list_units.append(ost_list_3_30)
summary_list_units.append(empty_str)

for key in n_dic_2_60.keys():
    n_list_2_60.append([key, n_dic_2_60[key][0], n_dic_2_60[key][1], n_dic_2_60[key][2], n_dic_2_60[key][3],
                        n_dic_2_60[key][4], n_dic_2_60[key][5], n_dic_2_60[key][6], n_dic_2_60[key][7],
                        n_dic_2_60[key][8], n_dic_2_60[key][9], n_dic_2_60[key][10], n_dic_2_60[key][11],
                        n_dic_2_60[key][12], n_dic_2_60[key][13], n_dic_2_60[key][14]])
    ITOG_list_2_60[2] += n_dic_2_60[key][1]
    ITOG_list_2_60[3] += n_dic_2_60[key][2]
    ITOG_list_2_60[4] += n_dic_2_60[key][3]
    ITOG_list_2_60[5] += n_dic_2_60[key][4]
    ITOG_list_2_60[6] += n_dic_2_60[key][5]
    ITOG_list_2_60[7] += n_dic_2_60[key][6]
    ITOG_list_2_60[8] += n_dic_2_60[key][7]
    ITOG_list_2_60[9] += n_dic_2_60[key][8]
    ITOG_list_2_60[10] += n_dic_2_60[key][9]
    ITOG_list_2_60[11] += n_dic_2_60[key][10]
    ITOG_list_2_60[12] += n_dic_2_60[key][11]
    ITOG_list_2_60[13] += n_dic_2_60[key][12]
    ITOG_list_2_60[14] += n_dic_2_60[key][13]
    ITOG_list_2_60[15] += n_dic_2_60[key][14]
n_list_2_60.append(ITOG_list_2_60)
ost_list_2_60 = ['', 'Остаток:', '', '', ITOG_list_2_60[2] - ITOG_list_2_60[4], ITOG_list_2_60[3] - ITOG_list_2_60[5],
                 ITOG_list_2_60[2] - ITOG_list_2_60[6], ITOG_list_2_60[3] - ITOG_list_2_60[7],
                 ITOG_list_2_60[2] - ITOG_list_2_60[8], ITOG_list_2_60[3] - ITOG_list_2_60[9],
                 ITOG_list_2_60[2] - ITOG_list_2_60[10],
                 ITOG_list_2_60[3] - ITOG_list_2_60[11], '', '', ITOG_list_2_60[14] - ITOG_list_2_60[15], '']
for i in n_list_2_60:
    summary_list_units.append(i)

summary_list_units.append(ost_list_2_60)
summary_list_units.append(empty_str)

for key in n_dic_2_70.keys():
    n_list_2_70.append([key, n_dic_2_70[key][0], n_dic_2_70[key][1], n_dic_2_70[key][2], n_dic_2_70[key][3],
                        n_dic_2_70[key][4], n_dic_2_70[key][5], n_dic_2_70[key][6], n_dic_2_70[key][7],
                        n_dic_2_70[key][8], n_dic_2_70[key][9], n_dic_2_70[key][10], n_dic_2_70[key][11],
                        n_dic_2_70[key][12], n_dic_2_70[key][13], n_dic_2_70[key][14]])
    ITOG_list_2_70[2] += n_dic_2_70[key][1]
    ITOG_list_2_70[3] += n_dic_2_70[key][2]
    ITOG_list_2_70[4] += n_dic_2_70[key][3]
    ITOG_list_2_70[5] += n_dic_2_70[key][4]
    ITOG_list_2_70[6] += n_dic_2_70[key][5]
    ITOG_list_2_70[7] += n_dic_2_70[key][6]
    ITOG_list_2_70[8] += n_dic_2_70[key][7]
    ITOG_list_2_70[9] += n_dic_2_70[key][8]
    ITOG_list_2_70[10] += n_dic_2_70[key][9]
    ITOG_list_2_70[11] += n_dic_2_70[key][10]
    ITOG_list_2_70[12] += n_dic_2_70[key][11]
    ITOG_list_2_70[13] += n_dic_2_70[key][12]
    ITOG_list_2_70[14] += n_dic_2_70[key][13]
    ITOG_list_2_70[15] += n_dic_2_70[key][14]
n_list_2_70.append(ITOG_list_2_70)
ost_list_2_70 = ['', 'Остаток:', '', '', ITOG_list_2_70[2] - ITOG_list_2_70[4], ITOG_list_2_70[3] - ITOG_list_2_70[5],
                 ITOG_list_2_70[2] - ITOG_list_2_70[6], ITOG_list_2_70[3] - ITOG_list_2_70[7],
                 ITOG_list_2_70[2] - ITOG_list_2_70[8], ITOG_list_2_70[3] - ITOG_list_2_70[9],
                 ITOG_list_2_70[2] - ITOG_list_2_70[10],
                 ITOG_list_2_70[3] - ITOG_list_2_70[11], '', '', ITOG_list_2_70[14] - ITOG_list_2_70[15], '']
for i in n_list_2_70:
    summary_list_units.append(i)
summary_list_units.append(ost_list_2_70)
summary_list_units.append(empty_str)

for i in n_list_3_110:
    summary_list_units.append(i)
summary_list_units.append(ost_list_3_110)
summary_list_units.append(empty_str)

iso_summary_table = [
    ['Изометрия', 'Тестпакет', 'Длина', 'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT',
     'Статус И.Д. по Хронологии', 'Статус И.Д. по RFI', 'Акт испытаний', 'Акт продувки', 'Уведомления',
     'Статус ДИГ', 'RFI ДИГ', 'Статус АКЗ', 'Тип изоляции', 'Пр. объём изоляции', 'RFI вата', 'RFI металл']]
for key in isotpdic.keys():
    status_akt_id = 'n/p'
    status_akt_test = 'не подписан'
    status_akt_blowing = 'не подписан'

    if 'CPECC' in isotpdic[key][4] and 'ФОП' not in isotpdic[key][4]:
        status_akt_test = 'Акт подписан'
    if 'CPECC' in isotpdic[key][5] and 'ФОП' not in isotpdic[key][5]:
        status_akt_blowing = 'Акт подписан'
    status_iso_hd = 'n/p'


    status_ncr_iso = 'Нет уведомлений'
    if key in iso_ncr_iso.keys():
        status_ncr_iso = iso_ncr_iso[key]

    iso_summary_table.append(
        [key, isotpdic[key][0], isotpdic[key][1], isotpdic[key][2], isotpdic[key][3], isotpdic[key][4],
         isotpdic[key][5], status_iso_hd, status_akt_id, status_akt_test, status_akt_blowing,
         status_ncr_iso, isotpdic[key][10], isotpdic[key][11], isotpdic[key][12], isotpdic[key][6], isotpdic[key][7],
         isotpdic[key][8], isotpdic[key][9]])


# Список для листа повторяющихся изометрий
double_iso_summary_table = [
    ['Изометрия', 'Тестпакет', 'Длина', 'Установка', 'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING',
     'RFI REINSTATEMENT']]
for key in sc_isotpdic.keys():
    double_iso_summary_table.append([sc_isotpdic[key][0], sc_isotpdic[key][1], sc_isotpdic[key][2], sc_isotpdic[key][7],
                                     sc_isotpdic[key][3], sc_isotpdic[key][4], sc_isotpdic[key][5],
                                     sc_isotpdic[key][6]])

# -------------------------------------------


workbook_summary = xlsxwriter.Workbook(f'Сводка по ФАЗЕ 2 на {datetime.datetime.now().strftime("%d.%m.%Y")}.xlsx')

ws0 = workbook_summary.add_worksheet('Краткая сводка по установкам')
ws0.set_column(0, 0, 12)
ws0.set_column(1, 1, 40)
ws0.set_column(4, 15, 12)
ws0.set_column(2, 2, 12)
ws0.set_column(3, 3, 12)

cell_format_green = workbook_summary.add_format()
cell_format_green.set_bg_color('#98FB98')
cell_format_blue = workbook_summary.add_format()
cell_format_blue.set_bg_color('#B0E0E6')
cell_format_hat = workbook_summary.add_format()
cell_format_hat.set_bg_color('#FFDAB9')
cell_format_date = workbook_summary.add_format()
cell_format_date.set_font_size(font_size=14)
for i, (one, two, three, four, five, six, seven, eight, nine, ten, eleven, twelve, thirteen,
        fourteen, fiften, sixten,) in enumerate(summary_list_units, start=2):
    if fourteen == 0:
        color = cell_format_green
    elif fourteen == 'Остаток ОС, ТП':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
    elif fourteen == '-':
        color = cell_format_date
        color.set_bold('bold')
    elif two == 'Итого:':
        color = cell_format_hat
        color.set_bold('bold')
    elif two == 'Остаток:':
        color = cell_format_hat
        color.set_bold('bold')
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass
    ws0.write(f'A{i}', one, color)
    ws0.write(f'B{i}', two, color)
    ws0.write(f'C{i}', three, color)
    ws0.write(f'D{i}', four, color)
    ws0.write(f'E{i}', fiften, color)
    ws0.write(f'F{i}', sixten, color)
    ws0.write(f'G{i}', five, color)
    ws0.write(f'H{i}', six, color)
    ws0.write(f'I{i}', seven, color)
    ws0.write(f'J{i}', eight, color)
    ws0.write(f'K{i}', nine, color)
    ws0.write(f'L{i}', ten, color)
    ws0.write(f'M{i}', eleven, color)
    ws0.write(f'N{i}', twelve, color)
    ws0.write(f'O{i}', thirteen, color)
    ws0.write(f'P{i}', fourteen, color)

# ws0.set_column('O:P', None, None, {'hidden': True})



# ------------------------------------------------------------
ws5 = workbook_summary.add_worksheet('Сводная информация по ТП')
ws5.set_column(0, 0, 30)
ws5.set_column(1, 5, 15)
ws5.set_column(6, 11, 22)
ws5.set_column(12, 17, 25)
ws5.set_column(18, 18, 13)
ws5.autofilter('A1:S1682')
for i, (testpack, ustan, flud, metr_ng, stat_id_1, inst_rfi, elev, twelw, thirt,
        fourteen, fifth, akt_id, akt_test, akt_blow, rep, aee, aeeedone) in enumerate(info_summary_table_phase2,
                                                                                      start=1):
    if testpack == 'Тестпакет':
        color = cell_format_hat
        color.set_bold('bold')
    elif akt_id:
        color = cell_format_green
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass

    ws5.write(f'A{i}', testpack, color)
    ws5.write(f'B{i}', ustan, color)
    ws5.write(f'C{i}', flud, color)
    ws5.write(f'D{i}', metr_ng, color)
    ws5.write(f'E{i}', stat_id_1, color)
    # ws5.write(f'F{i}', inst_rfi, color)
    ws5.write(f'F{i}', elev, color)
    ws5.write(f'G{i}', twelw, color)
    ws5.write(f'H{i}', thirt, color)
    ws5.write(f'I{i}', fourteen, color)
    ws5.write(f'J{i}', fifth, color)
    ws5.write(f'K{i}', akt_id, color)
    ws5.write(f'L{i}', akt_test, color)
    # ws5.write(f'N{i}', akt_blow, color)
    ws5.write(f'M{i}', rep, color)
    ws5.write(f'N{i}', aee, color)

# Сводка по изометриям

ws5 = workbook_summary.add_worksheet('Сводка по изометриям')
ws5.set_column(0, 0, 32)
ws5.set_column(1, 1, 28)
ws5.set_column(2, 2, 12)
ws5.set_column(3, 6, 22)
ws5.set_column(7, 9, 20)
ws5.set_column(10, 10, 13.8)
ws5.set_column(11, 11, 20)
ws5.set_column(12, 12, 20)
ws5.set_column(13, 14, 10.5)
ws5.set_column(15, 16, 20)

cell_format_ins = workbook_summary.add_format()
cell_format_ins.set_bg_color('#FFEBCD')

ws5.autofilter('A1:S20000')
for i, (testpack, ustan, flud, metr_ng, stat_id_1, stat_id_2, stat_id_3, inst_rfi, metr_inst, test_rfi, elev,
        twelw, one, two, three, four, five, six, seven) in enumerate(iso_summary_table, start=1):
    if testpack == 'Изометрия':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color_2 = cell_format_hat
        color_2.set_bold('bold')
        color_2.set_text_wrap(text_wrap=1)
    elif stat_id_3:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    ws5.write(f'A{i}', testpack, color)
    ws5.write(f'B{i}', ustan, color)
    ws5.write(f'C{i}', flud, color)
    ws5.write(f'D{i}', metr_ng, color)
    ws5.write(f'E{i}', stat_id_1, color)
    ws5.write(f'F{i}', stat_id_2, color)
    ws5.write(f'G{i}', stat_id_3, color)
    # ws5.write(f'H{i}', inst_rfi, color)
    # ws5.write(f'I{i}', metr_inst, color)
    ws5.write(f'H{i}', test_rfi, color)
    ws5.write(f'I{i}', elev, color)
    ws5.write(f'J{i}', twelw, color)
    ws5.write(f'K{i}', one, color)
    ws5.write(f'L{i}', two, color)
    ws5.write(f'M{i}', three, color_2)
    ws5.write(f'N{i}', four, color_2)
    ws5.write(f'O{i}', five, color_2)
    ws5.write(f'P{i}', six, color_2)
    ws5.write(f'Q{i}', seven, color_2)

ws01 = workbook_summary.add_worksheet('Double_iso')
ws01.set_column(0, 0, 37)
ws01.set_column(1, 1, 32)
ws01.set_column(2, 3, 12)
ws01.set_column(4, 7, 22)

ws01.autofilter('A1:S20000')

for i, (one, two, three, four, five, six, seven, eight) in enumerate(double_iso_summary_table, start=1):
    if one == 'Изометрия':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color_2 = cell_format_hat
        color_2.set_bold('bold')
        color_2.set_text_wrap(text_wrap=1)
    elif eight:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    ws01.write(f'A{i}', one, color)
    ws01.write(f'B{i}', two, color)
    ws01.write(f'C{i}', three, color)
    ws01.write(f'D{i}', four, color)
    ws01.write(f'E{i}', five, color)
    ws01.write(f'F{i}', six, color)
    ws01.write(f'G{i}', seven, color)
    ws01.write(f'H{i}', eight, color)

workbook_summary.close()

print('Итоговые файлы созданы.')
print('Прошёл ФАЗУ 2 \n\n ---------')


