# -*- coding: cp1251 -*-

import openpyxl as xl
import re
import xlsxwriter
from dateutil import parser
import datetime
import pandas as pd


# Создание общего словаря ТП-------------------------------------------
wb_phase1 = xl.load_workbook(r'C:\Users\ignatenkoia\PycharmProjects\GIT_PROJECTS\handlers_sg\БД ТП ФАЗА 1.xlsx')
sheet_TP_phase1 = wb_phase1['TP_PHASE1']
sheet_iso_tp_phase1 = wb_phase1['ISO_PHASE1']

testpackages_p1 = {}

for l in sheet_TP_phase1['A3':'K2764']:
    tp_number = ' '
    tp_short_code_BD = str(l[0].value).strip()
    title = str(l[2].value).strip()
    fluid = str(l[3].value).strip()
    category_tp = str(l[4].value)
    erection_rfi = ' ' + str(l[5].value).strip()
    test_rfi = ' ' + str(l[6].value).strip()
    airb_rfi = ' ' + str(l[7].value).strip()
    reinst_rfi = ' ' + str(l[8].value).strip()
    length = round(float(str(l[10].value)), 3)

    testpackages_p1[tp_short_code_BD] = [title, fluid, category_tp, length, '', '', erection_rfi, test_rfi, airb_rfi,
                                         reinst_rfi, tp_number]

print(f'В словаре {len(testpackages_p1.keys())}  ТП.')

isotpdic_p1 = {}
sc_isotpdic_p1 = {}

for i in sheet_iso_tp_phase1['A3':'S17651']:
    sc_iso_tp = str(i[3].value).strip() + str(i[0].value).strip()
    isometric = str(i[3].value).strip()
    testpack = str(i[0].value).strip()
    isolength = round(float(str(i[8].value)), 3)
    title_iso = str(i[4].value).strip()
    type_insulation = str(i[9].value).strip()
    try:
        area_insulation_tt_zra = round(float(i[10].value), 3)
    except:
        area_insulation_tt_zra = str(i[10].value).strip()
    try:
        area_ins_zra = round(float(i[11].value), 3)
    except:
        area_ins_zra = str(i[11].value).strip()
    try:
        count_ins_zra = round(float(i[12].value), 3)
    except:
        count_ins_zra = str(i[12].value).strip()
    try:
        area_insulation_tt = round(float(i[13].value), 3)
    except:
        area_insulation_tt = str(i[13].value).strip()

    rfi_min_vata_tt = ' ' + str(i[14].value).strip()
    rfi_metall_tt = ' ' + str(i[15].value).strip()
    rfi_foamglass_tt = ' ' + str(i[16].value).strip()
    rfi_metall_box = ' ' + str(i[17].value).strip()
    rfi_therm_cover = ' ' + str(i[18].value).strip()

    isotpdic_p1[isometric] = [testpack, isolength, title_iso, type_insulation, area_insulation_tt, area_ins_zra,
                              count_ins_zra, area_insulation_tt_zra, rfi_min_vata_tt, rfi_metall_tt, rfi_foamglass_tt,
                              rfi_metall_box, rfi_therm_cover, '']
    sc_isotpdic_p1[sc_iso_tp] = [isometric, testpack, isolength, title_iso, '', '', '', '']

wb_phase1.close()
print(f'Уникальных изометрий - {len(sc_isotpdic_p1.keys())} \n '
      f'Переходящих изометрий - {len(sc_isotpdic_p1.keys()) - len(isotpdic_p1.keys())}')
print('Общий словарь создан')
# Общий словарь создан--------------------------------------------------


# Добавление в словарь информации по движению ИД------------------------
wb_id = xl.load_workbook('Хронология движения ИД ТТ 1 Фаза.xlsx')
sheet_id = wb_id['ХРОНОЛОГИЯ']

status_id_p1 = {}
for i in sheet_id['C2':'L3171']:
    number_of_testpack = str(i[0].value).strip()
    date_prov_string = str(i[8].value)
    try:
        critical_comment = str(i[9].value)
    except:
        critical_comment = 'Не проверен'
    stat_crit = 'КРИТИКА'
    if 'нет' in critical_comment:
        stat_crit = 'ok'

    status_id_p1[number_of_testpack] = [date_prov_string, stat_crit]
print(len(status_id_p1.keys()))

for key in testpackages_p1.keys():
    a_1 = key.replace('YMT-', '')
    tp_for_id = a_1.replace('-HP', '')
    if tp_for_id in status_id_p1.keys():
        testpackages_p1[key][4] = status_id_p1[tp_for_id][0]
        testpackages_p1[key][5] = status_id_p1[tp_for_id][1]
    else:
        pass

wb_id.close()
print('Добавил информацию по Хронологии движения ИД в общий словарь')
# -------------------------------------------------------------------


# Проверка Журнал заявок АИС Р2 ФАЗА1-------------------------------------
df = pd.read_excel('Журнал заявок 1 фаза + спутники.xlsx')
df = df.sort_values(by='Дата назначения инспекции / Date of scheduled inspection', ascending=True)
df.to_excel('Журнал заявок 1 фаза + спутники.xlsx', index=0)

wb_journal_rfi_p1 = xl.load_workbook('Журнал заявок 1 фаза + спутники.xlsx')
sheet = wb_journal_rfi_p1['Sheet1']

replace_pattern_1 = ['-HT', '-VT', '-PT']
replace_pattern_2 = ['(T.T. REINSTATEMENT)', '(T.T. AIR BLOWING)', '(AIR BLOWING)', '(T.T AIR BLOWING',
                     '(T.T. ERECTION)', '(T.T.TEST)', '(T.T. TEST)',
                     '(T.T ERECTION)', '(T.T TEST)', '(T.T REINSTATEMENT)', '(T.T AIR BLOWING)', '(GPA AIR BLOWING)',
                     '(GPA TEST)',
                     '(GPA ERECTION)', '(GPA REINSTATEMENT)']
res_summary_p1 = {}
for i in sheet['B2':'AO30000']:
    rfi_number = str(i[1].value)
    tp_number = str(i[2].value)
    pkk = str(i[4].value)
    tp_shortname = ''
    tp_shortname_1 = ''
    for l in replace_pattern_2:
        if l in tp_number:
            tp_shortname_1 = tp_number.replace(l, '').strip()
        else:
            pass
    if '-HT' in tp_shortname_1:
        tp_shortname = tp_shortname_1.replace('-HT', '')
    elif '-PT' in tp_shortname_1:
        tp_shortname = tp_shortname_1.replace('-PT', '')
    elif '-VT' in tp_shortname_1:
        tp_shortname = tp_shortname_1.replace('-VT', '')
    else:
        tp_shortname = tp_shortname_1

    description_rfi = str(i[16].value)
    name_insp = str(i[26].value)
    list_iso = str(i[8].value)
    volume_meter = re.sub(r'[^0-9.]', '', str(i[18].value))
    category_cancelled = str(i[31].value)
    violation = str(i[34].value)
    comment = str(i[39].value)  # комментарий для сортировки Физ. объём подтверждён  на прочность и плотность

    if tp_shortname in testpackages_p1.keys():
        if 'Принято' in category_cancelled:
            if 'сборки технологических трубопроводов ГПА' in description_rfi:
                testpackages_p1[tp_shortname][9] = rfi_number
            if 'онтаж технологического трубопровода ГПА' in description_rfi:
                testpackages_p1[tp_shortname][6] = rfi_number
            if 'испытаний технологического трубопровода ГПА' in description_rfi:
                testpackages_p1[tp_shortname][7] = rfi_number
            if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                testpackages_p1[tp_shortname][6] = rfi_number
            if 'испыт' and 'рочност' in description_rfi:
                testpackages_p1[tp_shortname][7] = rfi_number
            if 'испытаний технологического трубопровода  на прочность' in description_rfi or 'Гидравлические испытания'\
                    in description_rfi:
                testpackages_p1[tp_shortname][7] = rfi_number
            if 'сборки технологических трубопроводов в проект' in description_rfi:
                testpackages_p1[tp_shortname][9] = rfi_number
            if 'родувка' in description_rfi:
                testpackages_p1[tp_shortname][8] = rfi_number
        else:
            if 'подтвержд' in comment:
                if 'сборки технологических трубопроводов ГПА' in description_rfi:
                    testpackages_p1[tp_shortname][9] = rfi_number + ' ФОП'
                if 'онтаж технологического трубопровода ГПА' in description_rfi:
                    testpackages_p1[tp_shortname][6] = rfi_number + ' ФОП'
                if 'испытаний технологического трубопровода ГПА' in description_rfi:
                    testpackages_p1[tp_shortname][7] = rfi_number + ' ФОП'
                if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                    testpackages_p1[tp_shortname][6] = rfi_number + ' ФОП'
                if 'испыт' and 'рочност' in description_rfi:
                    testpackages_p1[tp_shortname][7] = rfi_number + ' ФОП'
                if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                    testpackages_p1[tp_shortname][7] = rfi_number + ' ФОП'
                if 'сборки технологических трубопроводов в проект' in description_rfi:
                    testpackages_p1[tp_shortname][9] = rfi_number + ' ФОП'
                if 'родувка' in description_rfi:
                    testpackages_p1[tp_shortname][8] = rfi_number + ' ФОП'
            if 'зафиксирован' in comment:
                if 'испытаний технологического трубопровода ГПА' in description_rfi:
                    testpackages_p1[tp_shortname][7] = rfi_number + ' ФОП'
                if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                    testpackages_p1[tp_shortname][6] = rfi_number + ' ФОП'
                if 'испыт' and 'рочност' in description_rfi:
                    testpackages_p1[tp_shortname][7] = rfi_number + ' ФОП'
                if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                    testpackages_p1[tp_shortname][7] = rfi_number + ' ФОП'
                if 'сборки технологических трубопроводов в проект' in description_rfi:
                    testpackages_p1[tp_shortname][9] = rfi_number + ' ФОП'
                if 'родувка' in description_rfi:
                    testpackages_p1[tp_shortname][8] = rfi_number + ' ФОП'



    for isom in list_iso.split(';'):
        if isom.strip() + tp_shortname in sc_isotpdic_p1.keys():
            if 'Принято' in category_cancelled:
                if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                    sc_isotpdic_p1[isom.strip() + tp_shortname][4] = rfi_number
                if 'испыт' and 'рочност' in description_rfi:
                    sc_isotpdic_p1[isom.strip() + tp_shortname][5] = rfi_number
                if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                    sc_isotpdic_p1[isom.strip() + tp_shortname][5] = rfi_number
                if 'сборки технологических трубопроводов в проект' in description_rfi:
                    sc_isotpdic_p1[isom.strip() + tp_shortname][7] = rfi_number
                if 'родувка' in description_rfi:
                    sc_isotpdic_p1[isom.strip() + tp_shortname][6] = rfi_number
            else:
                if 'подтвержд' in comment:
                    if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][4] = rfi_number + ' ФОП'
                    if 'испыт' and 'рочност' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][5] = rfi_number + ' ФОП'
                    if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][5] = rfi_number + ' ФОП'
                    if 'сборки технологических трубопроводов в проект' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][7] = rfi_number + ' ФОП'
                    if 'родувка' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][6] = rfi_number + ' ФОП'
                if 'зафиксирован' in comment:
                    if 'Монтаж технологического трубопровода в рамках' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][4] = rfi_number + ' ФОП'
                    if 'испыт' and 'рочност' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][5] = rfi_number + ' ФОП'
                    if 'испытаний технологического трубопровода  на прочность' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][5] = rfi_number + ' ФОП'
                    if 'сборки технологических трубопроводов в проект' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][7] = rfi_number + ' ФОП'
                    if 'родувка' in description_rfi:
                        sc_isotpdic_p1[isom.strip() + tp_shortname][6] = rfi_number + ' ФОП'


        if isom.strip() in isotpdic_p1.keys():
            if 'дополн' in description_rfi:
                if 'Принято' in category_cancelled:
                    isotpdic_p1[isom.strip()][13] = rfi_number
                else:
                    if 'выдерж' in comment:
                        isotpdic_p1[isom.strip()][13] = rfi_number

    # ПРОВЕРКА ИЗОЛЯЦИИ-------------------------------------------
    if 'завершении работ по теплоизоляц' in pkk:
        if 'представлены не в полном объеме, представлены некорректные документы' in violation:
            if 'металлического кожуха фланцев и ЗРА' in description_rfi:
                for iso in list_iso.split(';'):
                    if iso in isotpdic_p1.keys():
                        isotpdic_p1[iso.strip()][11] = rfi_number
            if 'металлического кожуха согласно изометрическим' in description_rfi:
                for iso in list_iso.split(';'):
                    if iso in isotpdic_p1.keys():
                        isotpdic_p1[iso.strip()][9] = rfi_number
            if 'теплоизоляционного покрытия согласно изометрическим' in description_rfi:
                for iso in list_iso.split(';'):
                    if iso in isotpdic_p1.keys():
                        isotpdic_p1[iso.strip()][8] = rfi_number
            if 'FOAMGLAS' in tp_number or 'oamglas' in tp_number:
                for iso in list_iso.split(';'):
                    if iso in isotpdic_p1.keys():
                        isotpdic_p1[iso.strip()][10] = rfi_number
            if 'теплоизоляционной оболочки ( термочехлов)' in description_rfi:
                for iso in list_iso.split(';'):
                    if iso in isotpdic_p1.keys():
                        isotpdic_p1[iso.strip()][12] = rfi_number
        if 'документы, подтверждающие качество работ' in violation:
            if 'подтвержд' in comment:
                if 'теплоизоляционной оболочки ( термочехлов)' in description_rfi:
                    for iso in list_iso.split(';'):
                        if iso in isotpdic_p1.keys():
                            isotpdic_p1[iso.strip()][12] = rfi_number
                if 'металлического кожуха фланцев и ЗРА' in description_rfi:
                    for iso in list_iso.split(';'):
                        if iso in isotpdic_p1.keys():
                            isotpdic_p1[iso.strip()][11] = rfi_number
                if 'металлического кожуха согласно изометрическим' in description_rfi:
                    for iso in list_iso.split(';'):
                        if iso in isotpdic_p1.keys():
                            isotpdic_p1[iso.strip()][9] = rfi_number
                if 'теплоизоляционного покрытия согласно изометрическим' in description_rfi:
                    for iso in list_iso.split(';'):
                        if iso in isotpdic_p1.keys():
                            isotpdic_p1[iso.strip()][8] = rfi_number
                if 'FOAMGLAS' in tp_number or 'oamglas' in tp_number:
                    for iso in list_iso.split(';'):
                        if iso in isotpdic_p1.keys():
                            isotpdic_p1[iso.strip()][10] = rfi_number

# ------------------------------------------------------------

print('Проверил Журнал заявок')
wb_journal_rfi_p1.close()
# ------------------------------------------------------------------------
# -Проверка на уведомления-------------------------
wb_ncr = xl.load_workbook('Реестр уведомлений.xlsx')
sheet_ncr = wb_ncr['Предписания (Instructions)']
iso_ncr = {}
iso_ncr_p1 = {}
for i in sheet_ncr['B4':'V4500']:
    number_ncr = str(i[0].value)
    mark_execution = str(i[16].value)
    notification_items = str(i[1].value)
    type_violation = str(i[5].value)
    content_remarks = str(i[6].value).replace(' ', '')
    content_remarks_iso = re.findall(r'\d-\d-\d-\d\d-\d\d\d-\s?\w*\+?-[0-9A-Z][0-9A-Z]-\d\d\d\d-\d\d\d',
                                     content_remarks)
    content_remarks_joints = re.findall(r'\s{1}[Ss]\s?\-?\d*.\d*|\s{1}F\s?\-?\d*.\d*', str(i[6].value))
    joint_mark = []
    for i in content_remarks_joints:
        joint1 = i.replace(' ', '')
        joint = joint1.replace('-', '')
        joint_mark.append((re.sub(r'[\.\:\;]$', '', joint)).strip())
    if 'Нет' in mark_execution:
        if content_remarks_iso:
            for l in content_remarks_iso:
                iso_ncr_p1[l] = number_ncr
                try:
                    iso_ncr[isotpdic_p1[l][0]] = number_ncr
                except:
                    pass

for key in testpackages_p1.keys():
    if key in iso_ncr.keys():
        testpackages_p1[key].append(iso_ncr[key])
    else:
        testpackages_p1[key].append('Уведомлений нет')
wb_ncr.close()
print('Добавил информацию из Реестра уведомлений')
# --------------------------------------------------------


info_summary_table_phase1 = [
    ['TP ShortCode', 'TP LongCode', 'Установка', 'Среда', 'Категория', 'Длина', 'Дата проверки ИД',
     'Статус проверки ИД',
     'RFI ERECTION', 'RFI TEST', 'RFI AIR BLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]

for key in testpackages_p1.keys():
    info_summary_table_phase1.append(
        [key, testpackages_p1[key][10], testpackages_p1[key][0], testpackages_p1[key][1], testpackages_p1[key][2],
         testpackages_p1[key][3],
         testpackages_p1[key][4], testpackages_p1[key][5], testpackages_p1[key][6], testpackages_p1[key][7],
         testpackages_p1[key][8],
         testpackages_p1[key][9], testpackages_p1[key][11]])

info_summary_iso_phase1 = [
    ['Iso', 'TP ShortCode', 'Length', 'Title', 'Статус проверки ИД по ТП', 'RFI ERECTION', 'RFI TEST',
     'RFI AIR BLOWING', 'RFI REINSTATEMENT', 'RFI ДИГ', 'Тип изоляции', 'Площадь изоляции ТТ, м2',
     'Площадь изоляции ЗРА, м2',
     'Количество термочехлов, шт.', 'Площадь изоляции ТТ+ЗРА (короб), м2', 'RFI Мин-вата ТТ', 'RFI Металл ТТ',
     'RFI Пеностекло', 'RFI Металл Короб.', 'RFI Термочехлы', 'Статус уведомлений']]
for key in isotpdic_p1.keys():
    status_key_ncr = 'Уведомлений нет'
    if key in iso_ncr_p1.keys():
        status_key_ncr = iso_ncr_p1[key]
    try:
        info_summary_iso_phase1.append([key, isotpdic_p1[key][0], isotpdic_p1[key][1], isotpdic_p1[key][2],
                                        testpackages_p1[isotpdic_p1[key][0]][5],
                                        testpackages_p1[isotpdic_p1[key][0]][6],
                                        testpackages_p1[isotpdic_p1[key][0]][7],
                                        testpackages_p1[isotpdic_p1[key][0]][8],
                                        testpackages_p1[isotpdic_p1[key][0]][9],
                                        isotpdic_p1[key][13],
                                        isotpdic_p1[key][3], isotpdic_p1[key][4], isotpdic_p1[key][5],
                                        isotpdic_p1[key][6],
                                        isotpdic_p1[key][7], isotpdic_p1[key][8], isotpdic_p1[key][9],
                                        isotpdic_p1[key][10],
                                        isotpdic_p1[key][11], isotpdic_p1[key][12], status_key_ncr])
    except:
        pass

# # ---Проверка на проведение АЭ --------------------------------------
# wb_ae = xl.load_workbook('Испытания АЭ P2.xlsx')
# sheet_ae = wb_ae['Трубопроводы Р2']
# tp_done_ae = {}
# for i in sheet_ae['C200':'G1000']:
#     try:
#         date_ae = str(i[0].value)[0:10]
#         tp_ae = str(i[2].value).strip()
#         tp_done_ae[tp_ae] = date_ae
#     except:
#         pass
#
# print('Информация о проведении АЭ добавлена.')
# # ------------------------------------------------------------------


n_dic_1_60 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC4P+': ['Бутановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3P+': ['Пропановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC5+': ['С5+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC4+': ['С4+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNMP': ['Азот СД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOWWA': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WMMI': ['Водометанольная смесь', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSU': ['Поверхностная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HYDV': ['Пары углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
              }

n_dic_1_70 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_list_1_60 = [['', f'Статус по ТП 1-60 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                '', '', '', '', '',
                '', '', '-'], ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                               'Принят монтаж, ТП',
                               'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.',
                               'Принята продувка, ТП', 'Принята ОС, м.',
                               'Принята ОС, ТП', 'Остаток ОС, м.', 'Остаток ОС, ТП']]
ITOG_list_1_60 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_1_70 = [['', f'Статус по ТП 1-70 на {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                '', '', '', '', '',
                '', '', '-'], ['Код среды', 'Наименование среды', 'По проекту, м.', 'Кол-во ТП', 'Принят монтаж, м.',
                               'Принят монтаж, ТП',
                               'Приняты испыт-я, м.', 'Приняты испыт-я, ТП', 'Принята продувка, м.',
                               'Принята продувка, ТП', 'Принята ОС, м.',
                               'Принята ОС, ТП', 'Остаток ОС, м.', 'Остаток ОС, ТП']]
ITOG_list_1_70 = ['', 'Итого:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

brief_summary = [['Краткая сводка 1-60', 'Дата ', f'{datetime.datetime.now().strftime("%d.%m.%Y")}',
                  ' ', ' ']]

for key in testpackages_p1.keys():

    if testpackages_p1[key][0] == '1-60':
        n_dic_1_60[testpackages_p1[key][1]][1] += testpackages_p1[key][3]
        n_dic_1_60[testpackages_p1[key][1]][2] += 1
        if 'CC' in testpackages_p1[key][6]:
            n_dic_1_60[testpackages_p1[key][1]][3] += testpackages_p1[key][3]
            n_dic_1_60[testpackages_p1[key][1]][4] += 1
        if 'CC' in testpackages_p1[key][7]:
            n_dic_1_60[testpackages_p1[key][1]][5] += testpackages_p1[key][3]
            n_dic_1_60[testpackages_p1[key][1]][6] += 1
        if 'CC' not in testpackages_p1[key][7] and 'CC' in testpackages_p1[key][8]:
            n_dic_1_60[testpackages_p1[key][1]][5] += testpackages_p1[key][3]
            n_dic_1_60[testpackages_p1[key][1]][6] += 1
        if 'CC' in testpackages_p1[key][8]:
            n_dic_1_60[testpackages_p1[key][1]][7] += testpackages_p1[key][3]
            n_dic_1_60[testpackages_p1[key][1]][8] += 1
        if 'CC' in testpackages_p1[key][9]:
            n_dic_1_60[testpackages_p1[key][1]][9] += testpackages_p1[key][3]
            n_dic_1_60[testpackages_p1[key][1]][10] += 1
        n_dic_1_60[testpackages_p1[key][1]][11] = n_dic_1_60[testpackages_p1[key][1]][1] - \
                                                  n_dic_1_60[testpackages_p1[key][1]][9]
        n_dic_1_60[testpackages_p1[key][1]][12] = n_dic_1_60[testpackages_p1[key][1]][2] - \
                                                  n_dic_1_60[testpackages_p1[key][1]][10]

    if testpackages_p1[key][0] == '1-70':
        n_dic_1_70[testpackages_p1[key][1]][1] += testpackages_p1[key][3]
        n_dic_1_70[testpackages_p1[key][1]][2] += 1
        if 'CC' in testpackages_p1[key][6]:
            n_dic_1_70[testpackages_p1[key][1]][3] += testpackages_p1[key][3]
            n_dic_1_70[testpackages_p1[key][1]][4] += 1
        if 'CC' in testpackages_p1[key][7]:
            n_dic_1_70[testpackages_p1[key][1]][5] += testpackages_p1[key][3]
            n_dic_1_70[testpackages_p1[key][1]][6] += 1
        if 'CC' not in testpackages_p1[key][7] and 'CC' in testpackages_p1[key][8]:
            n_dic_1_70[testpackages_p1[key][1]][5] += testpackages_p1[key][3]
            n_dic_1_70[testpackages_p1[key][1]][6] += 1
        if 'CC' in testpackages_p1[key][8]:
            n_dic_1_70[testpackages_p1[key][1]][7] += testpackages_p1[key][3]
            n_dic_1_70[testpackages_p1[key][1]][8] += 1
        if 'CC' in testpackages_p1[key][9]:
            n_dic_1_70[testpackages_p1[key][1]][9] += testpackages_p1[key][3]
            n_dic_1_70[testpackages_p1[key][1]][10] += 1
        n_dic_1_70[testpackages_p1[key][1]][11] = n_dic_1_70[testpackages_p1[key][1]][1] - \
                                                  n_dic_1_70[testpackages_p1[key][1]][9]
        n_dic_1_70[testpackages_p1[key][1]][12] = n_dic_1_70[testpackages_p1[key][1]][2] - \
                                                  n_dic_1_70[testpackages_p1[key][1]][10]

for key in n_dic_1_60.keys():
    n_list_1_60.append([key, n_dic_1_60[key][0], n_dic_1_60[key][1], n_dic_1_60[key][2], n_dic_1_60[key][3],
                        n_dic_1_60[key][4], n_dic_1_60[key][5], n_dic_1_60[key][6], n_dic_1_60[key][7],
                        n_dic_1_60[key][8], n_dic_1_60[key][9], n_dic_1_60[key][10], n_dic_1_60[key][11],
                        n_dic_1_60[key][12]])
    ITOG_list_1_60[2] += n_dic_1_60[key][1]
    ITOG_list_1_60[3] += n_dic_1_60[key][2]
    ITOG_list_1_60[4] += n_dic_1_60[key][3]
    ITOG_list_1_60[5] += n_dic_1_60[key][4]
    ITOG_list_1_60[6] += n_dic_1_60[key][5]
    ITOG_list_1_60[7] += n_dic_1_60[key][6]
    ITOG_list_1_60[8] += n_dic_1_60[key][7]
    ITOG_list_1_60[9] += n_dic_1_60[key][8]
    ITOG_list_1_60[10] += n_dic_1_60[key][9]
    ITOG_list_1_60[11] += n_dic_1_60[key][10]
    ITOG_list_1_60[12] += n_dic_1_60[key][11]
    ITOG_list_1_60[13] += n_dic_1_60[key][12]

n_list_1_60.append(ITOG_list_1_60)
ost_list_1_60 = ['', 'Остаток:', '', '', ITOG_list_1_60[2] - ITOG_list_1_60[4], ITOG_list_1_60[3] - ITOG_list_1_60[5],
                 ITOG_list_1_60[2] - ITOG_list_1_60[6], ITOG_list_1_60[3] - ITOG_list_1_60[7],
                 ITOG_list_1_60[2] - ITOG_list_1_60[8], ITOG_list_1_60[3] - ITOG_list_1_60[9],
                 ITOG_list_1_60[2] - ITOG_list_1_60[10], ITOG_list_1_60[3] - ITOG_list_1_60[11], '', '']
n_list_1_60.append(ost_list_1_60)
empty_str = [' * ', ' * ', ' * ', '  *  ', '  *  ', '  *  ', '  *  ', '  ', '  ', '  ', '  ', '', '', '-']
n_list_1_60.append(empty_str)

brief_summary.append(['Всего ТП/м.', f'{ITOG_list_1_60[3]}', 'Остаток, ТП', f'{round(ITOG_list_1_60[2], 3)}', 'Остаток, м.'])
brief_summary.append(['Конструктив принят', f'{ITOG_list_1_60[5]}', f'{ost_list_1_60[5]}',
                      f'{round(ITOG_list_1_60[4], 3)}', f'{round(ost_list_1_60[4], 3)}'])
brief_summary.append(['Испытания зафиксированы', f'{ITOG_list_1_60[7]}', f'{ost_list_1_60[7]}',
                      f'{round(ITOG_list_1_60[6], 3)}', f'{round(ost_list_1_60[6], 3)}'])
brief_summary.append(['Продувка зафиксирована', f'{ITOG_list_1_60[9]}', f'{ost_list_1_60[9]}',
                      f'{round(ITOG_list_1_60[8], 3)}', f'{round(ost_list_1_60[8], 3)}'])
brief_summary.append(['Обратная сборка принята', f'{ITOG_list_1_60[11]}', f'{ost_list_1_60[11]}',
                      f'{round(ITOG_list_1_60[10], 3)}', f'{round(ost_list_1_60[10], 3)}'])
brief_summary.append(['*', '*', '*', '*', '*',])
brief_summary.append(['Краткая сводка 1-70', 'Дата ', f'{datetime.datetime.now().strftime("%d.%m.%Y")}',
                  ' ', ' '])

for key in n_dic_1_70.keys():
    n_list_1_70.append([key, n_dic_1_70[key][0], n_dic_1_70[key][1], n_dic_1_70[key][2], n_dic_1_70[key][3],
                        n_dic_1_70[key][4], n_dic_1_70[key][5], n_dic_1_70[key][6], n_dic_1_70[key][7],
                        n_dic_1_70[key][8], n_dic_1_70[key][9], n_dic_1_70[key][10], n_dic_1_70[key][11],
                        n_dic_1_70[key][12]])
    ITOG_list_1_70[2] += n_dic_1_70[key][1]
    ITOG_list_1_70[3] += n_dic_1_70[key][2]
    ITOG_list_1_70[4] += n_dic_1_70[key][3]
    ITOG_list_1_70[5] += n_dic_1_70[key][4]
    ITOG_list_1_70[6] += n_dic_1_70[key][5]
    ITOG_list_1_70[7] += n_dic_1_70[key][6]
    ITOG_list_1_70[8] += n_dic_1_70[key][7]
    ITOG_list_1_70[9] += n_dic_1_70[key][8]
    ITOG_list_1_70[10] += n_dic_1_70[key][9]
    ITOG_list_1_70[11] += n_dic_1_70[key][10]
    ITOG_list_1_70[12] += n_dic_1_70[key][11]
    ITOG_list_1_70[13] += n_dic_1_70[key][12]

n_list_1_70.append(ITOG_list_1_70)
ost_list_1_70 = ['', 'Остаток:', '', '', ITOG_list_1_70[2] - ITOG_list_1_70[4], ITOG_list_1_70[3] - ITOG_list_1_70[5],
                 ITOG_list_1_70[2] - ITOG_list_1_70[6], ITOG_list_1_70[3] - ITOG_list_1_70[7],
                 ITOG_list_1_70[2] - ITOG_list_1_70[8], ITOG_list_1_70[3] - ITOG_list_1_70[9],
                 ITOG_list_1_70[2] - ITOG_list_1_70[10], ITOG_list_1_70[3] - ITOG_list_1_70[11], '', '']
n_list_1_70.append(ost_list_1_70)

brief_summary.append(['Всего ТП/м.', f'{ITOG_list_1_70[3]}', 'Остаток, ТП', f'{round(ITOG_list_1_70[2], 3)}', 'Остаток, м.'])
brief_summary.append(['Конструктив принят', f'{ITOG_list_1_70[5]}', f'{ost_list_1_70[5]}',
                      f'{round(ITOG_list_1_70[4], 3)}', f'{round(ost_list_1_70[4], 3)}'])
brief_summary.append(['Испытания зафиксированы', f'{ITOG_list_1_70[7]}', f'{ost_list_1_70[7]}',
                      f'{round(ITOG_list_1_70[6], 3)}', f'{round(ost_list_1_70[6], 3)}'])
brief_summary.append(['Продувка зафиксирована', f'{ITOG_list_1_70[9]}', f'{ost_list_1_70[9]}',
                      f'{round(ITOG_list_1_70[8], 3)}', f'{round(ost_list_1_70[8], 3)}'])
brief_summary.append(['Обратная сборка принята', f'{ITOG_list_1_70[11]}', f'{ost_list_1_70[11]}',
                      f'{round(ITOG_list_1_70[10], 3)}', f'{round(ost_list_1_70[10], 3)}'])
brief_summary.append(['*', '*', '*', '*', '*',])

for i in n_list_1_70:
    n_list_1_60.append(i)
n_list_1_60.append(empty_str)

double_iso_summary_table_p1 = [['Изометрия', 'Тестпакет', 'Длина', 'Установка', 'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING',
                             'RFI REINSTATEMENT']]
for key in sc_isotpdic_p1.keys():
    double_iso_summary_table_p1.append([sc_isotpdic_p1[key][0], sc_isotpdic_p1[key][1], sc_isotpdic_p1[key][2],
                                        sc_isotpdic_p1[key][3], sc_isotpdic_p1[key][4], sc_isotpdic_p1[key][5],
                                        sc_isotpdic_p1[key][6], sc_isotpdic_p1[key][7]])
# ЗАПИСЬ В ФАЙЛ------------------------------------------------

workbook_summary_p1 = xlsxwriter.Workbook(f'Сводка по ФАЗЕ 1 на {datetime.datetime.now().strftime("%d.%m.%Y")}.xlsx')

ws_brief = workbook_summary_p1.add_worksheet('Краткая сводка')

ws_brief.set_column(0, 0, 35)
ws_brief.set_column(1, 4, 25)

cell_format_green = workbook_summary_p1.add_format()
cell_format_green.set_bg_color('#98FB98')
cell_format_blue = workbook_summary_p1.add_format()
cell_format_blue.set_bg_color('#B0E0E6')
cell_format_hat = workbook_summary_p1.add_format()
cell_format_hat.set_bg_color('#FFDAB9')
cell_format_date = workbook_summary_p1.add_format()
cell_format_date.set_font_size(font_size=14)

for i, (one, two, three, four, five) in enumerate(brief_summary, start=1):
    color_1 = cell_format_green
    color_1.set_bold('bold')
    if 'Краткая' in one or 'Всего' in one:
        color = cell_format_hat
        color.set_bold('bold')
        color_1 = cell_format_hat
        color_1.set_bold('bold')
    elif '*' in one:
        color_1 = cell_format_blue
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
        color_1.set_border(style=1)
        color_1.set_text_wrap(text_wrap=1)
    except:
        pass

    ws_brief.write(f'A{i}', one, color_1)
    ws_brief.write(f'B{i}', two, color)
    ws_brief.write(f'C{i}', three, color)
    ws_brief.write(f'D{i}', four, color)
    ws_brief.write(f'E{i}', five, color)


ws0 = workbook_summary_p1.add_worksheet('Cводка по установкам')
ws0.set_column(0, 0, 12)
ws0.set_column(1, 1, 40)
ws0.set_column(4, 13, 12)
ws0.set_column(2, 2, 12)
ws0.set_column(3, 3, 12)


for i, (one, two, three, four, five, six, seven, eight, nine, ten, eleven, twelve, thirteen,
        fourteen) in enumerate(n_list_1_60, start=2):
    if fourteen == 0:
        color = cell_format_green
    elif fourteen == 'Остаток ОС, ТП':
        color = cell_format_hat
        color.set_bold('bold')
    elif fourteen == '-':
        color = cell_format_date
        color.set_bold('bold')
    elif two == 'Итого:':
        color = cell_format_hat
        color.set_bold('bold')
    elif two == 'Остаток:':
        color = cell_format_hat
        color.set_bold('bold')
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass
    ws0.write(f'A{i}', one, color)
    ws0.write(f'B{i}', two, color)
    ws0.write(f'C{i}', three, color)
    ws0.write(f'D{i}', four, color)
    ws0.write(f'E{i}', five, color)
    ws0.write(f'F{i}', six, color)
    ws0.write(f'G{i}', seven, color)
    ws0.write(f'H{i}', eight, color)
    ws0.write(f'I{i}', nine, color)
    ws0.write(f'J{i}', ten, color)
    ws0.write(f'K{i}', eleven, color)
    ws0.write(f'L{i}', twelve, color)
    ws0.write(f'M{i}', thirteen, color)
    ws0.write(f'N{i}', fourteen, color)

ws5 = workbook_summary_p1.add_worksheet('Сводная информация')
ws5.set_column(0, 1, 30)
ws5.set_column(2, 6, 15)
ws5.set_column(7, 12, 22)
ws5.set_column(13, 15, 25)
ws5.autofilter('A1:O5000')

cell_format_ins = workbook_summary_p1.add_format()
cell_format_ins.set_bg_color('#FFEBCD')
cell_format_green = workbook_summary_p1.add_format()
cell_format_green.set_bg_color('#98FB98')
cell_format_blue = workbook_summary_p1.add_format()
cell_format_blue.set_bg_color('#B0E0E6')
cell_format_hat = workbook_summary_p1.add_format()
cell_format_hat.set_bg_color('#00CED1')
cell_format_date = workbook_summary_p1.add_format()
cell_format_date.set_font_size(font_size=16)
for i, (testpack, ustan, flud, metr_ng, stat_id_1, stat_id_2, inst_rfi, test_rfi, elev, twelw, thirt, fourteen,
        fifth) in enumerate(info_summary_table_phase1, start=1):
    if 'CPECC' in fourteen:
        color = cell_format_green
    elif fifth == 'Статус уведомлений':
        color = cell_format_hat
        color.set_bold('bold')
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        # color.set_text_wrap(text_wrap=1)
    except:
        pass

    ws5.write(f'A{i}', testpack, color)
    ws5.write(f'B{i}', ustan, color)
    ws5.write(f'C{i}', flud, color)
    ws5.write(f'D{i}', metr_ng, color)
    ws5.write(f'E{i}', stat_id_1, color)
    ws5.write(f'F{i}', stat_id_2, color)
    ws5.write(f'G{i}', inst_rfi, color)
    ws5.write(f'H{i}', test_rfi, color)
    ws5.write(f'I{i}', elev, color)
    ws5.write(f'J{i}', twelw, color)
    ws5.write(f'K{i}', thirt, color)
    ws5.write(f'L{i}', fourteen, color)
    ws5.write(f'M{i}', fifth, color)

ws3 = workbook_summary_p1.add_worksheet('Сводка по изометричкам')
ws3.set_column(0, 1, 30)
ws3.set_column(2, 4, 13)
ws3.set_column(5, 9, 22)

ws3.set_column(10, 10, 13)
ws3.set_column(11, 14, 15)
ws3.set_column(15, 20, 22)
ws3.autofilter('A1:T20000')

for i, (testpack, ustan, flud, metr_ng, stat_id_1, stat_id_2, inst_rfi, test_rfi, elev, ten, odinn, twelve, thirteen, fourten,
fiveten, sixten, seventen, eighten, nineten, twenty, t_one) in enumerate(info_summary_iso_phase1, start=1):
    if 'CPECC' in elev:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
        color.set_font_size(font_size=12)
        color_2.set_font_size(font_size=12)
    elif testpack == 'Iso':
        color_2 = cell_format_hat
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color.set_font_size(font_size=16)
        color_2.set_font_size(font_size=16)
    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
        color.set_font_size(font_size=12)
        color_2.set_font_size(font_size=12)
    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    ws3.write(f'A{i}', testpack, color)
    ws3.write(f'B{i}', ustan, color)
    ws3.write(f'C{i}', flud, color)
    ws3.write(f'D{i}', metr_ng, color)
    ws3.write(f'E{i}', stat_id_1, color)
    ws3.write(f'F{i}', stat_id_2, color)
    ws3.write(f'G{i}', inst_rfi, color)
    ws3.write(f'H{i}', test_rfi, color)
    ws3.write(f'I{i}', elev, color)
    ws3.write(f'J{i}', ten, color)
    ws3.write(f'K{i}', odinn, color_2)
    ws3.write(f'L{i}', twelve, color_2)
    ws3.write(f'M{i}', thirteen, color_2)
    ws3.write(f'N{i}', fourten, color_2)
    ws3.write(f'O{i}', fiveten, color_2)
    ws3.write(f'P{i}', sixten, color_2)
    ws3.write(f'Q{i}', seventen, color_2)
    ws3.write(f'R{i}', eighten, color_2)
    ws3.write(f'S{i}', nineten, color_2)
    ws3.write(f'T{i}', twenty, color_2)
    ws3.write(f'U{i}', t_one, color_2)

ws01 = workbook_summary_p1.add_worksheet('Double isometric')
ws01.set_column(0, 0, 37)
ws01.set_column(1, 1, 32)
ws01.set_column(2, 3, 12)
ws01.set_column(4, 7, 22)
ws01.autofilter('A1:S20000')
for i, (one, two, three, four, five, six, seven, eight) in enumerate(double_iso_summary_table_p1, start=1):
    if one == 'Изометрия':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color_2 = cell_format_hat
        color_2.set_bold('bold')
        color_2.set_text_wrap(text_wrap=1)
    elif eight:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    ws01.write(f'A{i}', one, color)
    ws01.write(f'B{i}', two, color)
    ws01.write(f'C{i}', three, color)
    ws01.write(f'D{i}', four, color)
    ws01.write(f'E{i}', five, color)
    ws01.write(f'F{i}', six, color)
    ws01.write(f'G{i}', seven, color)
    ws01.write(f'H{i}', eight, color)



workbook_summary_p1.close()
print('Создан итоговый файл')
