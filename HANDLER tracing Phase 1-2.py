import openpyxl as xl
import re
import xlsxwriter
from dateutil import parser
import datetime
import pandas as pd

wb_sputniks = xl.load_workbook(r'C:\Users\ignatenkoia\PycharmProjects\GIT_PROJECTS\handlers_sg\БД ТП ФАЗА 1, 2.xlsx')
sheet_sputniks = wb_sputniks['5.8']

sputnik_dic = {}
short_draw_sput = {}
for i in sheet_sputniks['E3':'E2797']:
    sputnik_dic[str(i[0].value)] = ['', '', '', 0, '', '', '', '', '']
for i in sheet_sputniks['A3':'Q2797']:
    sputnik_dic[str(i[4].value)][3] += float(str(i[9].value))
    sputnik_dic[str(i[4].value)][2] = str(i[0].value).strip()
    sputnik_dic[str(i[4].value)][1] = str(i[15].value).strip()
    sputnik_dic[str(i[4].value)][0] = str(i[16].value).strip()
    short_draw_sput[str(i[16].value).strip()] = str(i[4].value).strip()
    if i[13].value:
        if 'CC' not in str(i[13].value):
            sputnik_dic[str(i[4].value)][4] = 'CPECC-CC-' + str(i[13].value).strip()
        else:
            sputnik_dic[str(i[4].value)][4] = str(i[13].value).strip()
    if i[14].value:
        if 'CC' not in str(i[14].value):
            sputnik_dic[str(i[4].value)][5] = 'CPECC-CC-' + str(i[14].value).strip()
        else:
            sputnik_dic[str(i[4].value)][5] = str(i[14].value).strip()


df = pd.read_excel('Журнал заявок общий.xlsx')
df = df.sort_values(by='Дата назначения инспекции / Date of scheduled inspection', ascending=True)
df.to_excel('Журнал заявок общий.xlsx',index=0)



wb_tracing = xl.load_workbook('Журнал заявок общий.xlsx')
sheet_tracing = wb_tracing['Sheet1']
for i in sheet_tracing['B2':'AO250000']:
    if i[0].value:
        rfi_number = str(i[1].value)
        description_rfi = str(i[16].value)
        name_insp = str(i[26].value)
        list_iso = str(i[8].value)
        volume_meter = re.sub(r'[^0-9.]', '', str(i[18].value))
        category_cancelled = str(i[31].value)
        comment = str(i[39].value) #комментарий для сортировки Физ. объём подтверждён  на прочность и плотность
        violation = str(i[35].value)

        list_drawing = re.findall(r'0055-CPC-GGC-4\.\d\.\d\.\d\d\.\d\d\d-\w\w\d-ID-\d\d\d\d', description_rfi)
        list_drawing_wrong = re.findall(r'0055-CPC-GGC-4\.\d\.\d\.\d\d\.\d\d\d-\w\w\d-ID-A-\d\d\d\d', description_rfi)

        list_short_draw = re.findall(r'\d-\d\d-HWSM-\d\d\d-\d\d/\d-\d\d-HWSM-\d\d\d-\d\d|\d-\d\d-HWSM-\d\d\d-\d\d', description_rfi.replace(' ', '').strip())

        if 'испытаний теплоспут' in description_rfi:
            if 'документы, подтверждающие' in violation or 'представлены не в полном объеме' in violation:
                fop = ''
                if 'Не принято' == category_cancelled:
                    fop = ' ФОП'
                if list_short_draw:
                    for l in list_short_draw:
                        if l in short_draw_sput.keys():
                            sputnik_dic[short_draw_sput[l]][5] = rfi_number + fop
                if list_drawing_wrong:
                    for z in list_drawing_wrong:
                        try:
                            d_s = z.replace('-A', '')
                            sputnik_dic[d_s][5] = rfi_number + fop
                        except:
                            pass
                if list_drawing:
                    for z in list_drawing:
                        if z in sputnik_dic.keys():
                            sputnik_dic[z][5] = rfi_number + fop
                for z in list_iso.split(';'):
                    if z.strip() in short_draw_sput.keys():
                        sputnik_dic[short_draw_sput[z.strip()]][5] = rfi_number + fop
        if 'испытаний на теплоспутн' in description_rfi:
            if 'документы, подтверждающие' in violation or 'представлены не в полном объеме' in violation:
                fop = ''
                if 'Не принято' == category_cancelled:
                    fop = ' ФОП'
                if list_short_draw:
                    for l in list_short_draw:
                        if l in short_draw_sput.keys():
                            sputnik_dic[short_draw_sput[l]][5] = rfi_number + fop
                if list_drawing_wrong:
                    for z in list_drawing_wrong:
                        try:
                            d_s = z.replace('-A', '')
                            sputnik_dic[d_s][5] = rfi_number + fop
                        except:
                            pass
                if list_drawing:
                    for z in list_drawing:
                        if z in sputnik_dic.keys():
                            sputnik_dic[z][5] = rfi_number + fop
                for z in list_iso.split(';'):
                    if z.strip() in short_draw_sput.keys():
                        sputnik_dic[short_draw_sput[z.strip()]][5] = rfi_number + fop
        if 'онтаж теплоспутника технологич' in description_rfi:
            if 'подтвержд' in comment:
                fop = ''
                if 'Не принято' == category_cancelled:
                    fop = ' ФОП'
                if list_short_draw:
                    for l in list_short_draw:
                        if l in short_draw_sput.keys():
                            sputnik_dic[short_draw_sput[l]][4] = rfi_number + fop
                if list_drawing_wrong:
                    for z in list_drawing_wrong:
                        try:
                            d_s = z.replace('-A', '')
                            sputnik_dic[d_s][4] = rfi_number + fop
                        except:
                            pass
                if list_drawing:
                    for z in list_drawing:
                        if z in sputnik_dic.keys():
                            sputnik_dic[z][4] = rfi_number + fop
                for z in list_iso.split(';'):
                    if z.strip() in short_draw_sput.keys():
                        sputnik_dic[short_draw_sput[z.strip()]][4] = rfi_number + fop
        if 'Продувка теплоспутника' in description_rfi:
            if 'Не предоставлены документы' in violation:
                fop = ''
                if 'Не принято' == category_cancelled:
                    fop = ' ФОП'
                if list_short_draw:
                    for l in list_short_draw:
                        if l in short_draw_sput.keys():
                            sputnik_dic[short_draw_sput[l]][6] = rfi_number + fop
                if list_drawing_wrong:
                    for z in list_drawing_wrong:
                        try:
                            d_s = z.replace('-A', '')
                            sputnik_dic[d_s][6] = rfi_number + fop
                        except:
                            pass
                if list_drawing:
                    for z in list_drawing:
                        if z in sputnik_dic.keys():
                            sputnik_dic[z][6] = rfi_number + fop
                for z in list_iso.split(';'):
                    if z.strip() in short_draw_sput.keys():
                        sputnik_dic[short_draw_sput[z.strip()]][6] = rfi_number + fop
        if 'покрытия теплоспутник' in description_rfi:
            if 'Не предоставлены документы' in violation:
                fop = ''
                if 'Не принято' == category_cancelled:
                    fop = ' ФОП'
                if list_short_draw:
                    for l in list_short_draw:
                        if l in short_draw_sput.keys():
                            sputnik_dic[short_draw_sput[l]][7] = rfi_number + fop
                if list_drawing_wrong:
                    for z in list_drawing_wrong:
                        try:
                            d_s = z.replace('-A', '')
                            sputnik_dic[d_s][7] = rfi_number + fop
                        except:
                            pass
                if list_drawing:
                    for z in list_drawing:
                        if z in sputnik_dic.keys():
                            sputnik_dic[z][7] = rfi_number + fop
                for z in list_iso.split(';'):
                    if z.strip() in short_draw_sput.keys():
                        sputnik_dic[short_draw_sput[z.strip()]][7] = rfi_number + fop
        if 'кожуха теплоспутник' in description_rfi:
            if 'Не предоставлены документы' in violation:
                fop = ''
                if 'Не принято' == category_cancelled:
                    fop = ' ФОП'
                if list_short_draw:
                    for l in list_short_draw:
                        if l in short_draw_sput.keys():
                            sputnik_dic[short_draw_sput[l]][8] = rfi_number + fop
                if list_drawing_wrong:
                    for z in list_drawing_wrong:
                        try:
                            d_s = z.replace('-A', '')
                            sputnik_dic[d_s][8] = rfi_number + fop
                        except:
                            pass
                if list_drawing:
                    for z in list_drawing:
                        if z in sputnik_dic.keys():
                            sputnik_dic[z][8] = rfi_number + fop
                for z in list_iso.split(';'):
                    if z.strip() in short_draw_sput.keys():
                        sputnik_dic[short_draw_sput[z.strip()]][8] = rfi_number + fop

        if '64713' in rfi_number:
            print(list_short_draw)
            print(list_drawing_wrong)
            print(violation)
            print(list_drawing)
            for p in list_drawing_wrong:
                d_s = p.replace('-A', '')
                print(d_s)
    else:
        break


small_summary = [['Краткая справка', 'Общий объём, м.', 'Остаток монтажа, м.', 'Остаток испытаний, м.',
                  'Остаток продувки, м.'],
                 ['Unit 1-110', 0, 0, 0, 0],
                 ['Unit 1-30', 0, 0, 0, 0],
                 ['Unit 2-110', 0, 0, 0, 0],
                 ['Unit 2-30', 0, 0, 0, 0],
                 ['Unit 1-60', 0, 0, 0, 0],
                 ['Unit 1-70', 0, 0, 0, 0],
                 ['Unit 3-110', 0, 0, 0, 0],
                 ['Unit 3-30', 0, 0, 0, 0],
                 ['Unit 2-60', 0, 0, 0, 0],
                 ['Unit 2-70', 0, 0, 0, 0]]

summary_sputnik = [['Чертеж по ГОСТ', 'Чертеж', 'Установка', 'Титул', 'Длина', 'RFI  ERECTION', 'RFI TEST',
                    'RFI BLOWING', 'RFI ВАТА', 'RFI Металл']]
for key in sputnik_dic.keys():
    summary_sputnik.append([key, sputnik_dic[key][0], sputnik_dic[key][1], sputnik_dic[key][2], sputnik_dic[key][3],
                            sputnik_dic[key][4], sputnik_dic[key][5], sputnik_dic[key][6], sputnik_dic[key][7],
                            sputnik_dic[key][8]])



    if '1-110' in sputnik_dic[key][1]:
        small_summary[1][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[1][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[1][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[1][4] += sputnik_dic[key][3]
    if '1-30' in sputnik_dic[key][1]:
        small_summary[2][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[2][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[2][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[2][4] += sputnik_dic[key][3]
    if '2-110' in sputnik_dic[key][1]:
        small_summary[3][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[3][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[3][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[3][4] += sputnik_dic[key][3]
    if '2-30' in sputnik_dic[key][1]:
        small_summary[4][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[4][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[4][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[4][4] += sputnik_dic[key][3]
    if '1-60' in sputnik_dic[key][1]:
        small_summary[5][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[5][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[5][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[5][4] += sputnik_dic[key][3]
    if '1-70' in sputnik_dic[key][1]:
        small_summary[6][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[6][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[6][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[6][4] += sputnik_dic[key][3]
    if '3-110' in sputnik_dic[key][1]:
        small_summary[7][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[7][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[7][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[7][4] += sputnik_dic[key][3]
    if '3-30' in sputnik_dic[key][1]:
        small_summary[8][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[8][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[8][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[8][4] += sputnik_dic[key][3]
    if '2-60' in sputnik_dic[key][1]:
        small_summary[9][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[9][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[9][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[9][4] += sputnik_dic[key][3]
    if '2-70' in sputnik_dic[key][1]:
        small_summary[10][1] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][4]:
            small_summary[10][2] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][5]:
            small_summary[10][3] += sputnik_dic[key][3]
        if 'CPECC' not in sputnik_dic[key][6]:
            small_summary[10][4] += sputnik_dic[key][3]
    #print(key, sputnik_dic[key])

workbook_summary_sputnik = xlsxwriter.Workbook(f'Сводка по 1, 2 ФАЗЕ теплоспутник {datetime.datetime.now().strftime("%d.%m.%Y")}.xlsx')

cell_format_green = workbook_summary_sputnik.add_format()
cell_format_green.set_bg_color('#98FB98')
cell_format_blue = workbook_summary_sputnik.add_format()
cell_format_blue.set_bg_color('#B0E0E6')
cell_format_hat = workbook_summary_sputnik.add_format()
cell_format_hat.set_bg_color('#F0E68C')
cell_format_date = workbook_summary_sputnik.add_format()
cell_format_date.set_font_size(font_size=14)

ws0 = workbook_summary_sputnik.add_worksheet('Кратчайшая сводка')
ws0.set_column(0, 0, 18)
ws0.set_column(1, 1, 18)
ws0.set_column(2, 5, 20)


for i, (one, two, three, four, five) in enumerate(small_summary, start=1):
    color_2 = cell_format_hat
    color_2.set_bold('bold')
    color_2.set_border(style=1)

    if one == 'Краткая справка':
        color = cell_format_hat
        color.set_bold('bold')
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass
    ws0.write(f'A{i}', one, color_2)
    ws0.write(f'B{i}', two, color)
    ws0.write(f'C{i}', three, color)
    ws0.write(f'D{i}', four, color)
    ws0.write(f'E{i}', five, color)



ws11 = workbook_summary_sputnik.add_worksheet('Сводка по спутникам')
ws11.set_column(0, 1, 38)
ws11.set_column(2, 4, 12)
ws11.set_column(5, 9, 22)
ws11.autofilter('A1:J2000')


for i, (one, two, three, four, five, six, seven, eight, nine, ten) in enumerate(summary_sputnik, start=1):
    if one == 'Чертеж по ГОСТ':
        color = cell_format_hat
        color.set_bold('bold')
    elif seven:
        color = cell_format_green
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass
    ws11.write(f'A{i}', one, color)
    ws11.write(f'B{i}', two, color)
    ws11.write(f'C{i}', three, color)
    ws11.write(f'D{i}', four, color)
    ws11.write(f'E{i}', five, color)
    ws11.write(f'F{i}', six, color)
    ws11.write(f'G{i}', seven, color)
    ws11.write(f'H{i}', eight, color)
    ws11.write(f'I{i}', nine, color)
    ws11.write(f'J{i}', ten, color)



workbook_summary_sputnik.close()

print('Done')

print('Файл по спутникам создан.')
