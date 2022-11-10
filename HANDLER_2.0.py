import csv
import openpyxl as xl
import re
import datetime
import xlsxwriter
import pandas as pd


# home laptop
# directory_dbs_files = r'C:\Users\vanik\PycharmProjects\handlers_sg\out_files_for_dbs\\'

# work laptop
directory_dbs_files = r'C:\Users\ignatenkoia\Desktop\work\GIT_PROJECTS\handlers_sg\dbs\\'



file_db_isotp = 'iso_tp_db.csv'


isotp_dic = {}
tp_dic = {}
iso_dic = {}

"""
Сводные списки для финальной записи в сводки по фазам.
"""
summary_iso_tp_phase_1 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                           'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
                           'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
                           'Статус уведомлений']]

summary_iso_tp_phase_2 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                           'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
                           'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
                           'Статус уведомлений']]

summary_iso_tp_phase_3 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                           'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
                           'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
                           'Статус уведомлений']]

summary_tp_phase_1 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                           'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]

summary_tp_phase_2 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                           'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]

summary_tp_phase_3 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
                           'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]


"""
Наполнение словарей по изометрии+тестпакет и по тестпакетам исходной подтверждённой информацией.
"""
with open(directory_dbs_files + file_db_isotp, 'r') as read_db:
    readed_db = csv.reader(read_db, delimiter=';')

    for row in readed_db:
        isotp_dic[row[0]] = ['', '', '', '', '', '', '', 0, '', '', '', '', '', '', '', '', '', '']
        tp_dic[row[2]] = ['', '', '', '', '', 0, '', '', '', '', '', '']
        iso_dic[row[1]] = ['', '', '', '', '', '']

with open(directory_dbs_files + file_db_isotp, 'r') as read_db:
    readed_db = csv.reader(read_db, delimiter=';')

    for row in readed_db:

        iso_with_tp = row[0].strip()
        isometric = row[1].strip()
        testpackage = row[2].strip()
        phase = row[3].strip()
        line = row[4].strip()
        title = row[5].strip()
        unit = row[6].strip()
        fluid = row[7].strip()
        ggn_status = row[8].strip()
        iso_length = float(row[9].strip().replace(',', '.'))
        rfi_erection = row[10].strip()
        rfi_test = row[11].strip()
        rfi_airblowing = row[12].strip()
        rfi_reinstatement = row[13].strip()
        type_ins = row[14].strip()
        volume_ins = row[15].strip()
        rfi_ins_cotton = row[16].strip()
        rfi_ins_metall = row[17].strip()
        rfi_ins_box = row[18].strip()

        isotp_dic[iso_with_tp] = [testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length,
                                  rfi_erection, rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins,
                                  rfi_ins_cotton, rfi_ins_metall, rfi_ins_box, '', '', phase]

        tp_dic[testpackage][0] = testpackage
        tp_dic[testpackage][1] = title
        tp_dic[testpackage][2] = unit
        tp_dic[testpackage][3] = fluid
        tp_dic[testpackage][4] = ggn_status
        tp_dic[testpackage][5] += iso_length
        tp_dic[testpackage][6] = rfi_erection
        tp_dic[testpackage][7] = rfi_test
        tp_dic[testpackage][8] = rfi_airblowing
        tp_dic[testpackage][9] = rfi_reinstatement
        tp_dic[testpackage][11] = phase

        iso_dic[isometric][5] = testpackage



# Проверка Журнал заявок АИС Р2 ФАЗА2-------------------------------------
df = pd.read_excel('Журнал заявок общий.xlsx')
df = df.sort_values(by='Дата подачи / Date of submission', ascending=True)
df.to_excel('Журнал заявок общий.xlsx', index=0)

print('Журнал заявок отсортирован по дате отработки инспекции.')

wb_journal_rfi = xl.load_workbook('Журнал заявок общий.xlsx')
sheet_journal_rfi = wb_journal_rfi['Sheet1']

replace_pattern_1 = ['-HT', '-VT', '-PT']
replace_pattern_2 = ['(T.T. REINSTATEMENT)', '(T.T. AIR BLOWING)', '(AIR BLOWING)', '(T.T AIR BLOWING)', '(T.T AIR BLOWING',
                     '(T.T. ERECTION)', '(T.T .ERECTION)', '(T.T.TEST)', '(T.T. AIR BLOWIHG)', '(T.T. TEST)',
                     '(T.T ERECTION)', '(T.T TEST)', '(T.T REINSTATEMENT)',
                      '(GPA AIR BLOWING)', '(GPA TEST)',
                     '(GPA ERECTION)', '(GPA REINSTATEMENT)', '(T.T. REISTATEMENT)', '(T.T.REINSTATEMENT)',
                     '(T.T RE-INSTATEMENT)', '( T.T AIR BLOWING )', '( T.T AIR BLOWING )',
                     '(T.T.ERECTION)', '(T.T.TEST)', '(T.T.AIR BLOWING)', '(T.T.REINSTATEMENT)']

res_summary = {} # ?????????

for i in sheet_journal_rfi['B2':'AO550000']:
    if i[0].value:
        rfi_number = str(i[1].value).strip()
        tp_number = str(i[2].value).strip()
        pkk = str(i[4].value).strip()

        description_rfi = str(i[16].value)
        violation = str(i[35].value)
        name_insp = str(i[26].value)
        list_iso = str(i[8].value).replace(' ', '').split(';')
        volume_m = re.sub(r'[^0-9.]', '', str(i[18].value))
        category_cancelled = str(i[31].value).strip()
        comment = str(i[39].value)

        for typo in replace_pattern_2:
            if typo in tp_number:
                tp_number = tp_number.replace(typo, '').strip()

        for typo in replace_pattern_1:
            if typo in tp_number:
                tp_number = tp_number.replace(typo, '').strip()
        tp_number = tp_number.strip()



        if 'Монтаж технологического трубопровода в рамках' in description_rfi:
            if 'Принято' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][6] = rfi_number
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][8] = rfi_number
            if 'Принято с замечаниями' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][6] = rfi_number + " ПЗ"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][8] = rfi_number + " ПЗ"
            if 'подтвержд' in comment or 'подтвржд' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][6] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][8] = rfi_number + " ФОП"
            if 'зафиксирован' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][6] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][8] = rfi_number + " ФОП"



        if 'испытаний на прочность и плотность' in description_rfi or\
                'испытаний технологического трубопровода  на прочность' in description_rfi or\
                'испытаний технологического трубопровода на прочность' in description_rfi:

            if 'Принято' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][7] = rfi_number
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number
            if 'Принято с замечаниями' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][7] = rfi_number + " ПЗ"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number + " ПЗ"
            if 'подтвержд' in comment or 'подтвржд' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][7] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number + " ФОП"
            if 'зафиксирован' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][7] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][9] = rfi_number + " ФОП"

        if 'испыт' and 'рочност' in description_rfi:
            if 'Принято' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][7] = rfi_number
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number
            if 'Принято с замечаниями' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][7] = rfi_number + " ПЗ"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number + " ПЗ"
            if 'подтвержд' in comment or 'подтвржд' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][7] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number + " ФОП"
            if 'зафиксирован' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][7] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric}-{tp_number}'][9] = rfi_number + " ФОП"




        if 'родувка' in description_rfi and 'еплоспутн' not in description_rfi:
            if 'Принято' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][8] = rfi_number
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][10] = rfi_number
            if 'Принято с замечаниями' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][8] = rfi_number + " ПЗ"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][10] = rfi_number + " ПЗ"
            if 'подтвержд' in comment or 'подтвржд' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][8] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][10] = rfi_number + " ФОП"
            if 'зафиксирован' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][8] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][10] = rfi_number + " ФОП"



        if 'сборки технологических трубопроводов в проект' in description_rfi or \
                'сборки технологических трубопроводов в рамках' in description_rfi:

            if 'Принято' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][9] = rfi_number
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][11] = rfi_number
            if 'Принято с замечаниями' == category_cancelled:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][9] = rfi_number + " ПЗ"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][11] = rfi_number + " ПЗ"
            if 'подтвержд' in comment or 'подтвржд' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][9] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][11] = rfi_number + " ФОП"
            if 'зафиксирован' in comment:
                if tp_number in tp_dic.keys():
                    tp_dic[tp_number][9] = rfi_number + " ФОП"
                if list_iso:
                    for isometric in list_iso:
                        if f'{isometric.strip()}-{tp_number}' in isotp_dic.keys():
                            isotp_dic[f'{isometric.strip()}-{tp_number}'][11] = rfi_number + " ФОП"


        if 'дополн' in description_rfi:
            if 'Принято' == category_cancelled:
                if list_iso:
                    for isometric in list_iso:
                        if isometric.strip() in iso_dic.keys():
                            iso_dic[isometric.strip()][3] = rfi_number


        # ИЗОЛЯЦИЯ ПРОВЕРКА------------------------------ ТРУБА
        if 'завершении работ по теплоизоляц' in pkk:
            if 'представлены не в полном объеме, представлены некорректные документы' in violation or \
                    'документы, подтверждающие качество работ' in violation:

                if 'крепежа и герметизации металлического кожуха согласно изометрическим' in description_rfi:
                    if list_iso:
                        for iso in list_iso:
                            if iso.strip() in iso_dic.keys():
                                iso_dic[iso.strip()][1] = rfi_number

                if 'теплоизоляционного покрытия согласно изометрическим' in description_rfi:
                    if list_iso:
                        for iso in list_iso:
                            if iso.strip() in iso_dic.keys():
                                iso_dic[iso.strip()][0] = rfi_number



            # КОРОБА/ЧЕХЛЫ ФЛАНЦЫ ЗРА

                if 'смонтированного теплоизоляционного покрытия и металлического кожуха металлических коробов на ' \
                   'фланцах и ЗРА' in description_rfi or 'Проверка качества' \
                                                         ' смонтированной теплоизоляционной оболочки ( термочехлов) ' \
                                                         'согласно изометрическим' in description_rfi or 'Проверка ' \
                         'качества смонтированного теплоизоляционного покрытия и металлического ' \
                             'кожуха металлических коробов, теплоизоляционной оболочки ' \
                                     '( термочехлов) на фланцах' in description_rfi:

                    if list_iso:
                        for iso in list_iso:
                            if iso.strip() in iso_dic.keys():
                                iso_dic[iso.strip()][2] = rfi_number





# -Проверка на уведомления-------------------------
wb_ncr = xl.load_workbook('Реестр уведомлений.xlsx')
sheet_ncr = wb_ncr['Предписания (Instructions)']
iso_ncr = {}

for i in sheet_ncr['B4':'V55000']:
    if i[0].value:
        number_ncr = str(i[0].value)
        mark_execution = str(i[16].value)
        notification_items = str(i[1].value)
        type_violation = str(i[5].value)
        content_remarks = str(i[6].value).replace(' ', '')
        content_remarks_iso = re.findall(
            r'\d-\d-\d-\d\d-\d\d\d-\w*-\d\w-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-\w*-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC3P\+-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC3\+-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC4P\+-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC5\+-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC4\+-\d\d-\d\d\d\d-\d\d\d',
            content_remarks.replace(' ', '').replace('\n', '').replace('Р', 'P').replace('С', 'C').strip())
        if 'Нет' in mark_execution:
            if content_remarks_iso:
                for iso in content_remarks_iso:
                    try:
                        iso_dic[iso][4] = number_ncr
                        tp_dic[iso_dic[iso][5]][10] = number_ncr
                    except:
                        print(f'Не нашел в словаре изометрию {iso}')
    else:
        break




"""
Блок подготовки к записи в сводный файл
"""

n_dic_3_30 = {'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSU': ['Поверхностная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOWWA': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNMP': ['Азот СД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_3_110 = {'HWBR': ['Вода котлового контура, обратная (Т21)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWBS': ['Вода котлового контура, прямая (Т11)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHSGA': ['Товарный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHSGAHP': ['Товарный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NODRA': ['Дренаж', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UABA': ['Барьерный воздух ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHGAL': ['Топливный газ НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWWW': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'WMMI': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_2_60 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC4P+': ['Бутановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3P+': ['Пропановая фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC5+': ['С5+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC4+': ['С4+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNMP': ['Азот СД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCS': ['Подача Оборотная вода(В4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCR': ['Возврат Оборотная вода(В5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOWWA': ['Сточные воды', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWPO': ['Питьевая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGAD': ['Сухой природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WMMI': ['Водометанольная смесь', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSU': ['Поверхностная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HYDV': ['Пары углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
              }

n_dic_2_70 = {'NODRAH': ['Дренаж углеводородов', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['Природный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['Технический воздух', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['Топливный газ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['Азот НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['Азот ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['Теплофикационная вода, обратная (Т2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['Теплофикационная вода, прямая (Т1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['Газ регенерации, прямой', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['Газ регенерации, обратный', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['Конденсат (Т8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['Дизельное топливо', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHLGPT': ['Очищенная ШФЛУ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['Факельный сброс в общую фак. систему', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['Сброс в атмосферу', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3+': ['С3+ фракция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['Пар НД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAIN': ['Воздух КИП', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['Техническая вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['Пожарная вода', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHGAH': ['Топливный газ ВД', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['Некондиция', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}


# isotp_dic = {}
# tp_dic = {}
# iso_dic = {}
#
# """
# Сводные списки для финальной записи в сводки по фазам.
# """
# summary_iso_tp_phase_1 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
#                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
#                            'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
#                            'Статус уведомлений']]
#
# summary_iso_tp_phase_2 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
#                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
#                            'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
#                            'Статус уведомлений']]
#
# summary_iso_tp_phase_3 = [['Тестпакет', 'Изометрия', 'Линия', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
#                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Тип изоляции',
#                            'Объём изоляции', 'RFI Мин.вата', 'RFI Металл. кожух', 'RFI Короб/чехол', 'RFI ДИГ',
#                            'Статус уведомлений']]
#
# summary_tp_phase_1 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
#                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]
#
# summary_tp_phase_2 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
#                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]
#
# summary_tp_phase_3 = [['Тестпакет', 'Титул', 'Установка', 'Среда', 'Статус ГГН', 'Длина',
#                            'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT', 'Статус уведомлений']]




for isotp in isotp_dic.keys():

    isotp_dic[isotp][14] = iso_dic[isotp_dic[isotp][1]][0]
    isotp_dic[isotp][15] = iso_dic[isotp_dic[isotp][1]][1]
    isotp_dic[isotp][16] = iso_dic[isotp_dic[isotp][1]][2]
    isotp_dic[isotp][17] = iso_dic[isotp_dic[isotp][1]][3]
    isotp_dic[isotp][18] = iso_dic[isotp_dic[isotp][1]][4]


    if isotp_dic[isotp][19] == '1' and 'SUPER' not in isotp_dic[isotp][0]:
        summary_iso_tp_phase_1.append([isotp_dic[isotp][0], isotp_dic[isotp][1], isotp_dic[isotp][2],
                                       isotp_dic[isotp][3], isotp_dic[isotp][4], isotp_dic[isotp][5],
                                       isotp_dic[isotp][6], isotp_dic[isotp][7], isotp_dic[isotp][8],
                                       isotp_dic[isotp][9], isotp_dic[isotp][10], isotp_dic[isotp][11],
                                       isotp_dic[isotp][12], isotp_dic[isotp][13], isotp_dic[isotp][14],
                                       isotp_dic[isotp][15], isotp_dic[isotp][16], isotp_dic[isotp][17],
                                       isotp_dic[isotp][18]
                                       ])

    if isotp_dic[isotp][19] == '2':
        summary_iso_tp_phase_2.append([isotp_dic[isotp][0], isotp_dic[isotp][1], isotp_dic[isotp][2],
                                       isotp_dic[isotp][3], isotp_dic[isotp][4], isotp_dic[isotp][5],
                                       isotp_dic[isotp][6], isotp_dic[isotp][7], isotp_dic[isotp][8],
                                       isotp_dic[isotp][9], isotp_dic[isotp][10], isotp_dic[isotp][11],
                                       isotp_dic[isotp][12], isotp_dic[isotp][13], isotp_dic[isotp][14],
                                       isotp_dic[isotp][15], isotp_dic[isotp][16], isotp_dic[isotp][17],
                                       isotp_dic[isotp][18]
                                       ])

    if isotp_dic[isotp][19] == '3':
        summary_iso_tp_phase_3.append([isotp_dic[isotp][0], isotp_dic[isotp][1], isotp_dic[isotp][2],
                                       isotp_dic[isotp][3], isotp_dic[isotp][4], isotp_dic[isotp][5],
                                       isotp_dic[isotp][6], isotp_dic[isotp][7], isotp_dic[isotp][8],
                                       isotp_dic[isotp][9], isotp_dic[isotp][10], isotp_dic[isotp][11],
                                       isotp_dic[isotp][12], isotp_dic[isotp][13], isotp_dic[isotp][14],
                                       isotp_dic[isotp][15], isotp_dic[isotp][16], isotp_dic[isotp][17],
                                       isotp_dic[isotp][18]
                                       ])


# isotp_dic[iso_with_tp] = [
# testpackage, 0
# isometric,1
# line, 2
# title, 3
# unit, 4
# fluid, 5
# ggn_status,6
# iso_length,7
# rfi_erection, 8
# rfi_test, 9
# rfi_airblowing, 10
# rfi_reinstatement, 11
# type_ins, 12
# volume_ins 13
# rfi_ins_cotton, 14
# rfi_ins_metall, 15
# rfi_ins_box, 16
# rfi dig 17
# ncr 18
# phase 19
#
#         tp_dic[testpackage][0] = testpackage
#         tp_dic[testpackage][1] = title
#         tp_dic[testpackage][2] = unit
#         tp_dic[testpackage][3] = fluid
#         tp_dic[testpackage][4] = ggn_status
#         tp_dic[testpackage][5] += iso_length
#         tp_dic[testpackage][6] = rfi_erection
#         tp_dic[testpackage][7] = rfi_test
#         tp_dic[testpackage][8] = rfi_airblowing
#         tp_dic[testpackage][9] = rfi_reinstatement
#         tp_dic[testpackage][10] = ncr_status
#         tp_dic[testpackage][11] = phase


for tp in tp_dic.keys():

    if tp_dic[tp][11] == '1' and 'SUPER' not in tp_dic[tp][0]:
        summary_tp_phase_1.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                   tp_dic[tp][5], tp_dic[tp][6],tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                   tp_dic[tp][10]])

    if tp_dic[tp][11] == '2':
        summary_tp_phase_2.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                   tp_dic[tp][5], tp_dic[tp][6],tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                   tp_dic[tp][10]])

    if tp_dic[tp][11] == '3':
        summary_tp_phase_3.append([tp_dic[tp][0], tp_dic[tp][1], tp_dic[tp][2], tp_dic[tp][3], tp_dic[tp][4],
                                   tp_dic[tp][5], tp_dic[tp][6],tp_dic[tp][7], tp_dic[tp][8], tp_dic[tp][9],
                                   tp_dic[tp][10]])






workbook_summary = xlsxwriter.Workbook(f'Сводка по ФАЗАМ на {datetime.datetime.now().strftime("%d.%m.%Y")}.xlsx')


cell_format_green = workbook_summary.add_format()
cell_format_green.set_bg_color('#98FB98')
cell_format_blue = workbook_summary.add_format()
cell_format_blue.set_bg_color('#B0E0E6')
cell_format_hat = workbook_summary.add_format()
cell_format_hat.set_bg_color('#FFDAB9')
cell_format_date = workbook_summary.add_format()
cell_format_date.set_font_size(font_size=14)


# -------------------------------------------------------------------------------------------------
ws5 = workbook_summary.add_worksheet('Сводка ТП ФАЗА 1')

ws5.set_column(0, 0, 30)
ws5.set_column(1, 5, 15)
ws5.set_column(6, 11, 22)
ws5.set_column(12, 17, 25)
ws5.set_column(18, 18, 13)
ws5.autofilter('A1:S1682')

for i, (testpackage, title, unit, fluid, ggn_status, iso_length, rfi_erection, rfi_test,
        rfi_airblowing, rfi_reinstatement, ncr_status) in enumerate(summary_tp_phase_1, start=1):

    if testpackage == 'Тестпакет':
        color = cell_format_hat
        color.set_bold('bold')
    elif '-CC' in rfi_reinstatement:
        color = cell_format_green
    else:
        color = cell_format_blue

    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass

    ws5.write(f'A{i}', testpackage, color)
    ws5.write(f'B{i}', title, color)
    ws5.write(f'C{i}', unit, color)
    ws5.write(f'D{i}', fluid, color)
    ws5.write(f'E{i}', ggn_status, color)
    ws5.write(f'F{i}', iso_length, color)
    ws5.write(f'G{i}', rfi_erection, color)
    ws5.write(f'H{i}', rfi_test, color)
    ws5.write(f'I{i}', rfi_airblowing, color)
    ws5.write(f'J{i}', rfi_reinstatement, color)
    ws5.write(f'K{i}', ncr_status, color)



# -------------------------------------------------------------------------------------------------
ws1 = workbook_summary.add_worksheet('Сводка ТП ФАЗА 2')

ws1.set_column(0, 0, 30)
ws1.set_column(1, 5, 15)
ws1.set_column(6, 11, 22)
ws1.set_column(12, 17, 25)
ws1.set_column(18, 18, 13)
ws1.autofilter('A1:S1682')

for i, (testpackage, title, unit, fluid, ggn_status, iso_length, rfi_erection, rfi_test,
        rfi_airblowing, rfi_reinstatement, ncr_status) in enumerate(summary_tp_phase_2, start=1):

    if testpackage == 'Тестпакет':
        color = cell_format_hat
        color.set_bold('bold')
    elif '-CC' in rfi_reinstatement:
        color = cell_format_green
    else:
        color = cell_format_blue

    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass

    ws1.write(f'A{i}', testpackage, color)
    ws1.write(f'B{i}', title, color)
    ws1.write(f'C{i}', unit, color)
    ws1.write(f'D{i}', fluid, color)
    ws1.write(f'E{i}', ggn_status, color)
    ws1.write(f'F{i}', iso_length, color)
    ws1.write(f'G{i}', rfi_erection, color)
    ws1.write(f'H{i}', rfi_test, color)
    ws1.write(f'I{i}', rfi_airblowing, color)
    ws1.write(f'J{i}', rfi_reinstatement, color)
    ws1.write(f'K{i}', ncr_status, color)


# -------------------------------------------------------------------------------------------------
ws2 = workbook_summary.add_worksheet('Сводка ТП ФАЗА 3')

ws2.set_column(0, 0, 30)
ws2.set_column(1, 5, 15)
ws2.set_column(6, 11, 22)
ws2.set_column(12, 17, 25)
ws2.set_column(18, 18, 13)
ws2.autofilter('A1:S1682')

for i, (testpackage, title, unit, fluid, ggn_status, iso_length, rfi_erection, rfi_test,
        rfi_airblowing, rfi_reinstatement, ncr_status) in enumerate(summary_tp_phase_3, start=1):

    if testpackage == 'Тестпакет':
        color = cell_format_hat
        color.set_bold('bold')
    elif '-CC' in rfi_reinstatement:
        color = cell_format_green
    else:
        color = cell_format_blue

    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass

    ws2.write(f'A{i}', testpackage, color)
    ws2.write(f'B{i}', title, color)
    ws2.write(f'C{i}', unit, color)
    ws2.write(f'D{i}', fluid, color)
    ws2.write(f'E{i}', ggn_status, color)
    ws2.write(f'F{i}', iso_length, color)
    ws2.write(f'G{i}', rfi_erection, color)
    ws2.write(f'H{i}', rfi_test, color)
    ws2.write(f'I{i}', rfi_airblowing, color)
    ws2.write(f'J{i}', rfi_reinstatement, color)
    ws2.write(f'K{i}', ncr_status, color)






# Сводка по изометриям

wsi1 = workbook_summary.add_worksheet('Сводка по ИЗО ФАЗА 1')
wsi1.set_column(0, 1, 32)
wsi1.set_column(2, 7, 14)
wsi1.set_column(8, 11, 20)

wsi1.set_column(12, 13, 13)
wsi1.set_column(14, 16, 20)
wsi1.set_column(17, 18, 14)

cell_format_ins = workbook_summary.add_format()
cell_format_ins.set_bg_color('#FFEBCD')

wsi1.autofilter('A1:S20000')
for i, (testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length, rfi_erection,
        rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins, rfi_ins_cotton,
        rfi_ins_metall, rfi_ins_box, rfi_dig, ncr) in enumerate(summary_iso_tp_phase_1, start=1):

    if testpackage == 'Тестпакет':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color_2 = cell_format_hat
        color_2.set_bold('bold')
        color_2.set_text_wrap(text_wrap=1)

    elif rfi_reinstatement:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)

    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)

    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    wsi1.write(f'A{i}', testpackage, color)
    wsi1.write(f'B{i}', isometric, color)
    wsi1.write(f'C{i}', line, color)
    wsi1.write(f'D{i}', title, color)
    wsi1.write(f'E{i}', unit, color)
    wsi1.write(f'F{i}', fluid, color)
    wsi1.write(f'G{i}', ggn_status, color)
    wsi1.write(f'H{i}', iso_length, color)
    wsi1.write(f'I{i}', rfi_erection, color)
    wsi1.write(f'J{i}', rfi_test, color)
    wsi1.write(f'K{i}', rfi_airblowing, color)
    wsi1.write(f'L{i}', rfi_reinstatement, color)
    wsi1.write(f'M{i}', type_ins, color_2)
    wsi1.write(f'N{i}', volume_ins, color_2)
    wsi1.write(f'O{i}', rfi_ins_cotton, color_2)
    wsi1.write(f'P{i}', rfi_ins_metall, color_2)
    wsi1.write(f'Q{i}', rfi_ins_box, color_2)
    wsi1.write(f'R{i}', rfi_dig, color)
    wsi1.write(f'S{i}', ncr, color)


wsi2 = workbook_summary.add_worksheet('Сводка по ИЗО ФАЗА 2')
wsi2.set_column(0, 1, 32)
wsi2.set_column(2, 7, 14)
wsi2.set_column(8, 11, 20)

wsi2.set_column(12, 13, 13)
wsi2.set_column(14, 16, 20)
wsi2.set_column(17, 18, 14)

cell_format_ins = workbook_summary.add_format()
cell_format_ins.set_bg_color('#FFEBCD')

wsi2.autofilter('A1:S20000')
for i, (testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length, rfi_erection,
        rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins, rfi_ins_cotton,
        rfi_ins_metall, rfi_ins_box, rfi_dig, ncr) in enumerate(summary_iso_tp_phase_2, start=1):

    if testpackage == 'Тестпакет':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color_2 = cell_format_hat
        color_2.set_bold('bold')
        color_2.set_text_wrap(text_wrap=1)

    elif rfi_reinstatement:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)

    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)

    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    wsi2.write(f'A{i}', testpackage, color)
    wsi2.write(f'B{i}', isometric, color)
    wsi2.write(f'C{i}', line, color)
    wsi2.write(f'D{i}', title, color)
    wsi2.write(f'E{i}', unit, color)
    wsi2.write(f'F{i}', fluid, color)
    wsi2.write(f'G{i}', ggn_status, color)
    wsi2.write(f'H{i}', iso_length, color)
    wsi2.write(f'I{i}', rfi_erection, color)
    wsi2.write(f'J{i}', rfi_test, color)
    wsi2.write(f'K{i}', rfi_airblowing, color)
    wsi2.write(f'L{i}', rfi_reinstatement, color)
    wsi2.write(f'M{i}', type_ins, color_2)
    wsi2.write(f'N{i}', volume_ins, color_2)
    wsi2.write(f'O{i}', rfi_ins_cotton, color_2)
    wsi2.write(f'P{i}', rfi_ins_metall, color_2)
    wsi2.write(f'Q{i}', rfi_ins_box, color_2)
    wsi2.write(f'R{i}', rfi_dig, color)
    wsi2.write(f'S{i}', ncr, color)


wsi3 = workbook_summary.add_worksheet('Сводка по ИЗО ФАЗА 3')
wsi3.set_column(0, 1, 32)
wsi3.set_column(2, 7, 14)
wsi3.set_column(8, 11, 20)

wsi3.set_column(12, 13, 13)
wsi3.set_column(14, 16, 20)
wsi3.set_column(17, 18, 14)

cell_format_ins = workbook_summary.add_format()
cell_format_ins.set_bg_color('#FFEBCD')

wsi3.autofilter('A1:S20000')
for i, (testpackage, isometric, line, title, unit, fluid, ggn_status, iso_length, rfi_erection,
        rfi_test, rfi_airblowing, rfi_reinstatement, type_ins, volume_ins, rfi_ins_cotton,
        rfi_ins_metall, rfi_ins_box, rfi_dig, ncr) in enumerate(summary_iso_tp_phase_3, start=1):

    if testpackage == 'Тестпакет':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color_2 = cell_format_hat
        color_2.set_bold('bold')
        color_2.set_text_wrap(text_wrap=1)

    elif rfi_reinstatement:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)

    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)

    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    wsi3.write(f'A{i}', testpackage, color)
    wsi3.write(f'B{i}', isometric, color)
    wsi3.write(f'C{i}', line, color)
    wsi3.write(f'D{i}', title, color)
    wsi3.write(f'E{i}', unit, color)
    wsi3.write(f'F{i}', fluid, color)
    wsi3.write(f'G{i}', ggn_status, color)
    wsi3.write(f'H{i}', iso_length, color)
    wsi3.write(f'I{i}', rfi_erection, color)
    wsi3.write(f'J{i}', rfi_test, color)
    wsi3.write(f'K{i}', rfi_airblowing, color)
    wsi3.write(f'L{i}', rfi_reinstatement, color)
    wsi3.write(f'M{i}', type_ins, color_2)
    wsi3.write(f'N{i}', volume_ins, color_2)
    wsi3.write(f'O{i}', rfi_ins_cotton, color_2)
    wsi3.write(f'P{i}', rfi_ins_metall, color_2)
    wsi3.write(f'Q{i}', rfi_ins_box, color_2)
    wsi3.write(f'R{i}', rfi_dig, color)
    wsi3.write(f'S{i}', ncr, color)


workbook_summary.close()

print('Всё записал.')