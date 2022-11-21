# -*- coding: cp1251 -*-

import openpyxl as xl
import re
import xlsxwriter
from dateutil import parser
import datetime, time
import pandas as pd
from playsound import playsound


# �������� ������ ������� ��-------------------------------------------
wb_phase2 = xl.load_workbook('�� �� ���� 1, 2.xlsx')
sheet_phase2_TP = wb_phase2['������� ����������']
sheet_phase2_ISO = wb_phase2['TP']

list_categories_ae = ['��-I/ Ba-I', '�(�)-I', 'A(�)-I / A(b)-I', '�(�)-I / B(a)-I', 'B-I / C-I', '�(�)-I / B(b)-I',
                      '�-I', '�(�)-I/B(a)-I', 'A(�)-II / A(b)-II', '��-I/Ba-I', '��-II/ Ba-II', '��-I/ Bb-I',
                      'B-I/ V-I', 'B-I/V-I', 'A(�)-I/A(b)-I', '�(�)-I', '�(�)-I/B(b)-I', 'B-I/C-I', '��-II/Ba-II',
                      '��-I/Bb-I']

testpackages = {}
testpackages_ae = {}
for l in sheet_phase2_TP['A5':'K5000']:
    if l[0].value:
        tp_short_code_BD = str(l[0].value).strip()
        title = str(l[1].value).strip()
        fluid = str(l[3].value).strip()
        category_tp = str(l[7].value)
        vt_tp = str(l[4].value)
        try:
            length = round(float(str(l[5].value)), 3)
        except:
            pass
        testpackages[tp_short_code_BD] = [title, fluid, category_tp, length]

        if category_tp in list_categories_ae and '����������' not in vt_tp:
            testpackages_ae[tp_short_code_BD] = 1
        else:
            testpackages_ae[tp_short_code_BD] = 0
    else:
        break

print(f'{len(testpackages.keys())} �����������')

isotpdic = {}
sc_isotpdic = {}

for i in sheet_phase2_ISO['A3':'K12117']:
    if i[0].value:
        sc_iso_tp = str(i[1].value).strip()
        isometric = str(i[2].value).strip()
        testpack = str(i[0].value).strip()
        isolength_max = round(float(str(i[6].value)), 3)
        isolength = round(float(str(i[7].value)), 3)
        iso_type_ins = str(i[8].value).strip()
        iso_ins_volume = float(str(i[9].value))
        title_iso = str(i[3].value)
        isotpdic[isometric] = [testpack, isolength_max, '', '', '', '', iso_type_ins, iso_ins_volume, '', '', '', '',
                               '']
        sc_isotpdic[sc_iso_tp] = [isometric, testpack, isolength, '', '', '', '', title_iso]
    else:
        break

print(f'����� ������� ������. \n ���������� ��������� - {len(isotpdic.keys())} '
      f'\n ����������� ��������� - {len(sc_isotpdic.keys()) - len(isotpdic.keys())}')
# ����� ������� ������--------------------------------------------------


for key in testpackages.keys():
    testpackages[key].append('n/p')
    testpackages[key].append('n/p')
    testpackages[key].append('n/p')


# ------------------------------------------------------------
for key in testpackages.keys():
    testpackages[key].append('')
    testpackages[key].append(0)
    testpackages[key].append('')
    testpackages[key].append('')
    testpackages[key].append('')
    testpackages[key].append('')
# ------------------------------------------------------------------


# �������� ������ ������ ��� �2 ����2-------------------------------------
df = pd.read_excel('������ ������ �����.xlsx')
df = df.sort_values(by='���� ������ / Date of submission', ascending=True)
df.to_excel('������ ������ �����.xlsx', index=0)

wb_journal_rfi = xl.load_workbook('������ ������ �����.xlsx')
sheet = wb_journal_rfi['Sheet1']

replace_pattern_1 = ['-HT', '-VT', '-PT']
replace_pattern_2 = ['(T.T. REINSTATEMENT)', '(T.T. AIR BLOWING)', '(AIR BLOWING)', '(T.T AIR BLOWING',
                     '(T.T. ERECTION)', '(T.T.TEST)', '(T.T. TEST)',
                     '(T.T ERECTION)', '(T.T TEST)', '(T.T REINSTATEMENT)',
                     '(T.T AIR BLOWING)', '(GPA AIR BLOWING)', '(GPA TEST)',
                     '(GPA ERECTION)', '(GPA REINSTATEMENT)', '(T.T. REISTATEMENT)', '(T.T.REINSTATEMENT)',
                     '(T.T RE-INSTATEMENT)', '( T.T AIR BLOWING )', '( T.T AIR BLOWING )',
                     '(T.T.ERECTION)', '(T.T.TEST)', '(T.T.AIR BLOWING)', '(T.T.REINSTATEMENT)']
res_summary = {}
for i in sheet['B2':'AO550000']:
    if i[0].value:
        rfi_number = str(i[1].value)
        tp_number = str(i[2].value)
        pkk = str(i[4].value)
        tp_shortname = ''
        tp_shortname_1 = ''
        for l in replace_pattern_2:
            if l in tp_number:
                tp_shortname_1 = tp_number.replace(l, '').strip()
            else:
                pass
        if '-HT' in tp_shortname_1:
            tp_shortname = tp_shortname_1.replace('-HT', '')
        elif '-PT' in tp_shortname_1:
            tp_shortname = tp_shortname_1.replace('-PT', '')
        elif '-VT' in tp_shortname_1:
            tp_shortname = tp_shortname_1.replace('-VT', '')
        else:
            tp_shortname = tp_shortname_1

        description_rfi = str(i[16].value)
        violation = str(i[35].value)
        name_insp = str(i[26].value)
        list_iso = str(i[8].value)
        volume_meter = re.sub(r'[^0-9.]', '', str(i[18].value))
        category_cancelled = str(i[31].value)
        comment = str(i[39].value)


        if tp_shortname in testpackages.keys():
            if '�������' in category_cancelled:
                if '������� ������ ��������������� ������������� ���' in description_rfi:
                    testpackages[tp_shortname][11] = rfi_number
                if '����� ���������������� ������������ ���' in description_rfi:
                    testpackages[tp_shortname][9] = rfi_number
                if '��������� ���������������� ������������ ���' in description_rfi:
                    testpackages[tp_shortname][10] = rfi_number
                if '������ ���������������� ������������ � ������' in description_rfi:
                    testpackages[tp_shortname][9] = rfi_number
                if '��������� �� ��������� � ���������' in description_rfi:
                    testpackages[tp_shortname][10] = rfi_number
                if '��������� ���������������� ������������ �� ���������' in description_rfi:
                    testpackages[tp_shortname][10] = rfi_number
                if '��������� ���������������� ������������  �� ���������' in description_rfi:
                    testpackages[tp_shortname][10] = rfi_number
                if '������ ��������������� ������������� � ������' in description_rfi:
                    testpackages[tp_shortname][11] = rfi_number
                if '��������' in description_rfi:
                    testpackages[tp_shortname][12] = rfi_number
            else:
                if '���������' in comment:
                    if '������ ��������������� ������������� ���' in description_rfi:
                        testpackages[tp_shortname][11] = rfi_number + ' ���'
                    if '����� ���������������� ������������ ���' in description_rfi:
                        testpackages[tp_shortname][9] = rfi_number + ' ���'
                    if '��������� ���������������� ������������ ���' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ���'
                    if '������ ���������������� ������������ � ������' in description_rfi:
                        testpackages[tp_shortname][9] = rfi_number + ' ���'
                    if '��������� �� ��������� � ���������' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ���'
                    if '��������� ���������������� ������������ �� ���������' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number
                    if '��������� ���������������� ������������  �� ���������' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ���'
                    if '������ ��������������� ������������� � ������' in description_rfi:
                        testpackages[tp_shortname][11] = rfi_number + ' ���'
                    if '�������' in description_rfi:
                        testpackages[tp_shortname][12] = rfi_number + ' ���'
                if '������������' in comment:
                    if '��������� ���������������� ������������ ���' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ���'
                    if '������ ���������������� ������������ � ������' in description_rfi:
                        testpackages[tp_shortname][9] = rfi_number + ' ���'
                    if '��������� �� ��������� � ���������' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ���'
                    if '��������� ���������������� ������������ �� ���������' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number
                    if '��������� ���������������� ������������  �� ���������' in description_rfi:
                        testpackages[tp_shortname][10] = rfi_number + ' ���'
                    if '������ ��������������� ������������� � ������' in description_rfi:
                        testpackages[tp_shortname][11] = rfi_number + ' ���'
                    if '��������' in description_rfi:
                        testpackages[tp_shortname][12] = rfi_number + ' ���'

        for isom in list_iso.split(';'):
            if rfi_number == 'CPECC-CC-64685/1':
                isom = isom.replace(' �. 0', '').replace(' �. 1', '').replace(' �. 3', '')

            if isom.strip() in isotpdic.keys():
                if '�������' in category_cancelled:
                    if '������ ���������������� ������������ � ������' in description_rfi:
                        isotpdic[isom.strip()][2] = rfi_number
                    if '��������� �� ��������� � ���������' in description_rfi:
                        isotpdic[isom.strip()][3] = rfi_number
                    if '��������� ���������������� ������������  �� ���������' in description_rfi:
                        isotpdic[isom.strip()][3] = rfi_number
                    if '������ ��������������� ������������� � ������' in description_rfi:
                        isotpdic[isom.strip()][5] = rfi_number
                    if '�������' in description_rfi and '���������' not in description_rfi:
                        isotpdic[isom.strip()][4] = rfi_number
                    if '�������������� ���������' in description_rfi:
                        isotpdic[isom.strip()][11] = rfi_number
                        isotpdic[isom.strip()][10] = '�������'
                else:
                    if '���������' in comment or '��������' in comment:
                        if '������ ��������������� ������������� ���' in description_rfi:
                            isotpdic[isom.strip()][5] = rfi_number + ' ���'
                        if '����� ���������������� ������������ ���' in description_rfi:
                            isotpdic[isom.strip()][2] = rfi_number + ' ���'
                        if '��������� ���������������� ������������ ���' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ���'
                        if '������ ���������������� ������������ � ������' in description_rfi:
                            isotpdic[isom.strip()][2] = rfi_number + ' ���'
                        if '��������� �� ��������� � ���������' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ���'
                        if '��������� ���������������� ������������  �� ���������' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ���'
                        if '������ ��������������� ������������� � ������' in description_rfi:
                            isotpdic[isom.strip()][5] = rfi_number + ' ���'
                        if '�������' in description_rfi and '���������' not in description_rfi:
                            isotpdic[isom.strip()][4] = rfi_number + ' ���'
                        if '�������' in description_rfi:
                            isotpdic[isom.strip()][12] = rfi_number + ' ���'
                    if '������������' in comment:
                        if '��������� ���������������� ������������ ���' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ���'
                        if '������ ���������������� ������������ � ������' in description_rfi:
                            isotpdic[isom.strip()][2] = rfi_number + ' ���'
                        if '��������� �� ��������� � ���������' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ���'
                        if '��������� ���������������� ������������  �� ���������' in description_rfi:
                            isotpdic[isom.strip()][3] = rfi_number + ' ���'
                        if '������ ��������������� ������������� � ������' in description_rfi:
                            isotpdic[isom.strip()][5] = rfi_number + ' ���'
                        if '�������' in description_rfi and '���������' not in description_rfi:
                            isotpdic[isom.strip()][4] = rfi_number + ' ���'

            # �������� ��������� ������ � ��������. ����������� ������� ����������� ���������.
            if tp_shortname + isom.strip() in sc_isotpdic.keys():
                if '�������' in category_cancelled:
                    if '������ ���������������� ������������ � ������' in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][3] = rfi_number
                    if '��������� �� ��������� � ���������' in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number
                    if '��������� ���������������� ������������  �� ���������' in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number
                    if '������ ��������������� ������������� � ������' in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][6] = rfi_number
                    if '�������' in description_rfi and '���������' not in description_rfi:
                        sc_isotpdic[tp_shortname + isom.strip()][5] = rfi_number
                else:
                    if '���������' or '��������' in comment:
                        if '������ ���������������� ������������ � ������' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][3] = rfi_number + ' ���'
                        if '��������� �� ��������� � ���������' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number + ' ���'
                        if '��������� ���������������� ������������  �� ���������' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number + ' ���'
                        if '������ ��������������� ������������� � ������' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][6] = rfi_number + ' ���'
                        if '�������' in description_rfi and '���������' not in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][5] = rfi_number + ' ���'
                    if '������������' in comment:
                        if '������ ���������������� ������������ � ������' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][3] = rfi_number + ' ���'
                        if '��������� �� ��������� � ���������' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number + ' ���'
                        if '��������� ���������������� ������������  �� ���������' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][4] = rfi_number + ' ���'
                        if '������ ��������������� ������������� � ������' in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][6] = rfi_number + ' ���'
                        if '�������' in description_rfi and '���������' not in description_rfi:
                            sc_isotpdic[tp_shortname + isom.strip()][5] = rfi_number + ' ���'

        if '75530/2' or '75823' in rfi_number:
            violation = '������������ �� � ������ ������, ������������ ������������ ���������'

        # �������� ��������------------------------------
        if '���������� ����� �� �����������' in pkk:
            if '������������ �� � ������ ������, ������������ ������������ ���������' in violation:
                if '�������������� ������ �������� ��������������' in description_rfi:
                    for iso in list_iso.split(';'):
                        try:
                            isotpdic[iso.strip()][9] = rfi_number
                        except:
                            pass
                if '������������������ �������� �������� ��������������' in description_rfi:
                    for iso in list_iso.split(';'):
                        try:
                            isotpdic[iso.strip()][8] = rfi_number
                        except:
                            pass
            if '���������, �������������� �������� �����' in violation:
                if '���������' in comment:
                    if '�������������� ������ �������� ��������������' in description_rfi:
                        for iso in list_iso.split(';'):
                            try:
                                isotpdic[iso.strip()][9] = rfi_number
                            except:
                                pass
                    if '������������������ �������� �������� ��������������' in description_rfi:
                        for iso in list_iso.split(';'):
                            try:
                                isotpdic[iso.strip()][8] = rfi_number
                            except:
                                pass
    else:
        break

print('������ ������ ��������. ���������� ���������.')
wb_journal_rfi.close()
# ------------------------------------------------------------------------
# -�������� �� �����������-------------------------
wb_ncr = xl.load_workbook('������ �����������.xlsx')
sheet_ncr = wb_ncr['����������� (Instructions)']
iso_ncr = {}
iso_ncr_iso = {}
for i in sheet_ncr['B4':'V55000']:
    if i[0].value:
        number_ncr = str(i[0].value)
        mark_execution = str(i[16].value)
        notification_items = str(i[1].value)
        type_violation = str(i[5].value)
        content_remarks = str(i[6].value).replace(' ', '')
        content_remarks_iso = re.findall(
            r'\d-\d-\d-\d\d-\d\d\d-\w*-\d\w-\d\d\d\d-\d\d\d|\d-\d-\d-\d\d-\d\d\d-\w*-\d\d-\d\d\d\d-\d\d\d|\d-\d-\d-\d\d-\d\d\d-NHC3P\+-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC3\+-\d\d-\d\d\d\d-\d\d\d|\d-\d-\d-\d\d-\d\d\d-NHC4P\+-\d\d-\d\d\d\d-\d\d\d|\d-\d-\d-\d\d-\d\d\d-NHC5\+-\d\d-\d\d\d\d-\d\d\d|'
            r'\d-\d-\d-\d\d-\d\d\d-NHC4\+-\d\d-\d\d\d\d-\d\d\d',
            content_remarks.replace(' ', '').replace('\n', '').replace('�', 'P').replace('�', 'C').strip())
        if '���' in mark_execution:
            if content_remarks_iso:
                for l in content_remarks_iso:
                    try:
                        iso_ncr_iso[l] = number_ncr
                        iso_ncr[isotpdic[l][0]] = number_ncr
                    except:
                        pass
    else:
        break
for key in testpackages.keys():
    if key in iso_ncr.keys():
        testpackages[key].append(iso_ncr[key])
    else:
        testpackages[key].append('����������� ���')
wb_ncr.close()
print('������ ����������� ��������. ���������� ���������.')
# --------------------------------------------------------

n_dic_3_110 = {'HWBR': ['���� ��������� �������, �������� (�21)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWBS': ['���� ��������� �������, ������ (�11)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWRP': ['���������������� ����, �������� (�2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'HWSP': ['���������������� ����, ������ (�1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHNGA': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHNGAD': ['����� ��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHRGR': ['��� �����������, ��������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHRGS': ['��� �����������, ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHSGA': ['�������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NHSGAHP': ['�������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NODRA': ['������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NOVENA': ['����� � ���������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UABA': ['��������� ������ ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UAIN': ['������ ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UAPL': ['����������� ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UFGAW': ['��������� ����� � ����� ���. �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHG': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHGAH': ['��������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UHGAL': ['��������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UNHP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UNLP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'USLP': ['��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWFF': ['�������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWSW': ['����������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'UWWW': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'WMMI': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
               'NODRAH': ['������ �������������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_3_110_a = {'HWBR': ['���� ��������� �������, �������� (�21)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWBS': ['���� ��������� �������, ������ (�11)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWRP': ['���������������� ����, �������� (�2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWSP': ['���������������� ����, ������ (�1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHNGA': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHNGAD': ['����� ��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHRGR': ['��� �����������, ��������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHRGS': ['��� �����������, ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHSGA': ['�������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHSGAHP': ['�������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NODRA': ['������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NOVENA': ['����� � ���������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UABA': ['��������� ������ ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UAIN': ['������ ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UAPL': ['����������� ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UFGAW': ['��������� ����� � ����� ���. �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHG': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHGAH': ['��������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHGAL': ['��������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UNHP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UNLP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'USLP': ['��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWFF': ['�������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWSW': ['����������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWWW': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'WMMI': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NODRAH': ['������ �������������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}
n_dic_3_110_b = {'HWBR': ['���� ��������� �������, �������� (�21)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWBS': ['���� ��������� �������, ������ (�11)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWRP': ['���������������� ����, �������� (�2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'HWSP': ['���������������� ����, ������ (�1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHNGA': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHNGAD': ['����� ��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHRGR': ['��� �����������, ��������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHRGS': ['��� �����������, ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHSGA': ['�������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NHSGAHP': ['�������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NODRA': ['������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NOVENA': ['����� � ���������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UABA': ['��������� ������ ', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UAIN': ['������ ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UAPL': ['����������� ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UFGAW': ['��������� ����� � ����� ���. �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHG': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHGAH': ['��������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UHGAL': ['��������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UNHP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UNLP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'USLP': ['��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWFF': ['�������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWSW': ['����������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'UWWW': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'WMMI': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                 'NODRAH': ['������ �������������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_3_30 = {'UAIN': ['������ ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGAD': ['����� ��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['����� � ���������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['��� �����������, ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['��� �����������, ��������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['�������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['��������� (�8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['��������� ����� � ����� ���. �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['���������������� ����, ������ (�1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['���������������� ����, �������� (�2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WMMI': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NODRAH': ['������ �������������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCS': ['������ ��������� ����(�4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCR': ['������� ��������� ����(�5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['����������� ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['����������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWPO': ['�������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSU': ['������������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOWWA': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['��������� �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['����������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNMP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_2_60 = {'NODRAH': ['������ �������������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC4P+': ['��������� �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3P+': ['���������� �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['��������� ����� � ����� ���. �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC5+': ['�5+ �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3+': ['�3+ �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC4+': ['�4+ �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHLGPT': ['��������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNMP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCS': ['������ ��������� ����(�4)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WPCR': ['������� ��������� ����(�5)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['��������� (�8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAIN': ['������ ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['����������� ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['����������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['���������������� ����, ������ (�1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['���������������� ����, �������� (�2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['����������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['����� � ���������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOWWA': ['������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['�������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWPO': ['�������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGAD': ['����� ��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'WMMI': ['��������������� �����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['��������� �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['��� �����������, ��������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['��� �����������, ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSU': ['������������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HYDV': ['���� �������������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
              }

n_dic_2_70 = {'NODRAH': ['������ �������������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHNGA': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAPL': ['����������� ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHG': ['��������� ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNLP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UNHP': ['���� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWRP': ['���������������� ����, �������� (�2)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'HWSP': ['���������������� ����, ������ (�1)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGS': ['��� �����������, ������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHRGR': ['��� �����������, ��������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWCH': ['��������� (�8)', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHD': ['��������� �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHLGPT': ['��������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UFGAW': ['��������� ����� � ����� ���. �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NOVENA': ['����� � ���������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'NHC3+': ['�3+ �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'USLP': ['��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UAIN': ['������ ���', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWSW': ['����������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UWFF': ['�������� ����', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'UHGAH': ['��������� ��� ��', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
              'OFSP': ['����������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

n_dic_3_20 = {'UHD': ['��������� �������', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]}

info_summary_table_phase2 = [['���������', '���������', '�����', '���������', '�����', '���������� �� ��',
                              '��������� ��', '�������� ��',
                              'RFI ERECTION', 'RFI TEST', 'RFI AIR BLOWING', 'RFI REINSTATEMENT', '������ �����������',
                              '������ �������� �.�.', '��� ���������', '��� ��������', '�������� (��)']]


for key in testpackages.keys():
    status_akt_id = testpackages[key][6]
    status_akt_test = '�� ��������'
    status_akt_blowing = '�� ��������'

    if 'CPECC' in testpackages[key][9]:
        status_akt_id = 'n/p'
    if 'CPECC' in testpackages[key][12] and '���' not in testpackages[key][12]:
        status_akt_test = '��� ��������'
    if 'CPECC' in testpackages[key][12] and '���' not in testpackages[key][11]:
        status_akt_blowing = '��� ��������'
    repairs = 0

    need_ae = '�� ���������'
    done_ae = '�� ����������'
    if testpackages_ae[key] == 1:
        need_ae = '���������'


    info_summary_table_phase2.append(
        [key, testpackages[key][0], testpackages[key][1], testpackages[key][2], testpackages[key][3],
         testpackages[key][6], need_ae, done_ae,
         testpackages[key][9], testpackages[key][10], testpackages[key][12], testpackages[key][11],
         testpackages[key][13],
         status_akt_id, status_akt_test, status_akt_blowing, repairs])

    if testpackages[key][0] == '3-110':
        n_dic_3_110[testpackages[key][1]][1] += testpackages[key][3]
        n_dic_3_110[testpackages[key][1]][2] += 1
        if testpackages[key][9]:
            n_dic_3_110[testpackages[key][1]][3] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][4] += 1
        if testpackages[key][10]:
            n_dic_3_110[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][6] += 1
        if not testpackages[key][10] and testpackages[key][12]:
            n_dic_3_110[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][6] += 1
        if testpackages[key][12]:
            n_dic_3_110[testpackages[key][1]][7] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][8] += 1
        if testpackages[key][11]:
            n_dic_3_110[testpackages[key][1]][9] += testpackages[key][3]
            n_dic_3_110[testpackages[key][1]][10] += 1
        n_dic_3_110[testpackages[key][1]][11] = n_dic_3_110[testpackages[key][1]][1] - \
                                                n_dic_3_110[testpackages[key][1]][9]
        n_dic_3_110[testpackages[key][1]][12] = n_dic_3_110[testpackages[key][1]][2] - \
                                                n_dic_3_110[testpackages[key][1]][10]
        if testpackages_ae[key] == 1:
            n_dic_3_110[testpackages[key][1]][13] += 1

        if 'A' in key[20:23]:
            n_dic_3_110_a[testpackages[key][1]][1] += testpackages[key][3]
            n_dic_3_110_a[testpackages[key][1]][2] += 1
            if testpackages[key][9]:
                n_dic_3_110_a[testpackages[key][1]][3] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][4] += 1
            if testpackages[key][10]:
                n_dic_3_110_a[testpackages[key][1]][5] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][6] += 1
            if not testpackages[key][10] and testpackages[key][12]:
                n_dic_3_110_a[testpackages[key][1]][5] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][6] += 1
            if testpackages[key][12]:
                n_dic_3_110_a[testpackages[key][1]][7] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][8] += 1
            if testpackages[key][11]:
                n_dic_3_110_a[testpackages[key][1]][9] += testpackages[key][3]
                n_dic_3_110_a[testpackages[key][1]][10] += 1
            n_dic_3_110_a[testpackages[key][1]][11] = n_dic_3_110_a[testpackages[key][1]][1] - \
                                                      n_dic_3_110_a[testpackages[key][1]][9]
            n_dic_3_110_a[testpackages[key][1]][12] = n_dic_3_110_a[testpackages[key][1]][2] - \
                                                      n_dic_3_110_a[testpackages[key][1]][10]
            if testpackages_ae[key] == 1:
                n_dic_3_110_a[testpackages[key][1]][13] += 1

        if 'B' in key[20:23]:
            n_dic_3_110_b[testpackages[key][1]][1] += testpackages[key][3]
            n_dic_3_110_b[testpackages[key][1]][2] += 1
            if testpackages[key][9]:
                n_dic_3_110_b[testpackages[key][1]][3] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][4] += 1
            if testpackages[key][10]:
                n_dic_3_110_b[testpackages[key][1]][5] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][6] += 1
            if not testpackages[key][10] and testpackages[key][12]:
                n_dic_3_110_b[testpackages[key][1]][5] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][6] += 1
            if testpackages[key][12]:
                n_dic_3_110_b[testpackages[key][1]][7] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][8] += 1
            if testpackages[key][11]:
                n_dic_3_110_b[testpackages[key][1]][9] += testpackages[key][3]
                n_dic_3_110_b[testpackages[key][1]][10] += 1
            n_dic_3_110_b[testpackages[key][1]][11] = n_dic_3_110_b[testpackages[key][1]][1] - \
                                                      n_dic_3_110_b[testpackages[key][1]][9]
            n_dic_3_110_b[testpackages[key][1]][12] = n_dic_3_110_b[testpackages[key][1]][2] - \
                                                      n_dic_3_110_b[testpackages[key][1]][10]
            if testpackages_ae[key] == 1:
                n_dic_3_110_b[testpackages[key][1]][13] += 1

    if testpackages[key][0] == '3-30':
        n_dic_3_30[testpackages[key][1]][1] += testpackages[key][3]
        n_dic_3_30[testpackages[key][1]][2] += 1
        if testpackages[key][9]:
            n_dic_3_30[testpackages[key][1]][3] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][4] += 1
        if testpackages[key][10]:
            n_dic_3_30[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][6] += 1

        if not testpackages[key][10] and testpackages[key][12]:
            n_dic_3_30[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][6] += 1

        if testpackages[key][12]:
            n_dic_3_30[testpackages[key][1]][7] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][8] += 1
        if testpackages[key][11]:
            n_dic_3_30[testpackages[key][1]][9] += testpackages[key][3]
            n_dic_3_30[testpackages[key][1]][10] += 1
        n_dic_3_30[testpackages[key][1]][11] = n_dic_3_30[testpackages[key][1]][1] - \
                                               n_dic_3_30[testpackages[key][1]][9]
        n_dic_3_30[testpackages[key][1]][12] = n_dic_3_30[testpackages[key][1]][2] - \
                                               n_dic_3_30[testpackages[key][1]][10]
        if testpackages_ae[key] == 1:
            n_dic_3_30[testpackages[key][1]][13] += 1

    if testpackages[key][0] == '2-60':
        n_dic_2_60[testpackages[key][1]][1] += testpackages[key][3]
        n_dic_2_60[testpackages[key][1]][2] += 1
        if testpackages[key][9]:
            n_dic_2_60[testpackages[key][1]][3] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][4] += 1
        if testpackages[key][10]:
            n_dic_2_60[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][6] += 1

        if not testpackages[key][10] and testpackages[key][12]:
            n_dic_2_60[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][6] += 1

        if testpackages[key][12]:
            n_dic_2_60[testpackages[key][1]][7] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][8] += 1
        if testpackages[key][11]:
            n_dic_2_60[testpackages[key][1]][9] += testpackages[key][3]
            n_dic_2_60[testpackages[key][1]][10] += 1
        n_dic_2_60[testpackages[key][1]][11] = n_dic_2_60[testpackages[key][1]][1] - \
                                               n_dic_2_60[testpackages[key][1]][9]
        n_dic_2_60[testpackages[key][1]][12] = n_dic_2_60[testpackages[key][1]][2] - \
                                               n_dic_2_60[testpackages[key][1]][10]
        if testpackages_ae[key] == 1:
            n_dic_2_60[testpackages[key][1]][13] += 1

    if testpackages[key][0] == '2-70':
        n_dic_2_70[testpackages[key][1]][1] += testpackages[key][3]
        n_dic_2_70[testpackages[key][1]][2] += 1
        if testpackages[key][9]:
            n_dic_2_70[testpackages[key][1]][3] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][4] += 1
        if testpackages[key][10]:
            n_dic_2_70[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][6] += 1

        if not testpackages[key][10] and testpackages[key][12]:
            n_dic_2_70[testpackages[key][1]][5] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][6] += 1

        if testpackages[key][12]:
            n_dic_2_70[testpackages[key][1]][7] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][8] += 1
        if testpackages[key][11]:
            n_dic_2_70[testpackages[key][1]][9] += testpackages[key][3]
            n_dic_2_70[testpackages[key][1]][10] += 1
        n_dic_2_70[testpackages[key][1]][11] = n_dic_2_70[testpackages[key][1]][1] - \
                                               n_dic_2_70[testpackages[key][1]][9]
        n_dic_2_70[testpackages[key][1]][12] = n_dic_2_70[testpackages[key][1]][2] - \
                                               n_dic_2_70[testpackages[key][1]][10]
        if testpackages_ae[key] == 1:
            n_dic_2_70[testpackages[key][1]][13] += 1


n_list_3_110_a = [['', f'������ �� �� 3-110 A �� {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                   '', '', '', '', '',
                   '', '', '-', '', ''],
                  ['��� �����', '������������ �����', '�� �������, �.', '���-�� ��', '������ ������, �.',
                   '������ ������, ��',
                   '������� �����-�, �.', '������� �����-�, ��', '������� ��������, �.', '������� ��������, ��',
                   '������� ��, �.',
                   '������� ��, ��', '������� ��, �.', '������� ��, ��', '��������� ��', '�������� ��']]
ITOG_list_a = ['', '�����:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_3_110_b = [['', f'������ �� �� 3-110 B �� {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                   '', '', '', '', '',
                   '', '', '-', '', ''],
                  ['��� �����', '������������ �����', '�� �������, �.', '���-�� ��', '������ ������, �.',
                   '������ ������, ��',
                   '������� �����-�, �.', '������� �����-�, ��', '������� ��������, �.', '������� ��������, ��',
                   '������� ��, �.',
                   '������� ��, ��', '������� ��, �.', '������� ��, ��', '��������� ��', '�������� ��']]
ITOG_list_b = ['', '�����:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_3_110 = [['', f'������ �� �� 3-110 �� {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                 '', '', '', '', '',
                 '', '', '-', '', ''],
                ['��� �����', '������������ �����', '�� �������, �.', '���-�� ��', '������ ������, �.',
                 '������ ������, ��',
                 '������� �����-�, �.', '������� �����-�, ��', '������� ��������, �.', '������� ��������, ��',
                 '������� ��, �.',
                 '������� ��, ��', '������� ��, �.', '������� ��, ��', '��������� ��', '�������� ��']]
ITOG_list = ['', '�����:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_3_30 = [['', f'������ �� �� 3-30 �� {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                '', '', '', '', '',
                '', '', '-', '', ''],
               ['��� �����', '������������ �����', '�� �������, �.', '���-�� ��', '������ ������, �.',
                '������ ������, ��',
                '������� �����-�, �.', '������� �����-�, ��', '������� ��������, �.', '������� ��������, ��',
                '������� ��, �.',
                '������� ��, ��', '������� ��, �.', '������� ��, ��', '��������� ��', '�������� ��']]
ITOG_list_3_30 = ['', '�����:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_2_60 = [['', f'������ �� �� 2-60 �� {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                '', '', '', '', '',
                '', '', '-', '', ''],
               ['��� �����', '������������ �����', '�� �������, �.', '���-�� ��', '������ ������, �.',
                '������ ������, ��',
                '������� �����-�, �.', '������� �����-�, ��', '������� ��������, �.', '������� ��������, ��',
                '������� ��, �.',
                '������� ��, ��', '������� ��, �.', '������� ��, ��', '��������� ��', '�������� ��']]
ITOG_list_2_60 = ['', '�����:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

n_list_2_70 = [['', f'������ �� �� 2-70 �� {datetime.datetime.now().strftime("%d.%m.%Y")}', '', '', '', '',
                '', '', '', '', '',
                '', '', '-', '', ''],
               ['��� �����', '������������ �����', '�� �������, �.', '���-�� ��', '������ ������, �.',
                '������ ������, ��',
                '������� �����-�, �.', '������� �����-�, ��', '������� ��������, �.', '������� ��������, ��',
                '������� ��, �.',
                '������� ��, ��', '������� ��, �.', '������� ��, ��', '��������� ��', '�������� ��']]
ITOG_list_2_70 = ['', '�����:', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

summary_list_units = []

for key in n_dic_3_110.keys():
    n_list_3_110.append([key, n_dic_3_110[key][0], n_dic_3_110[key][1], n_dic_3_110[key][2], n_dic_3_110[key][3],
                         n_dic_3_110[key][4], n_dic_3_110[key][5], n_dic_3_110[key][6], n_dic_3_110[key][7],
                         n_dic_3_110[key][8], n_dic_3_110[key][9], n_dic_3_110[key][10], n_dic_3_110[key][11],
                         n_dic_3_110[key][12], n_dic_3_110[key][13], n_dic_3_110[key][14]])
    ITOG_list[2] += n_dic_3_110[key][1]
    ITOG_list[3] += n_dic_3_110[key][2]
    ITOG_list[4] += n_dic_3_110[key][3]
    ITOG_list[5] += n_dic_3_110[key][4]
    ITOG_list[6] += n_dic_3_110[key][5]
    ITOG_list[7] += n_dic_3_110[key][6]
    ITOG_list[8] += n_dic_3_110[key][7]
    ITOG_list[9] += n_dic_3_110[key][8]
    ITOG_list[10] += n_dic_3_110[key][9]
    ITOG_list[11] += n_dic_3_110[key][10]
    ITOG_list[12] += n_dic_3_110[key][11]
    ITOG_list[13] += n_dic_3_110[key][12]
    ITOG_list[14] += n_dic_3_110[key][13]
    ITOG_list[15] += n_dic_3_110[key][14]
n_list_3_110.append(ITOG_list)
ost_list_3_110 = ['', '�������:', '', '', ITOG_list[2] - ITOG_list[4], ITOG_list[3] - ITOG_list[5],
                  ITOG_list[2] - ITOG_list[6], ITOG_list[3] - ITOG_list[7],
                  ITOG_list[2] - ITOG_list[8], ITOG_list[3] - ITOG_list[9], ITOG_list[2] - ITOG_list[10],
                  ITOG_list[3] - ITOG_list[11], '', '', ITOG_list[14] - ITOG_list[15], '']

empty_str = [' * ', ' * ', ' * ', '  *  ', '  *  ', '  *  ', '  *  ', '  ', '  ', '  ', '  ', '', '', '-', '', '']



for key in n_dic_3_30.keys():
    n_list_3_30.append([key, n_dic_3_30[key][0], n_dic_3_30[key][1], n_dic_3_30[key][2], n_dic_3_30[key][3],
                        n_dic_3_30[key][4], n_dic_3_30[key][5], n_dic_3_30[key][6], n_dic_3_30[key][7],
                        n_dic_3_30[key][8], n_dic_3_30[key][9], n_dic_3_30[key][10], n_dic_3_30[key][11],
                        n_dic_3_30[key][12], n_dic_3_30[key][13], n_dic_3_30[key][14]])
    ITOG_list_3_30[2] += n_dic_3_30[key][1]
    ITOG_list_3_30[3] += n_dic_3_30[key][2]
    ITOG_list_3_30[4] += n_dic_3_30[key][3]
    ITOG_list_3_30[5] += n_dic_3_30[key][4]
    ITOG_list_3_30[6] += n_dic_3_30[key][5]
    ITOG_list_3_30[7] += n_dic_3_30[key][6]
    ITOG_list_3_30[8] += n_dic_3_30[key][7]
    ITOG_list_3_30[9] += n_dic_3_30[key][8]
    ITOG_list_3_30[10] += n_dic_3_30[key][9]
    ITOG_list_3_30[11] += n_dic_3_30[key][10]
    ITOG_list_3_30[12] += n_dic_3_30[key][11]
    ITOG_list_3_30[13] += n_dic_3_30[key][12]
    ITOG_list_3_30[14] += n_dic_3_30[key][13]
    ITOG_list_3_30[15] += n_dic_3_30[key][14]
n_list_3_30.append(ITOG_list_3_30)
ost_list_3_30 = ['', '�������:', '', '', ITOG_list_3_30[2] - ITOG_list_3_30[4], ITOG_list_3_30[3] - ITOG_list_3_30[5],
                 ITOG_list_3_30[2] - ITOG_list_3_30[6], ITOG_list_3_30[3] - ITOG_list_3_30[7],
                 ITOG_list_3_30[2] - ITOG_list_3_30[8], ITOG_list_3_30[3] - ITOG_list_3_30[9],
                 ITOG_list_3_30[2] - ITOG_list_3_30[10],
                 ITOG_list_3_30[3] - ITOG_list_3_30[11], '', '', ITOG_list_3_30[14] - ITOG_list_3_30[15], '']
for i in n_list_3_30:
    summary_list_units.append(i)
summary_list_units.append(ost_list_3_30)
summary_list_units.append(empty_str)

for key in n_dic_2_60.keys():
    n_list_2_60.append([key, n_dic_2_60[key][0], n_dic_2_60[key][1], n_dic_2_60[key][2], n_dic_2_60[key][3],
                        n_dic_2_60[key][4], n_dic_2_60[key][5], n_dic_2_60[key][6], n_dic_2_60[key][7],
                        n_dic_2_60[key][8], n_dic_2_60[key][9], n_dic_2_60[key][10], n_dic_2_60[key][11],
                        n_dic_2_60[key][12], n_dic_2_60[key][13], n_dic_2_60[key][14]])
    ITOG_list_2_60[2] += n_dic_2_60[key][1]
    ITOG_list_2_60[3] += n_dic_2_60[key][2]
    ITOG_list_2_60[4] += n_dic_2_60[key][3]
    ITOG_list_2_60[5] += n_dic_2_60[key][4]
    ITOG_list_2_60[6] += n_dic_2_60[key][5]
    ITOG_list_2_60[7] += n_dic_2_60[key][6]
    ITOG_list_2_60[8] += n_dic_2_60[key][7]
    ITOG_list_2_60[9] += n_dic_2_60[key][8]
    ITOG_list_2_60[10] += n_dic_2_60[key][9]
    ITOG_list_2_60[11] += n_dic_2_60[key][10]
    ITOG_list_2_60[12] += n_dic_2_60[key][11]
    ITOG_list_2_60[13] += n_dic_2_60[key][12]
    ITOG_list_2_60[14] += n_dic_2_60[key][13]
    ITOG_list_2_60[15] += n_dic_2_60[key][14]
n_list_2_60.append(ITOG_list_2_60)
ost_list_2_60 = ['', '�������:', '', '', ITOG_list_2_60[2] - ITOG_list_2_60[4], ITOG_list_2_60[3] - ITOG_list_2_60[5],
                 ITOG_list_2_60[2] - ITOG_list_2_60[6], ITOG_list_2_60[3] - ITOG_list_2_60[7],
                 ITOG_list_2_60[2] - ITOG_list_2_60[8], ITOG_list_2_60[3] - ITOG_list_2_60[9],
                 ITOG_list_2_60[2] - ITOG_list_2_60[10],
                 ITOG_list_2_60[3] - ITOG_list_2_60[11], '', '', ITOG_list_2_60[14] - ITOG_list_2_60[15], '']
for i in n_list_2_60:
    summary_list_units.append(i)

summary_list_units.append(ost_list_2_60)
summary_list_units.append(empty_str)

for key in n_dic_2_70.keys():
    n_list_2_70.append([key, n_dic_2_70[key][0], n_dic_2_70[key][1], n_dic_2_70[key][2], n_dic_2_70[key][3],
                        n_dic_2_70[key][4], n_dic_2_70[key][5], n_dic_2_70[key][6], n_dic_2_70[key][7],
                        n_dic_2_70[key][8], n_dic_2_70[key][9], n_dic_2_70[key][10], n_dic_2_70[key][11],
                        n_dic_2_70[key][12], n_dic_2_70[key][13], n_dic_2_70[key][14]])
    ITOG_list_2_70[2] += n_dic_2_70[key][1]
    ITOG_list_2_70[3] += n_dic_2_70[key][2]
    ITOG_list_2_70[4] += n_dic_2_70[key][3]
    ITOG_list_2_70[5] += n_dic_2_70[key][4]
    ITOG_list_2_70[6] += n_dic_2_70[key][5]
    ITOG_list_2_70[7] += n_dic_2_70[key][6]
    ITOG_list_2_70[8] += n_dic_2_70[key][7]
    ITOG_list_2_70[9] += n_dic_2_70[key][8]
    ITOG_list_2_70[10] += n_dic_2_70[key][9]
    ITOG_list_2_70[11] += n_dic_2_70[key][10]
    ITOG_list_2_70[12] += n_dic_2_70[key][11]
    ITOG_list_2_70[13] += n_dic_2_70[key][12]
    ITOG_list_2_70[14] += n_dic_2_70[key][13]
    ITOG_list_2_70[15] += n_dic_2_70[key][14]
n_list_2_70.append(ITOG_list_2_70)
ost_list_2_70 = ['', '�������:', '', '', ITOG_list_2_70[2] - ITOG_list_2_70[4], ITOG_list_2_70[3] - ITOG_list_2_70[5],
                 ITOG_list_2_70[2] - ITOG_list_2_70[6], ITOG_list_2_70[3] - ITOG_list_2_70[7],
                 ITOG_list_2_70[2] - ITOG_list_2_70[8], ITOG_list_2_70[3] - ITOG_list_2_70[9],
                 ITOG_list_2_70[2] - ITOG_list_2_70[10],
                 ITOG_list_2_70[3] - ITOG_list_2_70[11], '', '', ITOG_list_2_70[14] - ITOG_list_2_70[15], '']
for i in n_list_2_70:
    summary_list_units.append(i)
summary_list_units.append(ost_list_2_70)
summary_list_units.append(empty_str)

for i in n_list_3_110:
    summary_list_units.append(i)
summary_list_units.append(ost_list_3_110)
summary_list_units.append(empty_str)

iso_summary_table = [
    ['���������', '���������', '�����', 'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING', 'RFI REINSTATEMENT',
     '������ �.�. �� ����������', '������ �.�. �� RFI', '��� ���������', '��� ��������', '�����������',
     '������ ���', 'RFI ���', '������ ���', '��� ��������', '��. ����� ��������', 'RFI ����', 'RFI ������']]
for key in isotpdic.keys():
    status_akt_id = 'n/p'
    status_akt_test = '�� ��������'
    status_akt_blowing = '�� ��������'

    if 'CPECC' in isotpdic[key][4] and '���' not in isotpdic[key][4]:
        status_akt_test = '��� ��������'
    if 'CPECC' in isotpdic[key][5] and '���' not in isotpdic[key][5]:
        status_akt_blowing = '��� ��������'
    status_iso_hd = 'n/p'


    status_ncr_iso = '��� �����������'
    if key in iso_ncr_iso.keys():
        status_ncr_iso = iso_ncr_iso[key]

    iso_summary_table.append(
        [key, isotpdic[key][0], isotpdic[key][1], isotpdic[key][2], isotpdic[key][3], isotpdic[key][4],
         isotpdic[key][5], status_iso_hd, status_akt_id, status_akt_test, status_akt_blowing,
         status_ncr_iso, isotpdic[key][10], isotpdic[key][11], isotpdic[key][12], isotpdic[key][6], isotpdic[key][7],
         isotpdic[key][8], isotpdic[key][9]])


# ������ ��� ����� ������������� ���������
double_iso_summary_table = [
    ['���������', '���������', '�����', '���������', 'RFI ERECTION', 'RFI TEST', 'RFI AIRBLOWING',
     'RFI REINSTATEMENT']]
for key in sc_isotpdic.keys():
    double_iso_summary_table.append([sc_isotpdic[key][0], sc_isotpdic[key][1], sc_isotpdic[key][2], sc_isotpdic[key][7],
                                     sc_isotpdic[key][3], sc_isotpdic[key][4], sc_isotpdic[key][5],
                                     sc_isotpdic[key][6]])

# -------------------------------------------


workbook_summary = xlsxwriter.Workbook(f'������ �� ���� 2 �� {datetime.datetime.now().strftime("%d.%m.%Y")}.xlsx')

ws0 = workbook_summary.add_worksheet('������� ������ �� ����������')
ws0.set_column(0, 0, 12)
ws0.set_column(1, 1, 40)
ws0.set_column(4, 15, 12)
ws0.set_column(2, 2, 12)
ws0.set_column(3, 3, 12)

cell_format_green = workbook_summary.add_format()
cell_format_green.set_bg_color('#98FB98')
cell_format_blue = workbook_summary.add_format()
cell_format_blue.set_bg_color('#B0E0E6')
cell_format_hat = workbook_summary.add_format()
cell_format_hat.set_bg_color('#FFDAB9')
cell_format_date = workbook_summary.add_format()
cell_format_date.set_font_size(font_size=14)
for i, (one, two, three, four, five, six, seven, eight, nine, ten, eleven, twelve, thirteen,
        fourteen, fiften, sixten,) in enumerate(summary_list_units, start=2):
    if fourteen == 0:
        color = cell_format_green
    elif fourteen == '������� ��, ��':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
    elif fourteen == '-':
        color = cell_format_date
        color.set_bold('bold')
    elif two == '�����:':
        color = cell_format_hat
        color.set_bold('bold')
    elif two == '�������:':
        color = cell_format_hat
        color.set_bold('bold')
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass
    ws0.write(f'A{i}', one, color)
    ws0.write(f'B{i}', two, color)
    ws0.write(f'C{i}', three, color)
    ws0.write(f'D{i}', four, color)
    ws0.write(f'E{i}', fiften, color)
    ws0.write(f'F{i}', sixten, color)
    ws0.write(f'G{i}', five, color)
    ws0.write(f'H{i}', six, color)
    ws0.write(f'I{i}', seven, color)
    ws0.write(f'J{i}', eight, color)
    ws0.write(f'K{i}', nine, color)
    ws0.write(f'L{i}', ten, color)
    ws0.write(f'M{i}', eleven, color)
    ws0.write(f'N{i}', twelve, color)
    ws0.write(f'O{i}', thirteen, color)
    ws0.write(f'P{i}', fourteen, color)

# ws0.set_column('O:P', None, None, {'hidden': True})



# ------------------------------------------------------------
ws5 = workbook_summary.add_worksheet('������� ���������� �� ��')
ws5.set_column(0, 0, 30)
ws5.set_column(1, 5, 15)
ws5.set_column(6, 11, 22)
ws5.set_column(12, 17, 25)
ws5.set_column(18, 18, 13)
ws5.autofilter('A1:S1682')
for i, (testpack, ustan, flud, metr_ng, stat_id_1, inst_rfi, elev, twelw, thirt,
        fourteen, fifth, akt_id, akt_test, akt_blow, rep, aee, aeeedone) in enumerate(info_summary_table_phase2,
                                                                                      start=1):
    if testpack == '���������':
        color = cell_format_hat
        color.set_bold('bold')
    elif akt_id:
        color = cell_format_green
    else:
        color = cell_format_blue
    try:
        color.set_border(style=1)
        color.set_text_wrap(text_wrap=1)
    except:
        pass

    ws5.write(f'A{i}', testpack, color)
    ws5.write(f'B{i}', ustan, color)
    ws5.write(f'C{i}', flud, color)
    ws5.write(f'D{i}', metr_ng, color)
    ws5.write(f'E{i}', stat_id_1, color)
    # ws5.write(f'F{i}', inst_rfi, color)
    ws5.write(f'F{i}', elev, color)
    ws5.write(f'G{i}', twelw, color)
    ws5.write(f'H{i}', thirt, color)
    ws5.write(f'I{i}', fourteen, color)
    ws5.write(f'J{i}', fifth, color)
    ws5.write(f'K{i}', akt_id, color)
    ws5.write(f'L{i}', akt_test, color)
    # ws5.write(f'N{i}', akt_blow, color)
    ws5.write(f'M{i}', rep, color)
    ws5.write(f'N{i}', aee, color)

# ������ �� ����������

ws5 = workbook_summary.add_worksheet('������ �� ����������')
ws5.set_column(0, 0, 32)
ws5.set_column(1, 1, 28)
ws5.set_column(2, 2, 12)
ws5.set_column(3, 6, 22)
ws5.set_column(7, 9, 20)
ws5.set_column(10, 10, 13.8)
ws5.set_column(11, 11, 20)
ws5.set_column(12, 12, 20)
ws5.set_column(13, 14, 10.5)
ws5.set_column(15, 16, 20)

cell_format_ins = workbook_summary.add_format()
cell_format_ins.set_bg_color('#FFEBCD')

ws5.autofilter('A1:S20000')
for i, (testpack, ustan, flud, metr_ng, stat_id_1, stat_id_2, stat_id_3, inst_rfi, metr_inst, test_rfi, elev,
        twelw, one, two, three, four, five, six, seven) in enumerate(iso_summary_table, start=1):
    if testpack == '���������':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color_2 = cell_format_hat
        color_2.set_bold('bold')
        color_2.set_text_wrap(text_wrap=1)
    elif stat_id_3:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    ws5.write(f'A{i}', testpack, color)
    ws5.write(f'B{i}', ustan, color)
    ws5.write(f'C{i}', flud, color)
    ws5.write(f'D{i}', metr_ng, color)
    ws5.write(f'E{i}', stat_id_1, color)
    ws5.write(f'F{i}', stat_id_2, color)
    ws5.write(f'G{i}', stat_id_3, color)
    # ws5.write(f'H{i}', inst_rfi, color)
    # ws5.write(f'I{i}', metr_inst, color)
    ws5.write(f'H{i}', test_rfi, color)
    ws5.write(f'I{i}', elev, color)
    ws5.write(f'J{i}', twelw, color)
    ws5.write(f'K{i}', one, color)
    ws5.write(f'L{i}', two, color)
    ws5.write(f'M{i}', three, color_2)
    ws5.write(f'N{i}', four, color_2)
    ws5.write(f'O{i}', five, color_2)
    ws5.write(f'P{i}', six, color_2)
    ws5.write(f'Q{i}', seven, color_2)

ws01 = workbook_summary.add_worksheet('Double_iso')
ws01.set_column(0, 0, 37)
ws01.set_column(1, 1, 32)
ws01.set_column(2, 3, 12)
ws01.set_column(4, 7, 22)

ws01.autofilter('A1:S20000')

for i, (one, two, three, four, five, six, seven, eight) in enumerate(double_iso_summary_table, start=1):
    if one == '���������':
        color = cell_format_hat
        color.set_bold('bold')
        color.set_text_wrap(text_wrap=1)
        color_2 = cell_format_hat
        color_2.set_bold('bold')
        color_2.set_text_wrap(text_wrap=1)
    elif eight:
        color = cell_format_green
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    else:
        color = cell_format_blue
        color.set_text_wrap(text_wrap=0)
        color_2 = cell_format_ins
        color_2.set_text_wrap(text_wrap=0)
    try:
        color.set_border(style=1)
        color_2.set_border(style=1)
    except:
        pass

    ws01.write(f'A{i}', one, color)
    ws01.write(f'B{i}', two, color)
    ws01.write(f'C{i}', three, color)
    ws01.write(f'D{i}', four, color)
    ws01.write(f'E{i}', five, color)
    ws01.write(f'F{i}', six, color)
    ws01.write(f'G{i}', seven, color)
    ws01.write(f'H{i}', eight, color)

workbook_summary.close()

print('�������� ����� �������.')
print('������ ���� 2 \n\n ---------')


